from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.hashers import make_password, check_password
from django.db.models import Q, Count
from .models import Usuario, Rol, Padre, Nino, HogarComunitario, Regional, SolicitudMatriculacion, Discapacidad, VisitaTecnica, ActaVisitaTecnica
from django.utils import timezone
from django import forms
from django.contrib.auth.forms import SetPasswordForm
from .forms import AdminPerfilForm, MadrePerfilForm, PadrePerfilForm, NinoForm, PadreForm, CustomAuthForm, AdminForm, NinoSoloForm, BuscarPadreForm, CambiarPadreForm, AgendarVisitaTecnicaForm, ActaVisitaTecnicaForm, HogarFormulario2Form
from desarrollo.models import DesarrolloNino
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Rol, Usuario, MadreComunitaria, HogarComunitario
from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
from django.http import JsonResponse, HttpResponse
from .models import Ciudad
from django.core.paginator import Paginator
from django.template.loader import get_template
import io
from xhtml2pdf import pisa
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from datetime import date, datetime, timedelta
import calendar
from datetime import date
from django.core.mail import send_mail  # 🆕 Para envío de emails
from django.conf import settings  # 🆕 Para configuración de email
from functools import wraps

# Importar decoradores
from .decorators import rol_requerido

# Importar vistas del dashboard mejorado
from .views_dashboard import *
from calendar import monthrange
from django.shortcuts import render
from django.http import JsonResponse
from core.models import Nino
from planeaciones.models import Planeacion
from django.contrib.auth.decorators import login_required
from novedades.models import Novedad
from planeaciones.models import Planeacion
from datetime import datetime as _datetime, date as _date
from core.models import Asistencia
from desarrollo.models import SeguimientoDiario
import uuid
import json
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

# --- VISTAS PERSONALIZADAS DE AUTENTICACIÓN ---
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings

def custom_password_reset(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            # --- VALIDACIÓN PERSONALIZADA ---
            # Verificar si el correo existe en la base de datos
            associated_users = Usuario.objects.filter(Q(correo__iexact=email))
            if associated_users.exists():
                for user in associated_users:
                    # Lógica de envío de correo (la misma que usa Django por defecto)
                    subject = "Restablecimiento de contraseña para ICBF Conecta"
                    email_template_name = "password_reset/password_reset_email.html"
                    c = {
                        "email": user.email,
                        'domain': request.get_host(),
                        'site_name': 'ICBF Conecta',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                    }
                    email_content = render_to_string(email_template_name, c)
                    send_mail(subject, email_content, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)
                return redirect('password_reset_done')
            else:
                # Si el correo no existe, añadir un error al formulario
                form.add_error('email', 'No existe una cuenta asociada a este correo electrónico.')
    else:
        form = PasswordResetForm()
    return render(request, 'password_reset/password_reset_form.html', {'form': form})



# --- GENERAR REPORTE PDF DE MATRÍCULA DE UN NIÑO ---
@login_required
def reporte_matricula_nino_pdf(request, nino_id):
    import os
    from django.conf import settings
    
    nino = get_object_or_404(Nino, id=nino_id)
    padre = nino.padre
    hogar = nino.hogar
    usuario_generador = request.user.get_full_name() or request.user.username
    fecha_reporte = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    # Función auxiliar para obtener ruta absoluta de archivo
    def get_absolute_path(file_field):
        if file_field and hasattr(file_field, 'path'):
            try:
                return os.path.abspath(file_field.path)
            except:
                return None
        return None
    
    template = get_template('madre/reporte_ninos.html')
    context = {
        'nino': nino,
        'padre': padre,
        'hogar': hogar,
        'usuario_generador': usuario_generador,
        'fecha_reporte': fecha_reporte,
        # Rutas absolutas para las imágenes del niño
        'nino_foto_path': get_absolute_path(nino.foto),
        'nino_carnet_path': get_absolute_path(nino.carnet_vacunacion),
        'nino_eps_path': get_absolute_path(nino.certificado_eps),
        'nino_registro_path': get_absolute_path(nino.registro_civil_img),
        # Rutas absolutas para las imágenes del padre
        'padre_documento_path': get_absolute_path(padre.documento_identidad_img) if padre else None,
        'padre_sisben_path': get_absolute_path(padre.clasificacion_sisben) if padre else None,
    }
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_matricula_{nino.nombres}_{nino.apellidos}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def reporte_general_hogar_pdf(request):
    """Genera un reporte PDF con todos los niños del hogar de la madre comunitaria"""
    import os
    from django.conf import settings
    
    # Verificar que el usuario sea madre comunitaria
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Acceso denegado. No tienes los permisos necesarios.')
        return redirect('home')
    
    # Obtener el hogar de la madre comunitaria logueada
    try:
        madre = MadreComunitaria.objects.get(usuario=request.user)
        # El hogar se obtiene a través del related_name 'hogares_asignados'
        hogar = madre.hogares_asignados.first()
        if not hogar:
            messages.error(request, 'No tienes un hogar asignado.')
            return redirect('listar_ninos')
    except MadreComunitaria.DoesNotExist:
        messages.error(request, 'No se encontró información de madre comunitaria.')
        return redirect('listar_ninos')
    
    # Obtener todos los niños del hogar
    ninos = Nino.objects.filter(hogar=hogar).select_related('padre', 'padre__usuario').order_by('apellidos', 'nombres')
    
    if not ninos.exists():
        messages.warning(request, 'No hay niños matriculados en este hogar para generar el reporte.')
        return redirect('listar_ninos')
    
    usuario_generador = request.user.get_full_name() or request.user.username
    fecha_reporte = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    # Función auxiliar para obtener ruta absoluta de archivo
    def get_absolute_path(file_field):
        if file_field and hasattr(file_field, 'path'):
            try:
                return os.path.abspath(file_field.path)
            except:
                return None
        return None
    
    # Preparar datos de cada niño con sus rutas de documentos
    ninos_data = []
    for nino in ninos:
        padre = nino.padre
        nino_info = {
            'nino': nino,
            'padre': padre,
            'nino_foto_path': get_absolute_path(nino.foto),
            'nino_carnet_path': get_absolute_path(nino.carnet_vacunacion),
            'nino_eps_path': get_absolute_path(nino.certificado_eps),
            'nino_registro_path': get_absolute_path(nino.registro_civil_img),
            'padre_documento_path': get_absolute_path(padre.documento_identidad_img) if padre else None,
            'padre_sisben_path': get_absolute_path(padre.clasificacion_sisben) if padre else None,
        }
        ninos_data.append(nino_info)
    
    template = get_template('madre/reporte_general_hogar.html')
    context = {
        'hogar': hogar,
        'ninos_data': ninos_data,
        'total_ninos': ninos.count(),
        'usuario_generador': usuario_generador,
        'fecha_reporte': fecha_reporte,
    }
    
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{hogar.nombre_hogar.replace(" ", "_")}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def certificado_matricula_pdf(request, nino_id):
    """Genera un certificado de matrícula oficial en PDF"""
    import os
    import random
    from datetime import date
    import locale
    
    # Configurar locale para fechas en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
        except:
            pass
    
    nino = get_object_or_404(Nino, id=nino_id)
    padre = nino.padre
    hogar = nino.hogar
    
    # Calcular edad del niño
    hoy = date.today()
    edad = hoy.year - nino.fecha_nacimiento.year - ((hoy.month, hoy.day) < (nino.fecha_nacimiento.month, nino.fecha_nacimiento.day))
    
    # Generar código de verificación único
    codigo_verificacion = f"ICBF-{hogar.id:04d}-{nino.id:05d}-{random.randint(1000, 9999)}"
    
    # Obtener ruta absoluta del logo
    from django.conf import settings
    logo_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'img', 'logo.png')
    if not os.path.exists(logo_path):
        logo_path = None
    else:
        logo_path = os.path.abspath(logo_path)
    
    # Formatear fecha en español manualmente
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    fecha_emision = f"{hoy.day} de {meses[hoy.month - 1]} de {hoy.year}"
    
    # Formatear fechas del niño
    fecha_nac = nino.fecha_nacimiento
    fecha_nacimiento_texto = f"{fecha_nac.day} de {meses[fecha_nac.month - 1]} de {fecha_nac.year}"
    
    fecha_ing = nino.fecha_ingreso
    fecha_ingreso_texto = f"{fecha_ing.day} de {meses[fecha_ing.month - 1]} de {fecha_ing.year}"
    
    template = get_template('madre/certificado_matricula.html')
    context = {
        'nino': nino,
        'padre': padre,
        'hogar': hogar,
        'edad': edad,
        'codigo_verificacion': codigo_verificacion,
        'fecha_emision': fecha_emision,
        'fecha_nacimiento_texto': fecha_nacimiento_texto,
        'fecha_ingreso_texto': fecha_ingreso_texto,
        'año_actual': hoy.year,
        'logo_path': logo_path,
    }
    
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificado_matricula_{nino.nombres}_{nino.apellidos}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el certificado', status=500)
    return response

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Rol, Usuario, MadreComunitaria, HogarComunitario
from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
from django.http import JsonResponse
from .models import Ciudad
from django.core.paginator import Paginator

# --- Dependencias para PDF ---
from django.http import HttpResponse
from django.template.loader import get_template
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xhtml2pdf import pisa
# Asegúrate de importar todos los formularios y modelos necesarios

# ----------------------------------------------------
# 💡 DECORADOR: Restringir acceso por Rol
# ----------------------------------------------------
from functools import wraps

@login_required
def buscar_padre_por_documento(request):
    """
    Vista AJAX para buscar un padre por su documento.
    """
    documento = request.GET.get('documento')
    data = {'encontrado': False}

    if documento:
        usuario_padre = Usuario.objects.filter(documento=documento, rol__nombre_rol='padre').first()
        if usuario_padre:
            padre_profile = Padre.objects.filter(usuario=usuario_padre).first()
            data.update({
                'encontrado': True,
                'nombres': usuario_padre.nombres,
                'apellidos': usuario_padre.apellidos,
                'correo': usuario_padre.correo,
                'telefono': usuario_padre.telefono,
                'direccion': usuario_padre.direccion,
                'ocupacion': padre_profile.ocupacion if padre_profile else ''
            })
    return JsonResponse(data)

# -----------------------------------------------------------------
# 💡 NUEVA FUNCIÓN: Matricular Niño (CRUD Crear)
# -----------------------------------------------------------------

@login_required
def matricular_nino(request):
    # Solo madres comunitarias pueden acceder
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden matricular niños.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        padre_form = PadreForm(request.POST, request.FILES, prefix='padre')
        nino_form = NinoForm(request.POST, request.FILES, prefix='nino')
        if padre_form.is_valid() and nino_form.is_valid():
            try:
                with transaction.atomic():
                    rol_padre = Rol.objects.get(nombre_rol='padre')
                    doc_padre = padre_form.cleaned_data['documento']
                    tipo_doc = padre_form.cleaned_data['tipo_documento']
                    
                    # Buscar o crear el usuario del padre
                    usuario_padre, created = Usuario.objects.get_or_create(
                        documento=doc_padre,
                        rol=rol_padre,
                        defaults={
                            'tipo_documento': tipo_doc,
                            'nombres': padre_form.cleaned_data['nombres'],
                            'apellidos': padre_form.cleaned_data['apellidos'],
                            'correo': padre_form.cleaned_data['correo'],
                            'telefono': padre_form.cleaned_data['telefono'],
                            'direccion': padre_form.cleaned_data.get('direccion', ''),
                        }
                    )

                    if created:
                        usuario_padre.set_password(str(doc_padre))
                        usuario_padre.save()
                    else: # Si ya existía, actualizamos sus datos por si hay cambios
                        usuario_padre.tipo_documento = tipo_doc
                        usuario_padre.nombres = padre_form.cleaned_data['nombres']
                        usuario_padre.apellidos = padre_form.cleaned_data['apellidos']
                        usuario_padre.correo = padre_form.cleaned_data['correo']
                        usuario_padre.telefono = padre_form.cleaned_data['telefono']
                        usuario_padre.direccion = padre_form.cleaned_data.get('direccion', '')
                        usuario_padre.save()

                    # Buscar o crear el perfil del padre
                    padre_obj, _ = Padre.objects.get_or_create(usuario=usuario_padre)
                    padre_obj.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                    padre_obj.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                    padre_obj.estrato = padre_form.cleaned_data.get('estrato')
                    padre_obj.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                    padre_obj.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                    padre_obj.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                    
                    # Guardar archivos del padre si se proporcionan
                    if 'documento_identidad_img' in request.FILES:
                        padre_obj.documento_identidad_img = request.FILES['documento_identidad_img']
                    if 'clasificacion_sisben' in request.FILES:
                        padre_obj.clasificacion_sisben = request.FILES['clasificacion_sisben']
                    
                    padre_obj.save()

                    # Crear el niño y asociarlo
                    nino = nino_form.save(commit=False)
                    nino.hogar = hogar_madre
                    nino.padre = padre_obj
                    # Nuevos campos
                    nino.tipo_sangre = nino_form.cleaned_data.get('tipo_sangre')
                    nino.parentesco = nino_form.cleaned_data.get('parentesco')
                    nino.tiene_discapacidad = nino_form.cleaned_data.get('tiene_discapacidad', False)
                    nino.otra_discapacidad = nino_form.cleaned_data.get('otra_discapacidad', '')
                    nino.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                    nino.save()
                    # ManyToMany: tipos_discapacidad
                    if nino.tiene_discapacidad:
                        nino.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                    else:
                        nino.tipos_discapacidad.clear()
                    # Guardar archivos del niño (fotos y documentos)
                    if 'nino-foto' in request.FILES:
                        nino.foto = request.FILES['nino-foto']
                    if 'nino-carnet_vacunacion' in request.FILES:
                        nino.carnet_vacunacion = request.FILES['nino-carnet_vacunacion']
                    if 'nino-certificado_eps' in request.FILES:
                        nino.certificado_eps = request.FILES['nino-certificado_eps']
                    if 'nino-registro_civil_img' in request.FILES:
                        nino.registro_civil_img = request.FILES['nino-registro_civil_img']
                    nino.save()
                    
                    # Guardar el nombre del niño en la sesión para mostrarlo en el SweetAlert
                    request.session['matricula_exitosa'] = {
                        'nombre': f'{nino.nombres} {nino.apellidos}',
                        'mensaje': f'El niño {nino.nombres} {nino.apellidos} ha sido matriculado exitosamente en el hogar {hogar_madre.nombre_hogar}.'
                    }
                    messages.success(request, f'Niño {nino.nombres} matriculado correctamente.')
                    return redirect('listar_ninos')
            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        padre_form = PadreForm(prefix='padre')
        nino_form = NinoForm(prefix='nino')

    # TODO: CAMBIAR TEMPLATE - Para usar el nuevo template mejorado, cambia 'madre/nino_form.html' 
    # por 'madre/nino_form_nuevo.html' (que tiene mejor organización de campos)
    return render(request, 'madre/nino_form_nuevo.html', {
        'hogar_madre': hogar_madre,
        'nino_form': nino_form,
        'padre_form': padre_form,
        'modo_edicion': False
    })
def home(request):
    return render(request, 'home.html')
@login_required
def admin_dashboard(request):
    """
    Dashboard mejorado - redirige a la nueva versión completa
    Para mantener compatibilidad, puedes elegir cuál template usar
    """
    # Opción 1: Usar el nuevo dashboard completo
    return dashboard_admin(request)
    
    # Opción 2: Mantener el dashboard antiguo (comentar la línea anterior y descomentar esta)
    # return render(request, 'admin/dashboard.html')

@login_required
@rol_requerido('administrador')
def admin_reportes(request):
    """Renderiza la página de reportes para el administrador."""
    context = {
        'regionales_filtro': Regional.objects.all().order_by('nombre'),
        'escolaridad_choices': MadreComunitaria.NIVEL_ESCOLARIDAD_CHOICES
    }
    return render(request, 'admin/reportes.html', context)

# ---------- Formularios simples ----------
class MadreForm(forms.ModelForm):
    class Meta:
        model = Usuario
        fields = ['nombres', 'apellidos', 'documento', 'email', 'telefono', 'direccion']


class HogarForm(forms.ModelForm):
    class Meta:
        model = HogarComunitario
        fields = ['nombre_hogar', 'direccion', 'localidad', 'estado', 'regional', 'ciudad']
# En la sección de formularios simples en views.py
class AdministradorForm(forms.ModelForm):
    # Añadir el campo de contraseña, ya que no se incluye automáticamente
    contraseña = forms.CharField(widget=forms.PasswordInput, required=True)
    
    class Meta:
        model = Usuario
        # Elige los campos que quieres:
        fields = ['first_name', 'email', 'rol'] # Asumiendo estos nombres para Usuario

# ---------- CRUD MADRES ----------
@login_required
# En core/views.py

@rol_requerido('administrador')
def listar_madres(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre', '')
    query_documento = request.GET.get('documento', '')
    query_hogar = request.GET.get('hogar', '')

    # Obtener todos los perfiles de madre con sus datos relacionados
    madres_query = MadreComunitaria.objects.select_related('usuario').prefetch_related('hogares_asignados').all()

    if query_nombre:
        madres_query = madres_query.filter(Q(usuario__nombres__icontains=query_nombre) | Q(usuario__apellidos__icontains=query_nombre))
    if query_documento:
        madres_query = madres_query.filter(usuario__documento__icontains=query_documento)
    if query_hogar:
        madres_query = madres_query.filter(hogares_asignados__nombre_hogar__icontains=query_hogar)
    
    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(madres_query.distinct(), 5) # 5 madres por página
    page_number = request.GET.get('page')
    madres_paginadas = paginator.get_page(page_number)

    context = {
        'madres': madres_paginadas, # Enviar el objeto paginado a la plantilla
        'filtros': { # Devolver los filtros a la plantilla
            'nombre': query_nombre,
            'documento': query_documento,
            'hogar': query_hogar,
        },
        'paginator': paginator # Opcional, pero útil para la plantilla
    }
    return render(request, 'admin/madres_list.html', context)

@login_required
@rol_requerido('administrador')
def listar_hogares(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre_hogar', '')
    query_madre = request.GET.get('madre', '')
    query_regional = request.GET.get('regional', '')

    # Consulta base
    hogares = HogarComunitario.objects.select_related(
        'madre__usuario', 'regional'
    ).annotate(
        num_ninos=Count('ninos')
    ).order_by('regional__nombre', 'nombre_hogar')

    if query_nombre:
        hogares = hogares.filter(nombre_hogar__icontains=query_nombre)
    if query_madre:
        hogares = hogares.filter(Q(madre__usuario__nombres__icontains=query_madre) | Q(madre__usuario__apellidos__icontains=query_madre))
    if query_regional:
        hogares = hogares.filter(regional_id=query_regional)

    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(hogares, 5) # 5 hogares por página
    page_number = request.GET.get('page')
    hogares_paginados = paginator.get_page(page_number)

    context = {
        'hogares': hogares_paginados, # Enviar el objeto paginado
        'regionales_filtro': Regional.objects.all().order_by('nombre'), # Para el dropdown de filtros
        'filtros': {
            'nombre_hogar': query_nombre,
            'madre': query_madre,
            'regional': query_regional,
        },
        'paginator': paginator # Opcional
    }

    return render(request, 'admin/hogares_list.html', context)

# Importa los formularios correctos al inicio de tu views.py
# from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
# ...

@login_required
# En core/views.py (con la lógica de importaciones y Rol.objects.get_or_create)


@login_required
@rol_requerido('administrador')
def crear_madre(request):
    rol_madre, _ = Rol.objects.get_or_create(nombre_rol='madre_comunitaria')

    regionales = Regional.objects.all().order_by('nombre')
    if request.method == 'POST':
        # DEBUG: Imprimir datos recibidos
        print("=" * 80)
        print("DEBUG - Datos POST recibidos:")
        print(f"Regional: {request.POST.get('regional')}")
        print(f"Ciudad: {request.POST.get('ciudad')}")
        print(f"Estado: {request.POST.get('estado')}")
        print(f"Capacidad máxima: {request.POST.get('capacidad_maxima')}")
        print(f"Nombre hogar: {request.POST.get('nombre_hogar')}")
        print(f"Dirección: {request.POST.get('direccion')}")
        print("=" * 80)
        
        usuario_form = UsuarioMadreForm(request.POST, request.FILES)
        madre_profile_form = MadreProfileForm(request.POST, request.FILES)
        hogar_form = HogarForm(request.POST)  # ¡Esta es la línea que faltaba!
        error_step = 1

        # DEBUG: Imprimir errores de cada formulario
        print("=" * 80)
        print("DEBUG - Validación de formularios:")
        print(f"UsuarioForm válido: {usuario_form.is_valid()}")
        if not usuario_form.is_valid():
            print(f"  Errores: {usuario_form.errors}")
        print(f"MadreProfileForm válido: {madre_profile_form.is_valid()}")
        if not madre_profile_form.is_valid():
            print(f"  Errores: {madre_profile_form.errors}")
        print(f"HogarForm válido: {hogar_form.is_valid()}")
        if not hogar_form.is_valid():
            print(f"  Errores: {hogar_form.errors}")
        print("=" * 80)

        if usuario_form.is_valid() and madre_profile_form.is_valid() and hogar_form.is_valid():
            documento = usuario_form.cleaned_data.get('documento')
            # 💡 CORRECCIÓN: Priorizar el nombre del hogar del formulario. Si está vacío, generar uno.
            nombre_hogar = hogar_form.cleaned_data.get('nombre_hogar')
            if not nombre_hogar:
                nombre_hogar = "Hogar de " + usuario_form.cleaned_data.get('nombres', '').split(' ')[0]

            direccion_hogar = hogar_form.cleaned_data.get('direccion')
            localidad_bogota = hogar_form.cleaned_data.get('localidad_bogota')
            ciudad = hogar_form.cleaned_data.get('ciudad')

            # Validación de documento duplicado
            if Usuario.objects.filter(documento=documento).exists():
                messages.error(request, f"Ya existe un usuario con el documento {documento} registrado.")
                from .models import LocalidadBogota
                localidades_bogota = LocalidadBogota.objects.all().order_by('numero')
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form, 'madre_profile_form': madre_profile_form, 'hogar_form': hogar_form,
                    'initial_step': 1, 'regionales': regionales, 'localidades_bogota': localidades_bogota
                })

            # Validación de hogar duplicado por dirección
            if HogarComunitario.objects.filter(direccion__iexact=direccion_hogar).exists():
                messages.error(request, f"La dirección '{direccion_hogar}' ya está registrada para otro hogar.")
                error_step = 3

            if messages.get_messages(request):
                from .models import LocalidadBogota
                localidades_bogota = LocalidadBogota.objects.all().order_by('numero')
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form, 'madre_profile_form': madre_profile_form, 'hogar_form': hogar_form,
                    'initial_step': error_step, 'regionales': regionales, 'localidades_bogota': localidades_bogota
                })

            try:
                # Usar transacción atómica para asegurar que todo se crea o nada
                with transaction.atomic():
                    # 1️⃣ Crear usuario
                    usuario = usuario_form.save(commit=False)
                    usuario.rol = rol_madre
                    # La contraseña se establece aquí, puedes cambiarla si es necesario
                    usuario.set_password('123456')
                    usuario.is_active = True
                    usuario.save()

                    # 2️⃣ Crear perfil madre
                    madre_profile = madre_profile_form.save(commit=False)
                    madre_profile.usuario = usuario
                    madre_profile.save()

                    # 3️⃣ Crear hogar comunitario asociado a la madre
                    hogar = hogar_form.save(commit=False)
                    hogar.madre = madre_profile
                    hogar.nombre_hogar = nombre_hogar
                    # Mantener compatibilidad: guardar nombre de localidad en campo texto
                    if localidad_bogota:
                        hogar.localidad = localidad_bogota.nombre
                    else:
                        hogar.localidad = ciudad.nombre if ciudad else ''
                    # 🆕 El hogar inicia en estado "pendiente_visita" hasta que se realice la visita técnica
                    hogar.estado = 'pendiente_visita'
                    
                    # Guardar fecha de primera visita si está disponible
                    fecha_primera_visita = request.POST.get('fecha_primera_visita')
                    if fecha_primera_visita:
                        from datetime import datetime, date
                        fecha_visita_obj = datetime.strptime(fecha_primera_visita, '%Y-%m-%d').date()
                        
                        # ✅ Validar que no sea en el pasado
                        if fecha_visita_obj < date.today():
                            messages.error(request, '❌ La fecha de primera visita no puede ser en el pasado.')
                            from .models import LocalidadBogota
                            localidades_bogota = LocalidadBogota.objects.all().order_by('numero')
                            return render(request, 'admin/madres_form.html', {
                                'usuario_form': usuario_form, 
                                'madre_profile_form': madre_profile_form, 
                                'hogar_form': hogar_form,
                                'initial_step': 3, 
                                'regionales': regionales,
                                'localidades_bogota': localidades_bogota
                            })
                        
                        hogar.fecha_primera_visita = fecha_visita_obj
                    
                    hogar.save()

                    # 4️⃣ Crear visita técnica programada
                    if fecha_primera_visita:
                        from .models import VisitaTecnica
                        from datetime import datetime
                        fecha_visita_dt = datetime.strptime(fecha_primera_visita, '%Y-%m-%d')
                        
                        visita = VisitaTecnica.objects.create(
                            hogar=hogar,
                            fecha_programada=fecha_visita_dt,
                            tipo_visita='V1',
                            estado='agendada',
                            creado_por=request.user,
                            observaciones_agenda=f'Primera visita programada al crear el hogar {nombre_hogar}'
                        )
                        
                        # Enviar correo de notificación
                        try:
                            from django.core.mail import send_mail
                            from django.template.loader import render_to_string
                            from django.utils.html import strip_tags
                            
                            asunto = f'Visita Técnica Programada - {nombre_hogar}'
                            mensaje_html = f"""
                            <html>
                            <body style="font-family: Arial, sans-serif;">
                                <h2 style="color: #004080;">Visita Técnica Programada</h2>
                                <p>Estimado/a <strong>{usuario.nombres} {usuario.apellidos}</strong>,</p>
                                <p>Se ha programado la primera visita técnica para el hogar comunitario:</p>
                                <div style="background-color: #f0f8ff; padding: 15px; border-left: 4px solid #007bff; margin: 20px 0;">
                                    <p><strong>Hogar:</strong> {nombre_hogar}</p>
                                    <p><strong>Dirección:</strong> {direccion_hogar}</p>
                                    <p><strong>Fecha de Visita:</strong> {fecha_visita_dt.strftime('%d de %B de %Y')}</p>
                                </div>
                                <p>Durante esta visita se evaluarán las condiciones del hogar para determinar su capacidad y habilitación.</p>
                                <p><strong>Importante:</strong> Por favor asegúrese de estar presente en la fecha programada.</p>
                                <br>
                                <p>Atentamente,</p>
                                <p><strong>Sistema ICBF Conecta</strong></p>
                            </body>
                            </html>
                            """
                            mensaje_texto = strip_tags(mensaje_html)
                            
                            send_mail(
                                asunto,
                                mensaje_texto,
                                'noreply@icbf.gov.co',
                                [usuario.correo],
                                html_message=mensaje_html,
                                fail_silently=True
                            )
                            
                            visita.correo_enviado = True
                            from django.utils import timezone
                            visita.fecha_envio_correo = timezone.now()
                            visita.save()
                            
                        except Exception as e:
                            print(f"Error al enviar correo: {e}")

                    # 5️⃣ Crear convivientes del hogar
                    num_convivientes = request.POST.get('num_convivientes', 0)
                    try:
                        num_convivientes = int(num_convivientes)
                    except (ValueError, TypeError):
                        num_convivientes = 0
                    
                    for i in range(num_convivientes):
                        tipo_doc = request.POST.get(f'conviviente_{i}_tipo_documento')
                        numero_doc = request.POST.get(f'conviviente_{i}_numero_documento')
                        nombre = request.POST.get(f'conviviente_{i}_nombre_completo')
                        parentesco = request.POST.get(f'conviviente_{i}_parentesco')
                        antecedentes = request.FILES.get(f'conviviente_{i}_antecedentes')
                        
                        if tipo_doc and numero_doc and nombre and parentesco:
                            from .models import ConvivienteHogar
                            ConvivienteHogar.objects.create(
                                hogar=hogar,
                                tipo_documento=tipo_doc,
                                numero_documento=numero_doc,
                                nombre_completo=nombre,
                                parentesco=parentesco,
                                antecedentes_pdf=antecedentes
                            )

                # Mensaje de éxito informativo
                mensaje_exito = '✅ ¡Agente educativo y hogar creados exitosamente! '
                if fecha_primera_visita:
                    mensaje_exito += f'📅 Visita técnica programada para el {fecha_visita_dt.strftime("%d de %B de %Y")}. '
                    mensaje_exito += f'📧 Se ha enviado un correo de confirmación. '
                mensaje_exito += '⚠️ El hogar permanecerá en estado "Pendiente de Visita" hasta completar la evaluación técnica. '
                mensaje_exito += '🔑 Contraseña temporal: 123456'
                
                messages.success(request, mensaje_exito)
                return redirect('listar_madres')

            except Exception as e:
                messages.error(request, f"Ocurrió un error al guardar los datos: {str(e)}")

        else:
            if usuario_form.errors:
                error_step = 1
            elif madre_profile_form.errors:
                error_step = 2
            elif hogar_form.errors:
                error_step = 3
            messages.error(request, 'Error en los datos suministrados. Revise el paso marcado en azul.')

        from .models import LocalidadBogota
        localidades_bogota = LocalidadBogota.objects.all().order_by('numero')
        return render(request, 'admin/madres_form.html', {
            'usuario_form': usuario_form,
            'madre_profile_form': madre_profile_form,
            'hogar_form': hogar_form,
            'initial_step': error_step, # Para saber en qué paso del formulario mostrar el error
            'regionales': regionales,
            'localidades_bogota': localidades_bogota
        })

    # GET
    # 💡 MEJORA: Si el admin tiene una regional, la pre-seleccionamos en el formulario del hogar.
    initial_hogar = {} # Revertido a estado original
    
    # 🔧 Cargar explícitamente las localidades de Bogotá para el template
    from .models import LocalidadBogota
    localidades_bogota = LocalidadBogota.objects.all().order_by('numero')
    
    return render(request, 'admin/madres_form.html', {
        'usuario_form': UsuarioMadreForm(),
        'madre_profile_form': MadreProfileForm(),
        # Pasamos los datos iniciales al formulario del hogar
        'hogar_form': HogarForm(),
        'regionales': regionales,
        'localidades_bogota': localidades_bogota,  # 🆕 Pasar localidades explícitamente
        'initial_step': 1
    })
@login_required
@rol_requerido('administrador')
def editar_madre(request, id):
    # Obtener el usuario que es madre comunitaria
    usuario_madre = get_object_or_404(Usuario, id=id, rol__nombre_rol='madre_comunitaria')
    
    # Obtener el perfil de madre comunitaria
    madre_profile = usuario_madre.madre_profile
    
    # Obtener el hogar comunitario asociado
    hogar = HogarComunitario.objects.filter(madre=madre_profile).first()
    
    if not hogar:
        # Si por alguna razón no tiene hogar, es mejor redirigir.
        messages.error(request, 'Esta madre no tiene un hogar comunitario asignado.')
        return redirect('listar_madres')

    if request.method == 'POST':
        # PASAR request.FILES para permitir subida de nueva foto
        usuario_form = UsuarioMadreForm(request.POST, instance=usuario_madre)
        madre_profile_form = MadreProfileForm(request.POST, request.FILES, instance=madre_profile)
        hogar_form = HogarForm(request.POST, instance=hogar)
        
        if usuario_form.is_valid() and madre_profile_form.is_valid() and hogar_form.is_valid():
            # Validación para hogar duplicado, excluyendo el hogar actual
            nombre_hogar = hogar_form.cleaned_data['nombre_hogar']
            direccion_hogar = hogar_form.cleaned_data['direccion']
            localidad = hogar_form.cleaned_data.get('localidad')

            if HogarComunitario.objects.filter(nombre_hogar__iexact=nombre_hogar, localidad__iexact=localidad).exclude(id=hogar.id).exists():
                messages.error(request, f"Ya existe otro hogar llamado '{nombre_hogar}' en la localidad de '{localidad}'.")
            elif HogarComunitario.objects.filter(direccion__iexact=direccion_hogar).exclude(id=hogar.id).exists():
                messages.error(request, f"La dirección '{direccion_hogar}' ya está registrada para otro hogar.")

            # Si hay mensajes de error, renderizar de nuevo el formulario
            if messages.get_messages(request):
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form,
                    'madre_profile_form': madre_profile_form,
                    'hogar_form': hogar_form,
                    'modo_edicion': True,
                    'regionales': Regional.objects.all().order_by('nombre')
                })

            try:
                with transaction.atomic():
                    # Guardar usuario
                    usuario_actualizado = usuario_form.save()
                    
                    # Guardar perfil y hogar
                    madre_profile_form.save()
                    hogar_form.save()
                
                messages.success(request, '¡Madre comunitaria y hogar actualizados exitosamente!')
                return redirect('listar_madres') # Corregido para que el redirect esté dentro del try
            except Exception as e:
                messages.error(request, f'Error al guardar los cambios: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        usuario_form = UsuarioMadreForm(instance=usuario_madre)
        madre_profile_form = MadreProfileForm(instance=madre_profile)
        hogar_form = HogarForm(instance=hogar)

    regionales = Regional.objects.all().order_by('nombre')
    return render(request, 'admin/madres_form.html', {
        'usuario_form': usuario_form,
        'madre_profile_form': madre_profile_form,
        'hogar_form': hogar_form,
        'modo_edicion': True,
        'regionales': regionales
    })


@login_required
def eliminar_madre(request, id):
    # Obtener el usuario que es madre comunitaria
    usuario_madre = get_object_or_404(Usuario, id=id, rol__nombre_rol='madre_comunitaria')
    
    try:
        # Usar una transacción para asegurar la integridad de los datos
        with transaction.atomic():
            # El perfil de MadreComunitaria se eliminará en cascada cuando se elimine el Usuario,
            # pero el HogarComunitario está protegido.
            # Primero, eliminamos los hogares asociados a su perfil.
            if hasattr(usuario_madre, 'madre_profile'):
                HogarComunitario.objects.filter(madre=usuario_madre.madre_profile).delete()
            # Ahora sí podemos eliminar el usuario (y su perfil de madre en cascada)
            usuario_madre.delete()
        messages.success(request, '¡Madre comunitaria y su hogar asociado han sido eliminados exitosamente!')
    except Exception as e:
        messages.error(request, f"Ocurrió un error al intentar eliminar a la madre: {e}")
        
    return redirect('listar_madres')

@login_required
@rol_requerido('administrador')
def listar_administradores(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre', '')
    query_documento = request.GET.get('documento', '')

    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    # Incluimos la regional en la consulta para optimizar
    administradores = Usuario.objects.filter(rol=rol_admin).order_by('nombres')

    if query_nombre:
        administradores = administradores.filter(Q(nombres__icontains=query_nombre) | Q(apellidos__icontains=query_nombre))
    if query_documento:
        administradores = administradores.filter(documento__icontains=query_documento)

    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(administradores, 5) # 5 administradores por página
    page_number = request.GET.get('page')
    admins_paginados = paginator.get_page(page_number)

    context = {
        'administradores': admins_paginados, # Enviar el objeto paginado
        'filtros': {'nombre': query_nombre, 'documento': query_documento},
        'paginator': paginator # Opcional
    }
    return render(request, 'admin/administradores_list.html', context)


@login_required
@rol_requerido('administrador')
def crear_administrador(request):
    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    
    if request.method == 'POST':
        form = AdminForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            correo = form.cleaned_data['correo']
            contraseña = form.cleaned_data['contraseña']

            if not contraseña:
                messages.error(request, 'El campo contraseña es obligatorio para crear un nuevo administrador.')
                return render(request, 'admin/administradores_form.html', {'form': form})

            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exists():
                messages.error(request, 'Ya existe un usuario con ese documento o correo electrónico.')
                return render(request, 'admin/administradores_form.html', {'form': form})

            usuario = form.save(commit=False)
            usuario.rol = rol_admin
            usuario.set_password(contraseña)
            usuario.save()

            messages.success(request, '¡Administrador creado exitosamente!')
            return redirect('listar_administradores')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AdminForm()

    return render(request, 'admin/administradores_form.html', {'form': form})

@login_required
@rol_requerido('administrador')
def editar_administrador(request, id):
    admin = get_object_or_404(Usuario, id=id, rol__nombre_rol='administrador')
    if request.method == 'POST':
        form = AdminForm(request.POST, request.FILES, instance=admin)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            correo = form.cleaned_data['correo']

            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exclude(id=id).exists():
                messages.error(request, 'Ya existe otro usuario con ese documento o correo electrónico.')
                return render(request, 'admin/administradores_form.html', {'form': form, 'admin': admin})

            usuario = form.save(commit=False)
            nueva_contraseña = form.cleaned_data.get('contraseña')
            if nueva_contraseña:
                usuario.set_password(nueva_contraseña)
                if request.user.id == admin.id:
                    update_session_auth_hash(request, usuario)
            usuario.save()
            messages.success(request, '¡Administrador actualizado exitosamente!')
            return redirect('listar_administradores')
    else:
        form = AdminForm(instance=admin)

    return render(request, 'admin/administradores_form.html', {'form': form, 'admin': admin})

def _setup_excel_report_header(ws, title, record_count, num_columns):
    """
    Función auxiliar para configurar el encabezado personalizado en los reportes de Excel.
    """
    # --- ESTILOS ---
    title_font = Font(name='Poppins', bold=True, size=16)
    info_bar_font = Font(name='Poppins', bold=True, color='FFFFFF')
    info_bar_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    info_label_font = Font(name='Poppins', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- TÍTULO PRINCIPAL ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30

    # --- BARRA DE INFORMACIÓN ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)
    info_bar_cell = ws.cell(row=2, column=1, value="Información del reporte")
    info_bar_cell.font = info_bar_font
    info_bar_cell.fill = info_bar_fill
    info_bar_cell.alignment = center_alignment
    ws.row_dimensions[2].height = 25

    # --- Lógica para dividir las columnas de la sección informativa ---
    # Si hay 6 columnas, el punto medio será 3. El label irá de 1 a 3, y el valor de 4 a 6.
    mid_point_col = num_columns // 2

    # --- SECCIÓN INFORMATIVA ---
    # Fecha de Generación
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=mid_point_col)
    fecha_label = ws.cell(row=3, column=1, value="Fecha de Generación:")
    fecha_label.font = info_label_font

    ws.merge_cells(start_row=3, start_column=mid_point_col + 1, end_row=3, end_column=num_columns)
    fecha_value = ws.cell(row=3, column=mid_point_col + 1, value=timezone.now().strftime('%Y-%m-%d'))
    fecha_value.alignment = left_alignment

    # Aplicar borde a toda la fila 3
    for col in range(1, num_columns + 1):
        ws.cell(row=3, column=col).border = thin_border

    # Total de Registros
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=mid_point_col)
    total_label = ws.cell(row=4, column=1, value="Total de Registros:")
    total_label.font = info_label_font

    ws.merge_cells(start_row=4, start_column=mid_point_col + 1, end_row=4, end_column=num_columns)
    total_value = ws.cell(row=4, column=mid_point_col + 1, value=record_count)
    total_value.alignment = left_alignment

    # Aplicar borde a toda la fila 4
    for col in range(1, num_columns + 1):
        ws.cell(row=4, column=col).border = thin_border

    # Ajustar ancho de la primera columna
    ws.column_dimensions['A'].width = 25

    # Devolver la fila donde deben empezar los datos de la tabla
    return 5 # Los datos comenzarán en la fila 5

@login_required
@rol_requerido('administrador')
def reporte_administradores_excel(request):
    # Obtener filtros de la URL
    nombre = request.GET.get('nombre', '')
    documento = request.GET.get('documento', '')

    # Filtrar administradores
    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    administradores = Usuario.objects.filter(rol=rol_admin).order_by('nombres')
    if nombre:
        administradores = administradores.filter(Q(nombres__icontains=nombre) | Q(apellidos__icontains=nombre))
    if documento:
        administradores = administradores.filter(documento__icontains=documento)

    # --- Encabezados y configuración inicial ---
    headers = ['Nombres', 'Apellidos', 'Tipo Documento', 'Documento', 'Correo', 'Teléfono']
    num_columns = len(headers)

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Administradores'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Administradores", administradores.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25

    # --- Llenar datos ---
    for admin in administradores:
        start_row += 1
        ws.cell(row=start_row, column=1, value=admin.nombres)
        ws.cell(row=start_row, column=2, value=admin.apellidos)
        ws.cell(row=start_row, column=3, value=admin.get_tipo_documento_display())
        ws.cell(row=start_row, column=4, value=admin.documento)
        ws.cell(row=start_row, column=5, value=admin.correo)
        ws.cell(row=start_row, column=6, value=admin.telefono)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename="reporte_administradores.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('administrador')
def reporte_madres_excel(request):
    # Obtener filtros
    nombre = request.GET.get('nombre')
    hogar_asignado = request.GET.get('hogar') # Corregido para coincidir con el filtro de la lista
    escolaridad = request.GET.get('escolaridad', None)

    # Filtrar madres
    madres = MadreComunitaria.objects.select_related('usuario').prefetch_related('hogares_asignados').order_by('usuario__nombres')
    if nombre:
        madres = madres.filter(Q(usuario__nombres__icontains=nombre) | Q(usuario__apellidos__icontains=nombre))
    if hogar_asignado and hogar_asignado != '':
        madres = madres.filter(hogares_asignados__nombre_hogar__icontains=hogar_asignado)
    if escolaridad and escolaridad != '':
        madres = madres.filter(nivel_escolaridad=escolaridad)
    
    madres_list = madres.distinct()

    # --- Encabezados y configuración inicial ---
    headers = ['Nombres', 'Apellidos', 'Correo', 'Documento', 'Nivel de Escolaridad', 'Hogar Asignado']
    num_columns = len(headers)

    # --- Crear libro de Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Madres Comunitarias'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Madres Comunitarias", madres_list.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25
        
    # --- Llenar datos ---
    for madre in madres_list:
        hogar = madre.hogares_asignados.first()
        start_row += 1
        row_data = [madre.usuario.nombres, madre.usuario.apellidos, madre.usuario.correo, madre.usuario.documento, madre.get_nivel_escolaridad_display(), hogar.nombre_hogar if hogar else 'N/A']
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col_num, value=cell_value)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename="reporte_madres_comunitarias.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('administrador')
def reporte_hogares_excel(request):
    # Obtener filtros
    nombre_hogar = request.GET.get('nombre_hogar')
    regional_id = request.GET.get('regional', '')
    ciudad = request.GET.get('ciudad', '')
    madre_comunitaria = request.GET.get('madre', '') # Corregido para coincidir con el filtro de la lista
    ninos_matriculados = request.GET.get('ninos_matriculados')

    # Filtrar hogares
    hogares = HogarComunitario.objects.select_related('madre__usuario', 'regional', 'ciudad').annotate(
        num_ninos=Count('ninos')
    ).order_by('regional__nombre', 'nombre_hogar')
    if nombre_hogar:
        hogares = hogares.filter(nombre_hogar__icontains=nombre_hogar)
    if regional_id:
        hogares = hogares.filter(regional_id=regional_id)
    if ciudad:
        hogares = hogares.filter(ciudad__nombre__icontains=ciudad)
    if madre_comunitaria:
        hogares = hogares.filter(Q(madre__usuario__nombres__icontains=madre_comunitaria) | Q(madre__usuario__apellidos__icontains=madre_comunitaria))
    if ninos_matriculados:
        try:
            hogares = hogares.filter(num_ninos=int(ninos_matriculados))
        except (ValueError, TypeError):
            pass

    # --- Encabezados y configuración inicial ---
    headers = ['Nombre del Hogar', 'Madre Comunitaria', 'Regional', 'Ciudad', 'Dirección', 'Niños Matriculados', 'Capacidad', 'Estado']
    num_columns = len(headers)

    # --- Crear libro de Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hogares Comunitarios'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Hogares Comunitarios", hogares.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25

    # --- Llenar datos ---
    for hogar in hogares:
        start_row += 1
        row_data = [hogar.nombre_hogar, f"{hogar.madre.usuario.nombres} {hogar.madre.usuario.apellidos}", hogar.regional.nombre if hogar.regional else 'N/A', hogar.ciudad.nombre if hogar.ciudad else 'N/A', hogar.direccion, hogar.num_ninos, hogar.capacidad_maxima, hogar.get_estado_display()]
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col_num, value=cell_value)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename="reporte_hogares_comunitarios.xlsx"'
    wb.save(response)
    return response
    
@login_required
@rol_requerido('administrador')
def reporte_ninos_excel(request):
    # Obtener filtros de la URL
    hogar_nombre = request.GET.get('hogar', '')
    regional_id = request.GET.get('regional', '')
    madre_nombre = request.GET.get('madre', '')
    estado_nino = request.GET.get('estado', '')
    padre_nombre = request.GET.get('padre', '')

    # Consulta base
    ninos = Nino.objects.select_related(
        'hogar', 'hogar__madre__usuario', 'hogar__regional', 'padre__usuario'
    ).order_by('hogar__regional__nombre', 'hogar__nombre_hogar', 'apellidos')

    # Aplicar filtros
    if hogar_nombre:
        ninos = ninos.filter(hogar__nombre_hogar__icontains=hogar_nombre)
    if regional_id:
        ninos = ninos.filter(hogar__regional_id=regional_id)
    if madre_nombre:
        ninos = ninos.filter(
            Q(hogar__madre__usuario__nombres__icontains=madre_nombre) |
            Q(hogar__madre__usuario__apellidos__icontains=madre_nombre)
        )
    if padre_nombre:
        ninos = ninos.filter(
            Q(padre__usuario__nombres__icontains=padre_nombre) |
            Q(padre__usuario__apellidos__icontains=padre_nombre)
        )
    if estado_nino:
        ninos = ninos.filter(estado=estado_nino)

    # --- Encabezados y configuración inicial ---
    headers = [
        'Nombres', 'Apellidos', 'Fecha de Nacimiento', 'Edad (años)', 'Género',
        'Hogar Comunitario', 'Madre Comunitaria', 'Regional', 'Acudiente', 'Correo del Acudiente'
    ]
    num_columns = len(headers)

    # --- Crear libro de Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Niños Registrados'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte de Niños", ninos.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25

    # --- Llenar datos ---
    from datetime import date
    hoy = date.today()
    for nino in ninos:
        start_row += 1
        # --- CÁLCULO DE EDAD ---
        edad = (hoy.year - nino.fecha_nacimiento.year - ((hoy.month, hoy.day) < (nino.fecha_nacimiento.month, nino.fecha_nacimiento.day))) if nino.fecha_nacimiento else 'N/A'        

        # --- OBTENCIÓN DE NOMBRES RELACIONADOS (CORRECCIÓN FINAL) ---
        # Se concatena nombres y apellidos directamente, como se hace en otros reportes funcionales.
        madre_comunitaria_nombre = 'N/A'
        if nino.hogar and nino.hogar.madre and nino.hogar.madre.usuario:
            madre_comunitaria_nombre = f"{nino.hogar.madre.usuario.nombres} {nino.hogar.madre.usuario.apellidos}"

        acudiente_nombre = 'N/A'
        acudiente_correo = 'N/A'
        if nino.padre and nino.padre.usuario:
            acudiente_nombre = f"{nino.padre.usuario.nombres} {nino.padre.usuario.apellidos}"
            acudiente_correo = nino.padre.usuario.correo

        row_data = [
            nino.nombres, nino.apellidos, nino.fecha_nacimiento, edad, nino.get_genero_display(),
            nino.hogar.nombre_hogar if nino.hogar else 'N/A', madre_comunitaria_nombre, nino.hogar.regional.nombre if nino.hogar and nino.hogar.regional else 'N/A',
            acudiente_nombre, acudiente_correo
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col_num, value=cell_value)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename="reporte_ninos.xlsx"'
    wb.save(response)
    return response
@login_required
def eliminar_administrador(request, id):
    Usuario.objects.filter(id=id).delete()
    return redirect('listar_administradores')
# ... Tus otras funciones (home, admin_dashboard, crear_madre, etc.) ...

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Redirección por Rol (Se ejecuta después del login)
# ----------------------------------------------------
@login_required
def role_redirect(request):
    """
    Redirige al dashboard apropiado según el rol del usuario.
    Esta será la URL de redirección principal después de un login exitoso.
    """
    if not request.user.rol:
        return redirect('home')

    role = request.user.rol.nombre_rol.lower()

    if role == 'administrador':
        return redirect('dashboard_admin')
    elif role == 'madre_comunitaria':
        return redirect('madre_dashboard')
    elif role == 'padre':
        return redirect('padre_dashboard')

    return redirect('home')

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Dashboard de la Madre Comunitaria
# ----------------------------------------------------
@login_required
def madre_dashboard(request):
    from django.db.models import Count, Q, F, Value, CharField
    from django.utils import timezone
    from datetime import datetime, timedelta
    from planeaciones.models import Planeacion
    from novedades.models import Novedad
    import json
    
    if request.user.rol.nombre_rol != 'madre_comunitaria':
        return redirect('role_redirect')

    # Obtener hogar de la madre
    hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
    
    if not hogar_madre:
        return render(request, 'madre/dashboard.html', {'error': 'No tienes un hogar asignado.'})
    
    # Obtener todos los niños del hogar
    ninos = Nino.objects.filter(hogar=hogar_madre).select_related('padre', 'padre__usuario')
    total_ninos = ninos.count()
    
    # Información del hogar
    capacidad_maxima = hogar_madre.capacidad_maxima
    disponibles = capacidad_maxima - total_ninos
    porcentaje_ocupacion = round((total_ninos / capacidad_maxima) * 100) if capacidad_maxima > 0 else 0
    
    # Datos de género
    ninos_masculino = ninos.filter(genero='masculino').count()
    ninos_femenino = ninos.filter(genero='femenino').count()
    
    # Calcular edades
    hoy = datetime.now().date()
    edades_data = []
    edad_promedio = 0
    edad_minima = None
    edad_maxima = None
    
    for nino in ninos:
        if nino.fecha_nacimiento:
            edad = (hoy - nino.fecha_nacimiento).days // 365
            edades_data.append(edad)
            if edad_minima is None or edad < edad_minima:
                edad_minima = edad
            if edad_maxima is None or edad > edad_maxima:
                edad_maxima = edad
    
    if edades_data:
        edad_promedio = round(sum(edades_data) / len(edades_data), 1)
    
    # Distribución por edad (para gráfica)
    edad_0_2 = sum(1 for e in edades_data if e <= 2)
    edad_3_4 = sum(1 for e in edades_data if 3 <= e <= 4)
    edad_5 = sum(1 for e in edades_data if e >= 5)
    
    # Documentos
    ninos_con_cedula = ninos.filter(documento__isnull=False).count()
    ninos_con_vacunas = ninos.filter(carnet_vacunacion__isnull=False).count()
    ninos_con_afiliacion = ninos.filter(certificado_eps__isnull=False).count()
    
    # Asistencia últimas 4 semanas
    fecha_inicio = hoy - timedelta(days=28)
    asistencias = Asistencia.objects.filter(
        nino__hogar=hogar_madre,
        fecha__gte=fecha_inicio
    )
    
    asistencias_presentes = asistencias.filter(estado='Presente').count()
    asistencias_ausentes = asistencias.filter(estado='Ausente').count()
    total_dias_registro = (total_ninos * 28) if total_ninos > 0 else 1
    porcentaje_asistencia = round((asistencias_presentes / total_dias_registro) * 100) if total_dias_registro > 0 else 0
    
    # Datos de asistencia por niño (para tabla)
    ninos_asistencia = []
    for nino in ninos:
        asistencias_nino = Asistencia.objects.filter(
            nino=nino,
            fecha__gte=fecha_inicio
        )
        faltas = asistencias_nino.filter(estado='Ausente').count()
        asistencias_nino_count = asistencias_nino.filter(estado='Presente').count()
        porcentaje_nino = round((asistencias_nino_count / 28) * 100) if asistencias_nino_count > 0 else 0
        
        # Calcular edad
        edad = 0
        if nino.fecha_nacimiento:
            edad = (hoy - nino.fecha_nacimiento).days // 365
        
        # Determinar estado
        if faltas >= 7:
            estado = 'Alto riesgo'
            icono_estado = 'alert'
        elif faltas >= 4:
            estado = 'Advertencia'
            icono_estado = 'warning'
        else:
            estado = 'Normal'
            icono_estado = 'check'
        
        # Verificar documentos
        documentos_completos = 0
        if nino.documento:
            documentos_completos += 1
        if nino.carnet_vacunacion:
            documentos_completos += 1
        if nino.certificado_eps:
            documentos_completos += 1
        
        ninos_asistencia.append({
            'nino': nino,
            'edad': edad,
            'genero': nino.get_genero_display(),
            'faltas': faltas,
            'asistencias': asistencias_nino_count,
            'porcentaje': porcentaje_nino,
            'estado': estado,
            'icono_estado': icono_estado,
            'documentos': documentos_completos,
            'documentos_total': 3
        })
    
    # Novedades recientes
    fecha_novedades = hoy - timedelta(days=7)
    novedades_recientes = Novedad.objects.filter(
        nino__hogar=hogar_madre,
        fecha__gte=fecha_novedades
    ).select_related('nino').order_by('-fecha')[:3]
    
    # Planeaciones de esta semana
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    planeaciones_semana = Planeacion.objects.filter(
        madre=request.user,
        fecha__gte=inicio_semana,
        fecha__lte=inicio_semana + timedelta(days=6)
    ).count()
    
    # Datos para gráficas
    genero_data = json.dumps({
        'Masculino': ninos_masculino,
        'Femenino': ninos_femenino
    })
    
    edad_distribucion = json.dumps({
        '0-2 años': edad_0_2,
        '3-4 años': edad_3_4,
        '5+ años': edad_5
    })
    
    documentos_data = json.dumps({
        'Cédula': ninos_con_cedula,
        'Vacunas': ninos_con_vacunas,
        'Salud': ninos_con_afiliacion
    })
    
    # Calcular niños con documentos faltantes
    ninos_documentos_faltantes = []
    ninos_sin_documentos_completos = 0
    
    for nino in ninos:
        documentos_faltantes = []
        
        # Verificar registro civil
        if not nino.registro_civil_img:
            documentos_faltantes.append({
                'nombre': 'Registro Civil',
                'campo': 'registro_civil_img',
                'requerido': True
            })
        
        # Verificar carnet de vacunación
        if not nino.carnet_vacunacion:
            documentos_faltantes.append({
                'nombre': 'Carnet de Vacunación',
                'campo': 'carnet_vacunacion',
                'requerido': True
            })
        
        # Verificar certificado EPS
        if not nino.certificado_eps:
            documentos_faltantes.append({
                'nombre': 'Certificado de Afiliación a Salud (EPS)',
                'campo': 'certificado_eps',
                'requerido': True
            })
        
        # Verificar foto
        if not nino.foto:
            documentos_faltantes.append({
                'nombre': 'Foto del Niño',
                'campo': 'foto',
                'requerido': False
            })
        
        if documentos_faltantes:
            ninos_sin_documentos_completos += 1
            ninos_documentos_faltantes.append({
                'id': nino.id,
                'nombre': f"{nino.nombres} {nino.apellidos}",
                'documentos_faltantes': documentos_faltantes
            })
    
    # Convertir a JSON para JavaScript
    ninos_documentos_faltantes_json = json.dumps(ninos_documentos_faltantes)
    
    # Nombre completo de la madre
    nombre_madre = f"{request.user.nombres} {request.user.apellidos}"
    
    # Fecha de creación del hogar
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if hogar_madre.fecha_registro:
        fecha_registro_hogar = hogar_madre.fecha_registro
        fecha_creacion_hogar = f"{fecha_registro_hogar.day} de {meses[fecha_registro_hogar.month - 1]} de {fecha_registro_hogar.year}"
        hora_creacion_hogar = fecha_registro_hogar.strftime('%H:%M')
    else:
        fecha_creacion_hogar = "No disponible"
        hora_creacion_hogar = "--:--"
    
    context = {
        'nombre_madre': nombre_madre,
        'hogar_madre': hogar_madre,
        'hora_actual': hora_creacion_hogar,
        'fecha_actual': fecha_creacion_hogar,
        # Capacidad
        'total_ninos': total_ninos,
        'capacidad_maxima': capacidad_maxima,
        'disponibles': disponibles,
        'porcentaje_ocupacion': porcentaje_ocupacion,
        # Información de niños
        'ninos': ninos,
        'ninos_asistencia': ninos_asistencia,
        'ninos_masculino': ninos_masculino,
        'ninos_femenino': ninos_femenino,
        # Edades
        'edad_promedio': edad_promedio,
        'edad_minima': edad_minima,
        'edad_maxima': edad_maxima,
        # Documentos
        'ninos_con_cedula': ninos_con_cedula,
        'ninos_con_vacunas': ninos_con_vacunas,
        'ninos_con_afiliacion': ninos_con_afiliacion,
        'ninos_sin_documentos_completos': ninos_sin_documentos_completos,
        'ninos_documentos_faltantes': ninos_documentos_faltantes_json,
        # Asistencia
        'asistencias_presentes': asistencias_presentes,
        'asistencias_ausentes': asistencias_ausentes,
        'porcentaje_asistencia': porcentaje_asistencia,
        # Gráficas
        'genero_data': genero_data,
        'edad_distribucion': edad_distribucion,
        'documentos_data': documentos_data,
        # Otros
        'novedades_recientes': novedades_recientes,
        'planeaciones_semana': planeaciones_semana,
    }

    return render(request, 'madre/dashboard.html', context)


# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Dashboard del Padre de Familia
# ----------------------------------------------------
@login_required
def padre_dashboard(request):
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')
    # Importar modelos necesarios aquí para evitar importaciones circulares
    from novedades.models import Novedad
    from desarrollo.models import DesarrolloNino

    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos_qs = Nino.objects.filter(padre=padre).order_by('nombres')

        ninos_data = []
        for nino in ninos_qs:
            # Obtener la última asistencia registrada para el niño
            ultima_asistencia_obj = Asistencia.objects.filter(nino=nino).order_by('-fecha').first()
            if ultima_asistencia_obj:
                asistencia_info = {
                    'estado': ultima_asistencia_obj.estado,
                    'mensaje': f"El día {ultima_asistencia_obj.fecha.strftime('%d/%m/%Y')} estuvo {ultima_asistencia_obj.estado.lower()}."
                }
            else:
                asistencia_info = None

            # Obtener el último desarrollo y la última novedad
            ultimo_desarrollo = DesarrolloNino.objects.filter(nino=nino).order_by('-fecha_fin_mes').first()
            ultima_novedad = Novedad.objects.filter(nino=nino).order_by('-fecha').first()

            ninos_data.append({
                'nino': nino,
                'ultima_asistencia': asistencia_info,
                'ultimo_desarrollo': ultimo_desarrollo,
                'ultima_novedad': ultima_novedad
            })
        
        # 🆕 Obtener solicitudes de matrícula del padre (pendientes, en corrección, rechazadas recientes)
        solicitudes_activas = SolicitudMatriculacion.objects.filter(
            padre_solicitante=padre,
            estado__in=['pendiente', 'correccion']
        ).order_by('-fecha_creacion')
        
        # Agregar intentos_restantes a cada solicitud
        solicitudes_con_intentos = []
        for solicitud in solicitudes_activas:
            solicitud.intentos_restantes = 3 - solicitud.intentos_correccion
            solicitudes_con_intentos.append(solicitud)
        
        # Solicitudes rechazadas de los últimos 30 días (para mostrar feedback)
        hace_30_dias = timezone.now() - timedelta(days=30)
        solicitudes_rechazadas = SolicitudMatriculacion.objects.filter(
            padre_solicitante=padre,
            estado='rechazado',
            fecha_rechazo__gte=hace_30_dias
        ).order_by('-fecha_rechazo')[:3]  # Máximo 3 más recientes

        return render(request, 'padre/dashboard.html', {
            'ninos_data': ninos_data,
            'solicitudes_activas': solicitudes_con_intentos,
            'solicitudes_rechazadas': solicitudes_rechazadas,
        })
    except Padre.DoesNotExist:
        # Manejar el caso donde el padre no tiene un hijo asignado
        return render(request, 'padre/dashboard.html', {'error': 'No tienes un niño asignado.'})

def _get_logro_style(logro):
    """
    Retorna un color y un ícono para el logro mensual.
    """
    logro = (logro or "").lower()
    if "excelente" in logro or "alto" in logro:
        return "#27ae60", "fas fa-star"
    elif "medio" in logro or "regular" in logro:
        return "#f1c40f", "fas fa-check"
    elif "bajo" in logro or "alerta" in logro:
        return "#e74c3c", "fas fa-exclamation"
    else:
        return "#9B59B6", "fas fa-child"


# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Ver Desarrollo (Vista del Padre)
# ----------------------------------------------------
@login_required
def padre_ver_desarrollo(request, nino_id):
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')

    try:
        padre = Padre.objects.get(usuario=request.user)
        nino = get_object_or_404(Nino, id=nino_id, padre=padre)

        desarrollos_qs = DesarrolloNino.objects.filter(nino=nino).order_by('-fecha_fin_mes')

        mes_filtro = request.GET.get('mes', '')
        if mes_filtro:
            try:
                year, month = map(int, mes_filtro.split('-'))
                desarrollos_qs = desarrollos_qs.filter(fecha_fin_mes__year=year, fecha_fin_mes__month=month)
            except (ValueError, TypeError):
                mes_filtro = ''

        # Procesamiento de datos para la plantilla
        desarrollos_list = []
        for desarrollo in desarrollos_qs:
            accent_color, icono = _get_logro_style(desarrollo.logro_mes)

            desarrollos_list.append({
                'desarrollo': desarrollo,
                'accent_color': accent_color,
                'icono': icono,
            })

        # Paginación
        paginator = Paginator(desarrollos_list, 2) # 2 registros por página
        page_number = request.GET.get('page')
        desarrollos_paginados = paginator.get_page(page_number)

        return render(request, 'padre/desarrollo_list.html', {
            'nino': nino,
            'desarrollos': desarrollos_paginados,
            'filtros': {
                'mes': mes_filtro
            }
        })
    except (Padre.DoesNotExist, Nino.DoesNotExist):
        return redirect('padre_dashboard') # pragma: no cover

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

@login_required
def calendario_padres(request):
    tutor = request.user
    hoy = _date.today()

    year = int(request.GET.get("year", hoy.year))
    month = int(request.GET.get("month", hoy.month))

    first_day, total_days = monthrange(year, month)

    # Obtener los niños del tutor (perfil Padre)
    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos = Nino.objects.filter(padre=padre)
    except Padre.DoesNotExist:
        ninos = Nino.objects.none()

    # Planeaciones del mes
    planeaciones = Planeacion.objects.filter(fecha__year=year, fecha__month=month)

    # Novedades de todos los niños del padre (si existen)
    novedades = Novedad.objects.filter(
        nino__in=ninos,
        fecha__year=year,
        fecha__month=month
    ) if ninos.exists() else Novedad.objects.none()

    # Seguimientos de todos los niños del padre (si existen)
    seguimientos = SeguimientoDiario.objects.filter(
        nino__in=ninos,
        fecha__year=year,
        fecha__month=month
    ) if ninos.exists() else SeguimientoDiario.objects.none()

    eventos = {}

    for p in planeaciones:
        dia = p.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False})
        eventos[dia]["planeacion"] = True

    for n in novedades:
        dia = n.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False})
        eventos[dia]["novedad"] = True

    for s in seguimientos:
        dia = s.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False, "seguimiento": False})
        eventos[dia]["seguimiento"] = True

    return render(request, "padre/calendario_padres.html", {
        "year": year,
        "month": month,
        "month_name": MESES.get(month, str(month)),
        "first_day": first_day,
        "total_days": total_days,
        "eventos": eventos,
        "hoy": hoy,
    })


@login_required
def obtener_info(request):
    fecha = request.GET.get("fecha")
    # parsear fecha segura
    fecha_obj = None
    try:
        fecha_obj = _datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"planeacion": None, "novedad": None, "seguimientos": []})

    # Obtener los niños desde el perfil del Padre
    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos = Nino.objects.filter(padre=padre)
    except Padre.DoesNotExist:
        ninos = Nino.objects.none()

    planeacion = Planeacion.objects.filter(fecha=fecha_obj).first()
    novedades_del_dia = Novedad.objects.filter(nino__in=ninos, fecha=fecha_obj) if ninos.exists() else Novedad.objects.none()
    
    # --- NUEVA LÓGICA PARA SEGUIMIENTOS ---
    seguimientos_del_dia = SeguimientoDiario.objects.filter(
        nino__in=ninos, fecha=fecha_obj
    ).select_related('nino').prefetch_related('evaluaciones_dimension__dimension') if ninos.exists() else []

    seguimientos_data = []
    for s in seguimientos_del_dia:
        # Construir resumen para padres
        resumen = f"Hoy {s.nino.nombres} se mostró principalmente {s.get_estado_emocional_display().lower()} y su comportamiento fue {s.get_comportamiento_general_display().lower()}."
        if s.observacion_relevante and s.observaciones:
            resumen += f" La madre comunitaria observó: \"{s.observaciones}\"."

        # Obtener evaluaciones por dimensión
        evaluaciones = []
        for ev in s.evaluaciones_dimension.all():
            evaluaciones.append({
                "dimension": ev.dimension.nombre,
                "desempeno": ev.get_desempeno_display()
            })

        seguimientos_data.append({
            "nino_nombre": f"{s.nino.nombres} {s.nino.apellidos}",
            "fecha": s.fecha.strftime("%d/%m/%Y"),
            "resumen_dia": {
                "comportamiento": s.get_comportamiento_general_display(),
                "estado_emocional": s.get_estado_emocional_display(),
                "observacion_relevante": s.observaciones if s.observacion_relevante else None,
                "resumen_para_padres": resumen
            },
            "evaluaciones": evaluaciones,
            "valoracion_dia": s.valoracion,
            "valoracion_restante": 5 - (s.valoracion or 0)
        })

    # --- NUEVA LÓGICA PARA NOVEDADES ---
    novedades_data = []
    for n in novedades_del_dia:
        novedades_data.append({
            "tipo": n.get_tipo_display(),
            "descripcion": n.descripcion,
            "nino_nombre": f"{n.nino.nombres} {n.nino.apellidos}"
        })

    return JsonResponse({
        "planeacion": {
            "nombre": planeacion.nombre_experiencia if planeacion else None,
            "intencionalidad": planeacion.intencionalidad_pedagogica if planeacion else None,
            "materiales": planeacion.materiales_utilizar if planeacion else None
        } if planeacion else None,
        # Se envían como listas para manejar múltiples eventos por día
        "novedades": novedades_data,
        "seguimientos": seguimientos_data,
    })

@login_required
def ver_ficha_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    return render(request, 'madre/nino_ficha.html', {'nino': nino})

@login_required
def editar_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    # Asegurarse de que el padre y su usuario existan
    usuario_padre = nino.padre.usuario if nino.padre else None
    perfil_padre = nino.padre if nino.padre else None
    if request.method == 'POST':
        # Se instancian los formularios con los datos del POST, FILES y las instancias de los modelos
        nino_form = NinoForm(request.POST, request.FILES, instance=nino, prefix='nino')
        padre_form = PadreForm(request.POST, request.FILES, instance=usuario_padre, prefix='padre')

        if nino_form.is_valid() and padre_form.is_valid():
            # 💡 VALIDACIÓN: Verificar si el nuevo documento o correo ya existen en otro usuario.
            documento = padre_form.cleaned_data.get('documento')
            correo = padre_form.cleaned_data.get('correo')
            
            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exclude(id=usuario_padre.id).exists():
                messages.error(request, 'El documento o correo electrónico ingresado ya pertenece a otro usuario.')
            else:
                try:
                    with transaction.atomic():
                        # Guardar el niño con archivos
                        nino_actualizado = nino_form.save(commit=False)
                        # Actualizar campos personalizados
                        nino_actualizado.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                        # Manejar archivos del niño
                        if 'nino-foto' in request.FILES:
                            nino_actualizado.foto = request.FILES['nino-foto']
                        if 'nino-carnet_vacunacion' in request.FILES:
                            nino_actualizado.carnet_vacunacion = request.FILES['nino-carnet_vacunacion']
                        if 'nino-certificado_eps' in request.FILES:
                            nino_actualizado.certificado_eps = request.FILES['nino-certificado_eps']
                        if 'nino-registro_civil_img' in request.FILES:
                            nino_actualizado.registro_civil_img = request.FILES['nino-registro_civil_img']
                        nino_actualizado.save()
                        
                        # ManyToMany: tipos_discapacidad
                        if nino_actualizado.tiene_discapacidad:
                            nino_actualizado.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                        else:
                            nino_actualizado.tipos_discapacidad.clear()

                        # Guardar datos del Usuario del padre
                        usuario_actualizado = padre_form.save(commit=False)
                        usuario_actualizado.save()

                        # Actualizar el perfil del padre con archivos
                        if perfil_padre:
                            perfil_padre.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                            perfil_padre.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                            perfil_padre.estrato = padre_form.cleaned_data.get('estrato')
                            perfil_padre.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                            perfil_padre.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                            perfil_padre.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                            
                            # Manejar archivos del padre
                            if 'padre-documento_identidad_img' in request.FILES:
                                perfil_padre.documento_identidad_img = request.FILES['padre-documento_identidad_img']
                            if 'padre-clasificacion_sisben' in request.FILES:
                                perfil_padre.clasificacion_sisben = request.FILES['padre-clasificacion_sisben']
                            
                            perfil_padre.save()

                        messages.success(request, '¡La información del niño y su tutor ha sido actualizada!')
                        return redirect('listar_ninos')
                except Exception as e:
                    messages.error(request, f"Ocurrió un error inesperado: {e}")

        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')

    else:
        # Se instancian los formularios con los datos existentes para el método GET.
        nino_form = NinoForm(instance=nino, prefix='nino')

        # 💡 CORRECCIÓN: Pre-llenar el formulario del padre con los datos del usuario y del perfil.
        # Los datos del modelo Usuario se cargan con 'instance'.
        # Los datos del modelo Padre (como 'ocupacion') se cargan con 'initial'.
        initial_data_padre = {
            'documento': usuario_padre.documento if usuario_padre else ''
        }
        if perfil_padre:
            initial_data_padre['ocupacion'] = perfil_padre.ocupacion
            initial_data_padre['otra_ocupacion'] = perfil_padre.otra_ocupacion
        
        padre_form = PadreForm(instance=usuario_padre, prefix='padre', initial=initial_data_padre)

    # Obtener el hogar de la madre para el template
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        hogar_madre = nino.hogar  # Usar el hogar del niño como fallback

    return render(request, 'madre/nino_form_nuevo.html', {
        'nino_form': nino_form,
        'padre_form': padre_form,
        'nino': nino, # Se pasa el objeto nino para acceder a datos no editables en la plantilla
        'hogar_madre': hogar_madre,
        'modo_edicion': True,
        'titulo_form': 'Editar Información del Niño'
    })


@login_required
def editar_perfil(request):
    user = request.user
    rol = user.rol.nombre_rol

    # 1. Seleccionar el formulario y la instancia adecuados según el rol
    if rol == 'padre':
        # Para el padre, necesitamos la instancia de su perfil de Padre
        padre_profile, _ = Padre.objects.get_or_create(usuario=user)
        FormClass = PadrePerfilForm
        initial_data = {'ocupacion': padre_profile.ocupacion, 'estrato': padre_profile.estrato}
    elif rol == 'madre_comunitaria':
        FormClass = MadrePerfilForm
        initial_data = None
    else: # administrador
        FormClass = AdminPerfilForm
        initial_data = None

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, instance=user, initial=initial_data)
        if form.is_valid():
            # Guardar los datos del modelo Usuario
            user_instance = form.save(commit=False)
            # Si es un padre, guardar los campos adicionales en el modelo Padre
            if rol == 'padre':
                padre_profile.ocupacion = form.cleaned_data.get('ocupacion')
                padre_profile.estrato = form.cleaned_data.get('estrato')
                padre_profile.save()
            user_instance.save()
            messages.success(request, '¡Tu información ha sido actualizada exitosamente!')
            return redirect('editar_perfil')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        # Al cargar la página, inicializar el formulario con los datos existentes
        form = FormClass(instance=user, initial=initial_data)

    return render(request, 'perfil/editar_perfil.html', {'form': form})

@login_required
def listar_ninos(request):
    # Solo madres comunitarias pueden ver su listado
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Acceso denegado.')
        return redirect('home')
    try:
        hogar = HogarComunitario.objects.get(madre=request.user.madre_profile)
    except HogarComunitario.DoesNotExist:
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')
    
    ninos = Nino.objects.filter(hogar=hogar)
    
    # Contexto con información de matrícula exitosa si existe
    context = {
        'ninos': ninos
    }
    
    # Renderizar el template
    response = render(request, 'madre/nino_list.html', context)
    
    # Limpiar la sesión después de renderizar (para que el mensaje solo se muestre una vez)
    if 'matricula_exitosa' in request.session:
        del request.session['matricula_exitosa']
    if 'cambio_padre_exitoso' in request.session:
        del request.session['cambio_padre_exitoso']
    
    return response

@login_required
def panel_revision_solicitudes(request):
    """
    Vista para el panel de revisión de solicitudes de matriculación
    """
    # Verificar que sea madre comunitaria
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Acceso denegado.')
        return redirect('home')
    
    # Verificar que tenga hogar asignado
    try:
        hogar = HogarComunitario.objects.get(madre=request.user.madre_profile)
    except HogarComunitario.DoesNotExist:
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')
    
    return render(request, 'madre/panel_revision.html')


@login_required
def generar_reporte_ninos(request):
    # Aquí puedes generar el PDF o mostrar un mensaje temporal
    return render(request, 'madre/reporte_ninos.html')

@login_required
def eliminar_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    nino.delete()
    return redirect('listar_ninos')

@login_required
def gestion_ninos(request):
    # 1. Verificar rol y obtener el hogar de la madre
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria': # pragma: no cover
        messages.error(request, 'Acceso denegado.')
        return redirect('home')
    try:
        hogar = HogarComunitario.objects.get(madre=request.user.madre_profile)
    except HogarComunitario.DoesNotExist: # pragma: no cover
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')
    # 2. Filtrar los niños que pertenecen a ese hogar
    ninos_lista = Nino.objects.filter(hogar=hogar).order_by('nombres', 'apellidos')

    # 3. Aplicar paginación
    paginator = Paginator(ninos_lista, 3)  # 3 niños por página, como fue solicitado.
    page_number = request.GET.get('page')
    ninos_paginados = paginator.get_page(page_number)

    return render(request, 'madre/gestion_ninos_list.html', {'ninos': ninos_paginados})

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Cambiar Contraseña del Usuario
# ----------------------------------------------------
@login_required
def cambiar_contrasena(request):
    if request.method == 'POST':
        form = SetPasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Actualizar la sesión para que el usuario no sea deslogueado
            update_session_auth_hash(request, user)
            messages.success(request, '¡Tu contraseña ha sido actualizada exitosamente!')
            # Redirigir a la misma página para que el usuario vea el mensaje de éxito.
            return redirect('cambiar_contrasena')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = SetPasswordForm(request.user)
    
    return render(request, 'perfil/cambiar_contrasena.html', {'form': form})

@login_required
def detalles_madre_json(request, id):
    try:
        madre_perfil = get_object_or_404(MadreComunitaria, id=id)
        usuario = madre_perfil.usuario
        hogar = HogarComunitario.objects.filter(madre=madre_perfil).first() if hasattr(HogarComunitario, 'madre') else None

        # Obtener la URL de la foto de la madre
        if madre_perfil.foto_madre and hasattr(madre_perfil.foto_madre, 'url'):
            foto_madre_url = madre_perfil.foto_madre.url
        else:
            foto_madre_url = ''

        # Documentos adjuntos
        documentos = {
            'documento_identidad': madre_perfil.documento_identidad_pdf.url if madre_perfil.documento_identidad_pdf else None,
            'certificado_escolaridad': madre_perfil.certificado_escolaridad_pdf.url if madre_perfil.certificado_escolaridad_pdf else None,
            'certificado_antecedentes': madre_perfil.certificado_antecedentes_pdf.url if madre_perfil.certificado_antecedentes_pdf else None,
            'certificado_medico': madre_perfil.certificado_medico_pdf.url if madre_perfil.certificado_medico_pdf else None,
            'certificado_residencia': madre_perfil.certificado_residencia_pdf.url if madre_perfil.certificado_residencia_pdf else None,
            'cartas_recomendacion': madre_perfil.cartas_recomendacion_pdf.url if madre_perfil.cartas_recomendacion_pdf else None,
            'certificado_laboral': madre_perfil.certificado_laboral.url if madre_perfil.certificado_laboral else None,
            'carta_disponibilidad': madre_perfil.carta_disponibilidad.url if madre_perfil.carta_disponibilidad else None,
            'firma_digital': madre_perfil.firma_digital.url if madre_perfil.firma_digital else None,
        }

        data = {
            'usuario': {
                'nombres': usuario.nombres,
                'apellidos': usuario.apellidos,
                'tipo_documento': usuario.get_tipo_documento_display(),
                'documento': usuario.documento,
                'correo': usuario.correo,
                'telefono': usuario.telefono,
                'direccion': usuario.direccion,
            },
            'perfil': {
                'nivel_escolaridad': madre_perfil.nivel_escolaridad,
                'titulo_obtenido': madre_perfil.titulo_obtenido or "No especificado",
                'institucion': madre_perfil.institucion or "No especificada",
                'experiencia_previa': madre_perfil.experiencia_previa or "No especificada",
                'foto_madre_url': foto_madre_url,
                'no_retirado_icbf': madre_perfil.no_retirado_icbf,
                'fecha_registro': madre_perfil.fecha_registro.strftime('%d/%m/%Y %H:%M') if madre_perfil.fecha_registro else "N/A",
            },
            'documentos': documentos,
            'hogar': {
                'nombre_hogar': hogar.nombre_hogar if hogar else "Sin hogar asignado",
                'direccion': hogar.direccion if hogar else "N/A",
                'localidad': hogar.localidad if hogar else "N/A",
                'barrio': hogar.barrio if hogar else "N/A",
                'capacidad_maxima': getattr(hogar, 'capacidad_maxima', 'N/A') if hogar else "N/A",
                'regional': hogar.regional.nombre if hogar and hogar.regional else "N/A",
                'ciudad': hogar.ciudad.nombre if hogar and hogar.ciudad else "N/A",
            }
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def cargar_ciudades(request):

    """
    Retorna JSON con las ciudades de la regional indicada.
    Parámetro: ?regional_id=ID
    """
    regional_id = request.GET.get("regional_id")
    if not regional_id:
        return JsonResponse([], safe=False)
    try:
        ciudades = Ciudad.objects.filter(regional_id=regional_id).order_by('nombre')
        data = [{"id": c.id, "nombre": c.nombre} for c in ciudades]
        return JsonResponse(data, safe=False)
    except Exception:
        return JsonResponse([], safe=False)


# ----------------------------------------------------
# 🆕 NUEVAS VISTAS PARA MEJORAS DE MATRÍCULA
# ----------------------------------------------------

@login_required
def matricular_nino_a_padre_existente(request):
    """Matricular un niño nuevo a un padre que ya existe en el sistema"""
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden matricular niños.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        buscar_form = BuscarPadreForm(request.POST, prefix='buscar')
        nino_form = NinoSoloForm(request.POST, request.FILES, prefix='nino')
        
        # Verificar si se ha encontrado un padre
        padre_id = request.POST.get('padre_seleccionado')
        
        if padre_id and nino_form.is_valid():
            try:
                with transaction.atomic():
                    # Obtener el padre seleccionado
                    padre_obj = get_object_or_404(Padre, id=padre_id)
                    
                    # Crear el niño y asociarlo al padre
                    nino = nino_form.save(commit=False)
                    nino.hogar = hogar_madre
                    nino.padre = padre_obj
                    nino.tipo_sangre = nino_form.cleaned_data.get('tipo_sangre')
                    nino.parentesco = nino_form.cleaned_data.get('parentesco')
                    nino.tiene_discapacidad = nino_form.cleaned_data.get('tiene_discapacidad', False)
                    nino.otra_discapacidad = nino_form.cleaned_data.get('otra_discapacidad', '')
                    nino.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                    nino.save()
                    
                    # ManyToMany: tipos_discapacidad
                    if nino.tiene_discapacidad:
                        nino.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                    else:
                        nino.tipos_discapacidad.clear()
                    
                    messages.success(request, f'Niño {nino.nombres} matriculado correctamente al padre {padre_obj.usuario.get_full_name()}.')
                    return redirect('listar_ninos')
                    
            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado: {e}")
        else:
            if not padre_id:
                messages.error(request, 'Debe buscar y seleccionar un padre antes de continuar.')
            if not nino_form.is_valid():
                messages.error(request, 'Por favor corrige los errores en el formulario del niño.')
    else:
        buscar_form = BuscarPadreForm(prefix='buscar')
        nino_form = NinoSoloForm(prefix='nino')

    return render(request, 'madre/matricular_a_padre_existente.html', {
        'hogar_madre': hogar_madre,
        'buscar_form': buscar_form,
        'nino_form': nino_form,
    })


@login_required
def buscar_padre_ajax(request):
    """Vista AJAX para buscar padre por documento en matriculación a padre existente"""
    documento = request.GET.get('documento', '').strip()
    data = {'encontrado': False, 'padre': None}

    if documento:
        try:
            # Buscar usuario con rol de padre
            usuario_padre = Usuario.objects.filter(
                documento=documento, 
                rol__nombre_rol='padre',
                is_active=True
            ).first()
            
            if usuario_padre:
                try:
                    padre_profile = Padre.objects.get(usuario=usuario_padre)
                    data.update({
                        'encontrado': True,
                        'padre': {
                            'id': padre_profile.id,
                            'nombres': usuario_padre.nombres,
                            'apellidos': usuario_padre.apellidos,
                            'documento': usuario_padre.documento,
                            'correo': usuario_padre.correo,
                            'telefono': usuario_padre.telefono,
                            'direccion': usuario_padre.direccion,
                            'ocupacion': padre_profile.get_ocupacion_display() if padre_profile.ocupacion else 'No especificada'
                        }
                    })
                except Padre.DoesNotExist:
                    data['mensaje'] = 'Usuario encontrado pero sin perfil de padre completo.'
            else:
                data['mensaje'] = 'No se encontró un padre con ese documento.'
        except Exception as e:
            data['error'] = f'Error en la búsqueda: {str(e)}'

    return JsonResponse(data)


@login_required
def cambiar_padre_de_nino(request):
    """Cambiar la asignación de padre de un niño existente"""
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden cambiar asignaciones.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        cambiar_form = CambiarPadreForm(hogar=hogar_madre, data=request.POST, prefix='cambiar')
        buscar_form = BuscarPadreForm(request.POST, prefix='buscar')
        padre_form = PadreForm(request.POST, request.FILES, prefix='padre')
        
        # Verificar qué acción se está realizando
        accion = request.POST.get('accion')
        nino_id = request.POST.get('nino_seleccionado')
        nuevo_padre_id = request.POST.get('nuevo_padre_id')
        motivo_cambio = request.POST.get('motivo_cambio')
        observaciones_motivo = request.POST.get('observaciones_motivo', '')
        
        # Log para debugging
        print(f"[DEBUG] Acción: {accion}")
        print(f"[DEBUG] Niño ID: {nino_id}")
        print(f"[DEBUG] Nuevo Padre ID: {nuevo_padre_id}")
        print(f"[DEBUG] Motivo: {motivo_cambio}")
        print(f"[DEBUG] Observaciones: {observaciones_motivo}")
        print(f"[DEBUG] POST data: {request.POST}")
        
        if accion == 'cambiar_existente' and nino_id and nuevo_padre_id:
            # Validar que se haya seleccionado un motivo
            if not motivo_cambio:
                messages.error(request, 'Debe seleccionar el motivo del cambio de asignación.')
                return redirect('cambiar_padre_nino')
            
            # Cambiar a padre existente
            print(f"[DEBUG] Cambiando niño {nino_id} a padre existente {nuevo_padre_id}")
            try:
                with transaction.atomic():
                    nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
                    nuevo_padre = get_object_or_404(Padre, id=nuevo_padre_id)
                    padre_anterior = nino.padre
                    
                    print(f"[DEBUG] Niño encontrado: {nino.nombres} {nino.apellidos}")
                    print(f"[DEBUG] Padre anterior: {padre_anterior.usuario.get_full_name() if padre_anterior else 'Sin padre'}")
                    print(f"[DEBUG] Nuevo padre: {nuevo_padre.usuario.get_full_name()}")
                    
                    nino.padre = nuevo_padre
                    nino.save()
                    
                    mensaje_exito = f'El niño {nino.nombres} {nino.apellidos} ha sido asignado correctamente '
                    if padre_anterior:
                        mensaje_exito += f'del padre {padre_anterior.usuario.get_full_name()} '
                    mensaje_exito += f'al padre {nuevo_padre.usuario.get_full_name()}.'
                    
                    # Guardar en la sesión para el SweetAlert
                    request.session['cambio_padre_exitoso'] = {
                        'nombre': f'{nino.nombres} {nino.apellidos}',
                        'mensaje': mensaje_exito,
                        'motivo': motivo_cambio,
                        'observaciones': observaciones_motivo
                    }
                    
                    messages.success(request, mensaje_exito)
                    print(f"[DEBUG] Cambio exitoso - Motivo: {motivo_cambio}")
                    return redirect('listar_ninos')
            except Exception as e:
                print(f"[ERROR] Error al cambiar padre: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error al cambiar padre: {e}")
                
        elif accion == 'cambiar_nuevo' and nino_id and padre_form.is_valid():
            # Validar que se haya seleccionado un motivo
            if not motivo_cambio:
                messages.error(request, 'Debe seleccionar el motivo del cambio de asignación.')
                return redirect('cambiar_padre_nino')
            
            # Cambiar a nuevo padre
            print(f"[DEBUG] Creando nuevo padre y cambiando niño {nino_id}")
            try:
                with transaction.atomic():
                    nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
                    padre_anterior = nino.padre
                    
                    # Crear nuevo usuario padre
                    rol_padre = Rol.objects.get(nombre_rol='padre')
                    doc_padre = padre_form.cleaned_data['documento']
                    tipo_doc = padre_form.cleaned_data['tipo_documento']
                    
                    usuario_padre, created = Usuario.objects.get_or_create(
                        documento=doc_padre,
                        rol=rol_padre,
                        defaults={
                            'tipo_documento': tipo_doc,
                            'nombres': padre_form.cleaned_data['nombres'],
                            'apellidos': padre_form.cleaned_data['apellidos'],
                            'correo': padre_form.cleaned_data['correo'],
                            'telefono': padre_form.cleaned_data['telefono'],
                            'direccion': padre_form.cleaned_data.get('direccion', ''),
                        }
                    )

                    if created:
                        usuario_padre.set_password(str(doc_padre))
                        usuario_padre.save()

                    # Crear perfil del padre
                    padre_obj, _ = Padre.objects.get_or_create(usuario=usuario_padre)
                    padre_obj.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                    padre_obj.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                    padre_obj.estrato = padre_form.cleaned_data.get('estrato')
                    padre_obj.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                    padre_obj.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                    padre_obj.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                    padre_obj.save()
                    
                    # Actualizar niño
                    nino.padre = padre_obj
                    nino.save()
                    
                    messages.success(request, 
                        f'El niño {nino.nombres} {nino.apellidos} ha sido asignado correctamente '
                        f'del padre {padre_anterior.usuario.get_full_name()} '
                        f'al nuevo padre {padre_obj.usuario.get_full_name()}.'
                    )
                    return redirect('listar_ninos')
            except Exception as e:
                print(f"[ERROR] Error al crear nuevo padre y cambiar: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error al cambiar padre: {e}")
        else:
            print(f"[DEBUG] Datos incompletos o formulario inválido")
            print(f"[DEBUG] Acción válida: {accion in ['cambiar_existente', 'cambiar_nuevo']}")
            print(f"[DEBUG] Niño ID presente: {bool(nino_id)}")
            if accion == 'cambiar_existente':
                print(f"[DEBUG] Nuevo padre ID presente: {bool(nuevo_padre_id)}")
            if accion == 'cambiar_nuevo':
                print(f"[DEBUG] Padre form válido: {padre_form.is_valid()}")
                if not padre_form.is_valid():
                    print(f"[DEBUG] Errores del formulario: {padre_form.errors}")
            messages.error(request, 'Datos incompletos para realizar el cambio.')
    else:
        cambiar_form = CambiarPadreForm(hogar=hogar_madre, prefix='cambiar')
        buscar_form = BuscarPadreForm(prefix='buscar')
        padre_form = PadreForm(prefix='padre')

    return render(request, 'madre/cambiar_padre_nino.html', {
        'hogar_madre': hogar_madre,
        'cambiar_form': cambiar_form,
        'buscar_form': buscar_form,
        'padre_form': padre_form,
    })

@login_required
def actualizar_foto_perfil(request):
    """
    Vista AJAX para que la madre comunitaria actualice su foto de perfil
    directamente desde el navbar.
    """
    if request.method == 'POST' and request.user.rol.nombre_rol == 'madre_comunitaria':
        try:
            madre_profile = get_object_or_404(MadreComunitaria, usuario=request.user)
            
            # El nombre del input en el form debe ser 'foto_madre'
            new_photo = request.FILES.get('foto_madre')

            if new_photo:
                madre_profile.foto_madre = new_photo
                madre_profile.save()
                
                # Devolver la URL de la nueva foto
                return JsonResponse({'success': True, 'new_photo_url': madre_profile.foto_madre.url})
            else:
                return JsonResponse({'success': False, 'error': 'No se recibió ningún archivo.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Método no permitido o rol incorrecto.'}, status=405)


@login_required
def subir_documentos_nino(request):
    """
    Vista AJAX para subir documentos faltantes de un niño desde el dashboard de la madre.
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'success': False, 'error': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener el ID del niño
            nino_id = request.POST.get('nino_id')
            if not nino_id:
                return JsonResponse({'success': False, 'error': 'No se especificó el niño.'}, status=400)
            
            # Verificar que el niño pertenezca al hogar de la madre
            nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
            
            # Actualizar documentos
            documentos_actualizados = []
            
            # Registro civil
            if 'registro_civil_img' in request.FILES:
                nino.registro_civil_img = request.FILES['registro_civil_img']
                documentos_actualizados.append('Registro Civil')
            
            # Carnet de vacunación
            if 'carnet_vacunacion' in request.FILES:
                nino.carnet_vacunacion = request.FILES['carnet_vacunacion']
                documentos_actualizados.append('Carnet de Vacunación')
            
            # Certificado EPS
            if 'certificado_eps' in request.FILES:
                nino.certificado_eps = request.FILES['certificado_eps']
                documentos_actualizados.append('Certificado EPS')
            
            # Foto
            if 'foto' in request.FILES:
                nino.foto = request.FILES['foto']
                documentos_actualizados.append('Foto')
            
            # Guardar cambios
            if documentos_actualizados:
                nino.save()
                return JsonResponse({
                    'success': True,
                    'message': f'Documentos actualizados: {", ".join(documentos_actualizados)}',
                    'documentos_actualizados': documentos_actualizados
                })
            else:
                return JsonResponse({'success': False, 'error': 'No se seleccionó ningún documento para subir.'}, status=400)
            
        except Nino.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'El niño no existe o no pertenece a tu hogar.'}, status=404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Error al procesar los documentos: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


# ============================================
# SISTEMA DE INVITACIONES DE MATRICULACIÓN
# ============================================

@login_required
def padre_ver_solicitud_matricula(request, solicitud_id):
    """
    Vista para que el padre vea el detalle de su solicitud de matrícula desde el dashboard.
    Solo puede ver sus propias solicitudes.
    """
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')
    
    try:
        padre = Padre.objects.get(usuario=request.user)
        solicitud = get_object_or_404(
            SolicitudMatriculacion,
            id=solicitud_id,
            padre_solicitante=padre
        )
        
        # Calcular intentos restantes
        solicitud.intentos_restantes = 3 - solicitud.intentos_correccion
        
        return render(request, 'padre/solicitud_detalle.html', {
            'solicitud': solicitud,
        })
        
    except Padre.DoesNotExist:
        messages.error(request, 'No tienes un perfil de padre asignado.')
        return redirect('padre_dashboard')


@login_required
def padre_corregir_solicitud(request, solicitud_id):
    """
    Vista para que el padre corrija su solicitud de matrícula directamente desde el dashboard.
    Solo muestra y permite editar los campos que la madre marcó para corrección.
    """
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')
    
    try:
        padre = Padre.objects.get(usuario=request.user)
        solicitud = get_object_or_404(
            SolicitudMatriculacion,
            id=solicitud_id,
            padre_solicitante=padre,
            estado='correccion'
        )
        
        # Verificar que no haya excedido el límite de intentos
        if solicitud.intentos_correccion >= 3:
            messages.error(
                request,
                'Has excedido el límite de 3 intentos de corrección. '
                'Esta solicitud será rechazada automáticamente. Por favor, envía una nueva solicitud.'
            )
            return redirect('padre_dashboard')
        
        if request.method == 'POST':
            try:
                from core.models import Discapacidad
                from django.contrib.contenttypes.models import ContentType
                from notifications.models import Notification
                
                # Solo procesar campos que están en campos_corregir
                campos_corregir = solicitud.campos_corregir or []
                
                if not campos_corregir:
                    messages.warning(request, 'No hay campos para corregir en esta solicitud.')
                    return redirect('padre_dashboard')
                
                # Actualizar solo los campos marcados para corrección
                campos_texto = [
                    'nombres_nino', 'apellidos_nino', 'documento_nino',
                    'genero_nino', 'tipo_sangre_nino', 'parentesco', 'observaciones_nino'
                ]
                
                campos_archivos = [
                    'foto_nino', 'carnet_vacunacion_nino', 'certificado_eps_nino', 'registro_civil_nino'
                ]
                
                # Actualizar campos de texto
                for campo in campos_corregir:
                    if campo in campos_texto:
                        valor = request.POST.get(campo, '').strip()
                        if valor:
                            setattr(solicitud, campo, valor)
                        else:
                            messages.error(request, f'El campo {campo} es obligatorio.')
                            return redirect('padre_corregir_solicitud', solicitud_id=solicitud_id)
                
                # Actualizar fecha de nacimiento si está en corrección
                if 'fecha_nacimiento_nino' in campos_corregir:
                    fecha_nac = request.POST.get('fecha_nacimiento_nino')
                    if fecha_nac:
                        from datetime import datetime
                        solicitud.fecha_nacimiento_nino = datetime.strptime(fecha_nac, '%Y-%m-%d').date()
                    else:
                        messages.error(request, 'La fecha de nacimiento es obligatoria.')
                        return redirect('padre_corregir_solicitud', solicitud_id=solicitud_id)
                
                # Actualizar archivos (documentos)
                for campo in campos_corregir:
                    if campo in campos_archivos:
                        if campo in request.FILES:
                            setattr(solicitud, campo, request.FILES[campo])
                        else:
                            # Si es un archivo marcado para corrección, DEBE cargarse uno nuevo
                            campos_nombres = {
                                'foto_nino': 'Foto del niño',
                                'carnet_vacunacion_nino': 'Carnet de vacunación',
                                'certificado_eps_nino': 'Certificado EPS',
                                'registro_civil_nino': 'Registro civil'
                            }
                            messages.error(
                                request,
                                f'Debe cargar un archivo nuevo para: {campos_nombres.get(campo, campo)}'
                            )
                            return redirect('padre_corregir_solicitud', solicitud_id=solicitud_id)
                
                # Manejar discapacidad si está en corrección
                if 'tiene_discapacidad' in campos_corregir or 'tipos_discapacidad' in campos_corregir:
                    tiene_discapacidad = request.POST.get('tiene_discapacidad', 'false')
                    solicitud.tiene_discapacidad = tiene_discapacidad == 'true'
                    
                    if solicitud.tiene_discapacidad:
                        tipos_ids = request.POST.getlist('tipos_discapacidad')
                        solicitud.tipos_discapacidad = [int(id) for id in tipos_ids if id.isdigit()]
                        solicitud.otra_discapacidad = request.POST.get('otra_discapacidad', '').strip()
                    else:
                        solicitud.tipos_discapacidad = None
                        solicitud.otra_discapacidad = None
                
                # Cambiar estado a pendiente y limpiar campos de corrección
                solicitud.estado = 'pendiente'
                solicitud.campos_corregir = None
                solicitud.save()
                
                # Crear notificación para la madre comunitaria
                content_type = ContentType.objects.get_for_model(SolicitudMatriculacion)
                Notification.objects.create(
                    recipient=solicitud.hogar.madre.usuario,
                    title=f"Correcciones Realizadas: {solicitud.nombres_nino}",
                    message=f"El padre {padre.usuario.get_full_name()} ha realizado las correcciones solicitadas.",
                    level='info',
                    content_type=content_type,
                    object_id=solicitud.id
                )
                
                # Enviar email informativo a la madre
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    asunto = f'Correcciones Realizadas - {solicitud.hogar.nombre_hogar}'
                    mensaje_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px;">
                        <h2 style="color: #28a745;">Correcciones Realizadas</h2>
                        <p>Estimada madre comunitaria,</p>
                        <p>El padre <strong>{padre.usuario.get_full_name()}</strong> ha realizado las correcciones solicitadas para la matrícula de <strong>{solicitud.nombres_nino}</strong>.</p>
                        <p>Por favor, revisa nuevamente la solicitud en tu panel de revisión.</p>
                        <br>
                        <p>Saludos cordiales,<br><strong>Sistema ICBF Conecta</strong></p>
                    </body>
                    </html>
                    """
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.hogar.madre.usuario.correo],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
                except Exception as e:
                    print(f"Error al enviar email a madre: {e}")
                
                messages.success(
                    request,
                    f'¡Correcciones enviadas exitosamente! La solicitud de {solicitud.nombres_nino} está nuevamente en revisión.'
                )
                return redirect('padre_dashboard')
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error al procesar las correcciones: {str(e)}')
                return redirect('padre_corregir_solicitud', solicitud_id=solicitud_id)
        
        # GET: Mostrar formulario de corrección
        from core.models import Discapacidad
        discapacidades = Discapacidad.objects.all()
        
        # Calcular intentos restantes
        intentos_restantes = 3 - solicitud.intentos_correccion
        
        return render(request, 'padre/solicitud_corregir.html', {
            'solicitud': solicitud,
            'campos_corregir': solicitud.campos_corregir or [],
            'intentos_restantes': intentos_restantes,
            'discapacidades': discapacidades,
        })
        
    except Padre.DoesNotExist:
        messages.error(request, 'No tienes un perfil de padre asignado.')
        return redirect('padre_dashboard')


# 🆕 NUEVA FUNCIONALIDAD: SOLICITUD INICIADA POR PADRE (FORMULARIO COMPLETO)
@login_required
def padre_solicitar_matricula(request):
    """
    Permite al padre/tutor crear una solicitud de matrícula COMPLETA.
    El padre llena TODOS los datos del niño + sube documentos en un solo paso.
    Sistema valida cupos ANTES de guardar.
    """
    # Importar modelos necesarios al inicio
    from core.models import Nino, Discapacidad
    
    if request.method == 'POST':
        try:
            # Verificar que sea padre
            if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'padre':
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Solo los padres/tutores pueden crear solicitudes de matrícula.'
                }, status=403)
            
            # Obtener perfil del padre
            try:
                padre_profile = request.user.padre_profile
            except Padre.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'No se encontró el perfil de padre asociado.'
                }, status=400)
            
            # ========== OBTENER DATOS DEL FORMULARIO ==========
            hogar_id = request.POST.get('hogar_id')
            
            # Datos básicos del niño
            nombres_nino = request.POST.get('nombres_nino', '').strip()
            apellidos_nino = request.POST.get('apellidos_nino', '').strip()
            tipo_documento_nino = request.POST.get('tipo_documento_nino', '').strip()  # Guardará en observaciones por ahora
            documento_nino = request.POST.get('documento_nino', '').strip()
            fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip()
            genero = request.POST.get('genero', '').strip()
            pais_nacimiento_id = request.POST.get('pais_nacimiento', '').strip()  # Guardará en observaciones por ahora
            tipo_sangre = request.POST.get('tipo_sangre', '').strip()
            
            # Discapacidad
            tiene_discapacidad = request.POST.get('tiene_discapacidad') == 'on'
            tipos_discapacidad_ids = request.POST.getlist('tipos_discapacidad[]') if tiene_discapacidad else []
            otra_discapacidad = request.POST.get('otra_discapacidad', '').strip() if tiene_discapacidad else ''
            
            # Documentos
            foto_nino = request.FILES.get('foto_nino')
            carnet_vacunacion = request.FILES.get('carnet_vacunacion_nino')
            certificado_eps = request.FILES.get('certificado_eps_nino')
            registro_civil = request.FILES.get('registro_civil_nino')
            
            # ========== VALIDACIONES ==========
            errores = []
            
            if not all([hogar_id, nombres_nino, apellidos_nino, fecha_nacimiento, genero]):
                errores.append('Todos los campos obligatorios deben ser completados.')
            
            if not tipo_documento_nino:
                errores.append('El tipo de documento es obligatorio.')
            
            # Validar edad del niño (entre 1 y 5 años)
            if fecha_nacimiento:
                try:
                    fecha_nac_obj = datetime.strptime(fecha_nacimiento, '%Y-%m-%d').date()
                    hoy = timezone.now().date()
                    edad_dias = (hoy - fecha_nac_obj).days
                    edad_anios = edad_dias / 365.25
                    
                    if edad_anios < 1:
                        errores.append('El niño debe tener al menos 1 año de edad para ser matriculado.')
                    elif edad_anios > 5:
                        errores.append('El niño no puede tener más de 5 años de edad. Los hogares comunitarios solo admiten niños entre 1 y 5 años.')
                except ValueError:
                    errores.append('La fecha de nacimiento no tiene un formato válido.')
            
            if not foto_nino:
                errores.append('La foto del niño es obligatoria.')
            if not carnet_vacunacion:
                errores.append('El carné de vacunación es obligatorio.')
            if not certificado_eps:
                errores.append('El certificado de EPS es obligatorio.')
            if not registro_civil:
                errores.append('El registro civil es obligatorio.')
            
            if errores:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': ' '.join(errores)
                }, status=400)
            
            # 🔍 VALIDAR HOGAR
            ninos_padre = Nino.objects.filter(padre=padre_profile, estado='activo').select_related('hogar')
            
            if ninos_padre.exists():
                hogar_correcto = ninos_padre.first().hogar
                if str(hogar_correcto.id) != str(hogar_id):
                    return JsonResponse({
                        'status': 'error',
                        'mensaje': f'Solo puedes solicitar matrícula para el hogar donde ya tienes hijos registrados: {hogar_correcto.nombre_hogar}'
                    }, status=400)
                hogar = hogar_correcto
            else:
                try:
                    hogar = HogarComunitario.objects.get(id=hogar_id, estado='aprobado')
                except HogarComunitario.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'mensaje': 'El hogar comunitario seleccionado no existe o no está activo.'
                    }, status=404)
            
            # 🚨 VALIDAR CUPOS ANTES DE CREAR LA SOLICITUD
            # Buscar capacidad en múltiples campos (por compatibilidad con diferentes flujos)
            capacidad_total = hogar.capacidad or hogar.capacidad_calculada or hogar.capacidad_maxima or 0
            ninos_activos = Nino.objects.filter(hogar=hogar, estado='activo').count()
            cupos_disponibles = capacidad_total - ninos_activos
            
            if cupos_disponibles <= 0:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': f'Lo sentimos, el hogar {hogar.nombre_hogar} no tiene cupos disponibles en este momento. Capacidad: {capacidad_total}, Niños activos: {ninos_activos}.'
                }, status=400)
            
            # 🆕 VALIDAR QUE NO EXISTA UN NIÑO YA MATRICULADO CON ESE DOCUMENTO EN EL HOGAR
            if documento_nino:
                nino_existente = Nino.objects.filter(
                    documento=documento_nino,
                    hogar=hogar,
                    estado='activo'
                ).first()
                
                if nino_existente:
                    return JsonResponse({
                        'status': 'error',
                        'mensaje': f'Ya existe un niño matriculado con el documento {documento_nino} en el hogar {hogar.nombre_hogar}.'
                    }, status=400)
            
            # Validar solicitudes duplicadas
            solicitud_existente = SolicitudMatriculacion.objects.filter(
                padre_solicitante=padre_profile,
                hogar=hogar,
                estado__in=['pendiente', 'correccion']
            ).first()
            
            if solicitud_existente:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': f'Ya tienes una solicitud pendiente para el hogar {hogar.nombre_hogar}. Por favor espera la respuesta.'
                }, status=400)
            
            # ========== CREAR SOLICITUD COMPLETA ==========
            token = str(uuid.uuid4())
            
            # Preparar observaciones adicionales con info de tipo_documento y país
            observaciones_extra = []
            if tipo_documento_nino:
                observaciones_extra.append(f"Tipo Documento: {tipo_documento_nino}")
            if pais_nacimiento_id:
                # Buscar el nombre del país en las CHOICES (Nino ya importado al inicio)
                pais_nombre = dict(Nino.PAIS_NACIMIENTO_CHOICES).get(pais_nacimiento_id, pais_nacimiento_id)
                observaciones_extra.append(f"País de Nacimiento: {pais_nombre}")
            
            observaciones_completas = " | ".join(observaciones_extra) if observaciones_extra else ""
            
            solicitud = SolicitudMatriculacion.objects.create(
                hogar=hogar,
                email_acudiente=request.user.correo,
                token=token,
                fecha_expiracion=timezone.now() + timedelta(hours=48),
                estado='pendiente',
                tipo_solicitud='solicitud_padre',
                padre_solicitante=padre_profile,
                
                # Datos del niño
                nombres_nino=nombres_nino,
                apellidos_nino=apellidos_nino,
                documento_nino=documento_nino,
                fecha_nacimiento_nino=fecha_nacimiento,
                genero_nino=genero,
                tipo_sangre_nino=tipo_sangre,
                observaciones_nino=observaciones_completas,
                
                # Discapacidad
                tiene_discapacidad=tiene_discapacidad,
                tipos_discapacidad=tipos_discapacidad_ids if tiene_discapacidad else None,
                otra_discapacidad=otra_discapacidad,
                
                # Documentos
                foto_nino=foto_nino,
                carnet_vacunacion_nino=carnet_vacunacion,
                certificado_eps_nino=certificado_eps,
                registro_civil_nino=registro_civil,
                
                # Datos del padre (heredados)
                tipo_documento_padre=request.user.tipo_documento,
                documento_padre=request.user.documento,
                nombres_padre=request.user.nombres,
                apellidos_padre=request.user.apellidos,
                correo_padre=request.user.correo,
                telefono_padre=request.user.telefono,
                direccion_padre=request.user.direccion,
                barrio_padre=request.user.barrio,
                nivel_educativo_padre=request.user.nivel_educativo,
                ocupacion_padre=padre_profile.ocupacion,
                departamento_padre=request.user.departamento_residencia,
                ciudad_padre=request.user.ciudad_residencia,
                localidad_bogota_padre=request.user.localidad_bogota,
            )
            
            # Marcar cupos como validados
            solicitud.cupos_validados = True
            solicitud.tiene_cupos_disponibles = True
            solicitud.save()
            
            # 🔔 CREAR NOTIFICACIÓN para la madre
            try:
                from notifications.models import Notification
                from django.contrib.contenttypes.models import ContentType
                
                titulo = f'🆕 Nueva Solicitud de Matrícula Completa'
                mensaje_notif = (
                    f'{request.user.get_full_name()} ha enviado una solicitud completa de matrícula para {nombres_nino} {apellidos_nino}. '
                    f'✅ Todos los datos y documentos están incluidos. '
                    f'Quedan {cupos_disponibles - 1} cupo(s) disponible(s).'
                )
                
                content_type = ContentType.objects.get_for_model(SolicitudMatriculacion)
                
                Notification.objects.create(
                    recipient=hogar.madre.usuario,
                    title=titulo,
                    message=mensaje_notif,
                    level='info',
                    content_type=content_type,
                    object_id=solicitud.id
                )
            except Exception as e:
                print(f"Error al crear notificación: {e}")
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': f'¡Solicitud enviada exitosamente! La madre comunitaria del hogar {hogar.nombre_hogar} revisará tu solicitud. Quedan {cupos_disponibles - 1} cupo(s) disponible(s).',
                'solicitud_id': solicitud.id
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al crear la solicitud: {str(e)}'
            }, status=500)
    
    # GET: Mostrar formulario completo
    try:
        padre_profile = request.user.padre_profile
    except Padre.DoesNotExist:
        return render(request, 'padre/solicitar_matricula.html', {
            'error': 'No se encontró tu perfil de padre. Contacta al administrador.'
        })
    
    # Obtener hogar del padre
    hogar_padre = None
    ninos_en_hogar = Nino.objects.filter(padre=padre_profile, estado='activo').select_related('hogar')
    
    if ninos_en_hogar.exists():
        hogar_padre = ninos_en_hogar.first().hogar
        hogares_disponibles = None
    else:
        hogares_disponibles = HogarComunitario.objects.filter(estado='aprobado').select_related('madre__usuario')
    
    # Obtener países (desde CHOICES del modelo Nino) y discapacidades
    paises = Nino.PAIS_NACIMIENTO_CHOICES[1:]  # Excluir la opción vacía
    discapacidades = Discapacidad.objects.all().order_by('nombre')
    
    return render(request, 'padre/solicitar_matricula.html', {
        'hogar_padre': hogar_padre,
        'hogares_disponibles': hogares_disponibles,
        'paises': paises,
        'discapacidades': discapacidades,
    })


@login_required
def enviar_invitacion_matricula(request):
    """
    2.1. Endpoint: Enviar invitación de matriculación
    URL: POST /solicitudes/enviar-invitacion/
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos para realizar esta acción.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener email del acudiente
            email_acudiente = request.POST.get('email_acudiente', '').strip()
            
            # Validar que el correo sea válido y termine en @gmail.com
            if not email_acudiente:
                return JsonResponse({'status': 'error', 'mensaje': 'El correo electrónico es obligatorio.'}, status=400)
            
            if not email_acudiente.endswith('@gmail.com'):
                return JsonResponse({'status': 'error', 'mensaje': 'Solo se permiten correos de Gmail (@gmail.com).'}, status=400)
            
            # Validar formato de email
            import re
            email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if not re.match(email_regex, email_acudiente):
                return JsonResponse({'status': 'error', 'mensaje': 'El formato del correo electrónico no es válido.'}, status=400)
            
            # VALIDACIÓN 3: Verificar que no exista una solicitud pendiente o en corrección para este correo
            solicitud_existente = SolicitudMatriculacion.objects.filter(
                email_acudiente=email_acudiente,
                hogar=hogar_madre,
                estado__in=['pendiente', 'correccion']
            ).first()
            
            if solicitud_existente:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': f'Ya existe una solicitud activa para el correo {email_acudiente}. Por favor, espera a que se complete o expire.'
                }, status=400)
            
            # Generar token único (UUID4)
            token = str(uuid.uuid4())
            
            # Generar fecha de expiración (48 horas desde ahora)
            fecha_expiracion = timezone.now() + timedelta(hours=48)
            
            # Crear la solicitud de matriculación
            solicitud = SolicitudMatriculacion.objects.create(
                hogar=hogar_madre,
                email_acudiente=email_acudiente,
                token=token,
                fecha_expiracion=fecha_expiracion,
                estado='pendiente'
            )
            
            # Construir el link de invitación
            protocolo = 'https' if request.is_secure() else 'http'
            dominio = request.get_host()
            link_invitacion = f"{protocolo}://{dominio}/matricula/publico/{token}/"
            
            # Enviar email al acudiente
            asunto = f'Invitación de Matriculación - {hogar_madre.nombre_hogar}'
            
            # Renderizar el template de email
            from django.template.loader import render_to_string
            mensaje_html = render_to_string('emails/invitacion_matricula.html', {
                'hogar': hogar_madre,
                'link': link_invitacion,
                'fecha_expiracion': fecha_expiracion,
            })
            
            # Enviar email
            from django.core.mail import send_mail
            from django.conf import settings
            
            send_mail(
                asunto,
                '',  # Mensaje de texto plano (vacío porque usamos HTML)
                settings.DEFAULT_FROM_EMAIL,
                [email_acudiente],
                fail_silently=False,
                html_message=mensaje_html
            )
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': f'Invitación enviada exitosamente a {email_acudiente}. El enlace es válido por 48 horas.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al enviar la invitación: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'}, status=405)


@login_required
def listar_solicitudes_matricula(request):
    """
    2.2. Endpoint: Cargar solicitudes pendientes para el panel de revisión
    URL: GET /solicitudes/pendientes/
    """
    try:
        # Verificar que sea madre comunitaria
        if request.user.rol.nombre_rol != 'madre_comunitaria':
            return JsonResponse({'error': 'No tienes permisos para ver las solicitudes.'}, status=403)
        
        # Obtener el hogar de la madre
        hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
        if not hogar_madre:
            return JsonResponse({'error': 'No tienes un hogar asignado.'}, status=400)
        
        # Obtener solicitudes del hogar (pendientes y en corrección)
        solicitudes = SolicitudMatriculacion.objects.filter(
            hogar=hogar_madre,
            estado__in=['pendiente', 'correccion']
        ).order_by('-fecha_creacion')
        
        # Serializar datos
        datos = []
        for s in solicitudes:
            datos.append({
                'id': s.id,
                'nombre_nino': s.nombres_nino or 'Sin completar',
                'email_acudiente': s.email_acudiente,
                'fecha_envio': s.fecha_creacion.strftime('%d/%m/%Y %H:%M') if s.fecha_creacion else '',
                'estado': s.estado,
                'tiene_datos': bool(s.nombres_nino),  # True si ya llenó el formulario
                # 🆕 Nuevos campos para solicitudes de padre
                'tipo_solicitud': s.tipo_solicitud,
                'cupos_validados': s.cupos_validados,
                'tiene_cupos_disponibles': s.tiene_cupos_disponibles,
                'padre_solicitante': s.padre_solicitante.usuario.get_full_name() if s.padre_solicitante else None,
            })
        
        return JsonResponse({'success': True, 'solicitudes': datos})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al cargar solicitudes: {str(e)}'}, status=500)


@login_required
def detalle_solicitud_matricula(request, solicitud_id):
    """
    2.3. Endpoint: Ver detalle completo (formulario enviado por el acudiente)
    URL: GET /solicitudes/<id>/detalle/
    """
    try:
        # Verificar que sea madre comunitaria
        if request.user.rol.nombre_rol != 'madre_comunitaria':
            return JsonResponse({'error': 'No tienes permisos.'}, status=403)
        
        # Obtener el hogar de la madre
        hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
        if not hogar_madre:
            return JsonResponse({'error': 'No tienes un hogar asignado.'}, status=400)
        
        # Obtener la solicitud
        solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
        
        # Construir URLs de documentos
        def get_document_url(field):
            if field and hasattr(field, 'url'):
                return request.build_absolute_uri(field.url)
            return None
        
        # Serializar datos completos
        datos = {
            'id': solicitud.id,
            'email_acudiente': solicitud.email_acudiente,
            'estado': solicitud.estado,
            'fecha_creacion': solicitud.fecha_creacion.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_creacion else '',
            'fecha_expiracion': solicitud.fecha_expiracion.strftime('%d/%m/%Y %H:%M') if solicitud.fecha_expiracion else '',
            
            # 🆕 Nuevos campos
            'tipo_solicitud': solicitud.tipo_solicitud,
            'cupos_validados': solicitud.cupos_validados,
            'tiene_cupos_disponibles': solicitud.tiene_cupos_disponibles,
            'padre_solicitante': {
                'nombres': solicitud.padre_solicitante.usuario.nombres if solicitud.padre_solicitante else '',
                'apellidos': solicitud.padre_solicitante.usuario.apellidos if solicitud.padre_solicitante else '',
                'documento': solicitud.padre_solicitante.usuario.documento if solicitud.padre_solicitante else '',
            } if solicitud.padre_solicitante else None,
            
            # Datos del niño
            'nino': {
                'nombres': solicitud.nombres_nino or '',
                'apellidos': solicitud.apellidos_nino or '',
                'documento': solicitud.documento_nino or '',
                'fecha_nacimiento': solicitud.fecha_nacimiento_nino.strftime('%Y-%m-%d') if solicitud.fecha_nacimiento_nino else '',
                'genero': solicitud.genero_nino or '',
                'tipo_sangre': solicitud.tipo_sangre_nino or '',
                'parentesco': solicitud.parentesco or '',
                'observaciones': solicitud.observaciones_nino or '',
            },
            
            # Datos del padre/acudiente
            'padre': {
                'tipo_documento': solicitud.tipo_documento_padre or '',
                'documento': solicitud.documento_padre or '',
                'nombres': solicitud.nombres_padre or '',
                'apellidos': solicitud.apellidos_padre or '',
                'telefono': solicitud.telefono_padre or '',
                'correo': solicitud.correo_padre or '',
                'direccion': solicitud.direccion_padre or '',
                'barrio': solicitud.barrio_padre or '',
                'ocupacion': solicitud.ocupacion_padre or '',
                'nivel_educativo': solicitud.nivel_educativo_padre or '',
            },
            
            # Documentos (URLs)
            'documentos': {
                'foto_nino': get_document_url(solicitud.foto_nino),
                'carnet_vacunacion_nino': get_document_url(solicitud.carnet_vacunacion_nino),
                'certificado_eps_nino': get_document_url(solicitud.certificado_eps_nino),
                'registro_civil_nino': get_document_url(solicitud.registro_civil_nino),
                'documento_identidad_padre': get_document_url(solicitud.documento_identidad_padre),
                'clasificacion_sisben_padre': get_document_url(solicitud.clasificacion_sisben_padre),
            },
            
            # Campos de corrección/rechazo
            'campos_corregir': solicitud.campos_corregir or [],
            'motivo_rechazo': solicitud.motivo_rechazo or '',
        }
        
        return JsonResponse(datos)
        
    except SolicitudMatriculacion.DoesNotExist:
        return JsonResponse({'error': 'Solicitud no encontrada.'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Error al cargar el detalle: {str(e)}'}, status=500)


@login_required
def aprobar_solicitud_matricula(request):
    """
    2.4. Endpoint: Aprobar solicitud
    URL: POST /solicitudes/<id>/aprobar/
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener ID de la solicitud
            solicitud_id = request.POST.get('solicitud_id')
            if not solicitud_id:
                return JsonResponse({'status': 'error', 'mensaje': 'ID de solicitud no proporcionado.'}, status=400)
            
            # Obtener la solicitud
            solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
            
            # Verificar que tenga datos completos
            if not solicitud.nombres_nino:
                return JsonResponse({'status': 'error', 'mensaje': 'La solicitud no tiene datos completos para aprobar.'}, status=400)
            
            # VALIDACIÓN 1: Verificar que no exista un niño con el mismo documento en este hogar
            if solicitud.documento_nino:
                nino_existente = Nino.objects.filter(
                    documento=solicitud.documento_nino,
                    hogar=hogar_madre
                ).first()
                if nino_existente:
                    return JsonResponse({
                        'status': 'error',
                        'mensaje': f'Ya existe un niño con documento {solicitud.documento_nino} matriculado en este hogar.'
                    }, status=400)
            
            # Usar transacción para crear padre y niño
            with transaction.atomic():
                # 🆕 LÓGICA DIFERENCIADA SEGÚN TIPO DE SOLICITUD
                if solicitud.tipo_solicitud == 'solicitud_padre':
                    # SOLICITUD INICIADA POR PADRE: El padre ya existe, solo crear el niño
                    if not solicitud.padre_solicitante:
                        return JsonResponse({
                            'status': 'error',
                            'mensaje': 'Error: No se encontró el padre solicitante en la solicitud.'
                        }, status=400)
                    
                    padre = solicitud.padre_solicitante
                    
                else:
                    # INVITACIÓN DE MADRE: Crear o buscar padre (lógica original)
                    # Obtener o crear el rol de padre
                    rol_padre, _ = Rol.objects.get_or_create(nombre_rol='padre')
                    
                    # VALIDACIÓN 2: Verificar si ya existe un usuario con ese documento o correo
                    usuario_padre = Usuario.objects.filter(documento=solicitud.documento_padre).first()
                    usuario_correo = Usuario.objects.filter(correo=solicitud.correo_padre).first()
                    
                    # Si existe usuario con ese correo pero diferente documento, es un error
                    if usuario_correo and usuario_correo.documento != solicitud.documento_padre:
                        return JsonResponse({
                            'status': 'error',
                            'mensaje': f'El correo {solicitud.correo_padre} ya está registrado con otro documento.'
                        }, status=400)
                    
                    if usuario_padre:
                        # Si el usuario ya existe, verificar si tiene perfil de padre
                        padre = Padre.objects.filter(usuario=usuario_padre).first()
                        if not padre:
                            # Crear perfil de padre si no existe
                            padre = Padre.objects.create(
                                usuario=usuario_padre,
                                ocupacion=solicitud.ocupacion_padre or '',
                                documento_identidad_img=solicitud.documento_identidad_padre,
                                clasificacion_sisben=solicitud.clasificacion_sisben_padre,
                            )
                    else:
                        # Crear nuevo usuario para el padre
                        usuario_padre = Usuario.objects.create(
                            documento=solicitud.documento_padre,
                            tipo_documento=solicitud.tipo_documento_padre or 'CC',
                            nombres=solicitud.nombres_padre,
                            apellidos=solicitud.apellidos_padre,
                            telefono=solicitud.telefono_padre,
                            correo=solicitud.correo_padre,
                            # 🆕 Datos geográficos
                            departamento_residencia=solicitud.departamento_padre,
                            ciudad_residencia=solicitud.ciudad_padre,
                            localidad_bogota=solicitud.localidad_bogota_padre,
                            direccion=solicitud.direccion_padre or '',
                            barrio=solicitud.barrio_padre or '',
                            nivel_educativo=solicitud.nivel_educativo_padre or '',
                            rol=rol_padre,
                            is_active=True,
                        )
                        
                        # Establecer contraseña hasheada
                        if solicitud.password_padre:
                            usuario_padre.password = solicitud.password_padre
                            usuario_padre.save()
                        
                        # Crear perfil de padre
                        padre = Padre.objects.create(
                            usuario=usuario_padre,
                            ocupacion=solicitud.ocupacion_padre or '',
                            documento_identidad_img=solicitud.documento_identidad_padre,
                            clasificacion_sisben=solicitud.clasificacion_sisben_padre,
                        )
                
                # Crear niño (mismo proceso para ambos tipos de solicitud)
                nino = Nino.objects.create(
                    padre=padre,
                    hogar=hogar_madre,
                    nombres=solicitud.nombres_nino,
                    apellidos=solicitud.apellidos_nino,
                    documento=solicitud.documento_nino,
                    fecha_nacimiento=solicitud.fecha_nacimiento_nino,
                    genero=solicitud.genero_nino,
                    tipo_sangre=solicitud.tipo_sangre_nino or '',
                    parentesco=solicitud.parentesco or '',
                    observaciones_medicas=solicitud.observaciones_nino or '',
                    foto=solicitud.foto_nino,
                    registro_civil_img=solicitud.registro_civil_nino,
                    carnet_vacunacion=solicitud.carnet_vacunacion_nino,
                    certificado_eps=solicitud.certificado_eps_nino,
                    estado='activo',
                )
                
                # Actualizar solicitud
                solicitud.estado = 'aprobado'
                solicitud.fecha_aprobacion = timezone.now()
                solicitud.save()
                
                # Marcar token como usado para prevenir reutilización
                solicitud.marcar_token_usado()
            
            # Enviar email de notificación
            try:
                from django.template.loader import render_to_string
                from django.core.mail import send_mail
                from django.conf import settings
                
                # 🆕 LÓGICA DIFERENCIADA PARA EMAILS
                if solicitud.tipo_solicitud == 'solicitud_padre':
                    # Email informativo para padre autenticado
                    asunto = f'Solicitud Aprobada - {hogar_madre.nombre_hogar}'
                    mensaje_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
                        <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                            <h2 style="color: #28A745; border-bottom: 3px solid #28A745; padding-bottom: 10px;">¡Solicitud Aprobada!</h2>
                            <p>Estimado <strong>{solicitud.padre_solicitante.usuario.get_full_name()}</strong>,</p>
                            <p>Nos complace informarte que tu solicitud de matrícula para <strong>{solicitud.nombres_nino} {solicitud.apellidos_nino}</strong> ha sido <strong style="color: #28A745;">APROBADA</strong>.</p>
                            
                            <div style="background-color: #D4EDDA; border-left: 4px solid #28A745; padding: 15px; margin: 20px 0;">
                                <h3 style="margin-top: 0; color: #155724;">Detalles de la Matrícula</h3>
                                <p style="margin: 5px 0;"><strong>Hogar Comunitario:</strong> {hogar_madre.nombre_hogar}</p>
                                <p style="margin: 5px 0;"><strong>Niño/a:</strong> {solicitud.nombres_nino} {solicitud.apellidos_nino}</p>
                                <p style="margin: 5px 0;"><strong>Documento:</strong> {solicitud.documento_nino}</p>
                            </div>
                            
                            <p><strong>Próximos pasos:</strong></p>
                            <ul>
                                <li>Ya puedes ver a tu hijo/a en tu dashboard</li>
                                <li>Podrás consultar asistencias, desarrollo y novedades</li>
                                <li>La madre comunitaria se pondrá en contacto contigo para coordinar el ingreso</li>
                            </ul>
                            
                            <p style="margin-top: 30px;">Saludos cordiales,<br><strong>{hogar_madre.nombre_hogar}</strong></p>
                        </div>
                    </body>
                    </html>
                    """
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.padre_solicitante.usuario.correo],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
                else:
                    # Email original para invitacion_madre
                    asunto = f'Solicitud de Matriculación Aprobada - {hogar_madre.nombre_hogar}'
                    mensaje_html = render_to_string('emails/solicitud_aprobada.html', {
                        'hogar': hogar_madre,
                        'nombre_nino': solicitud.nombres_nino,
                        'documento_padre': solicitud.documento_padre,
                    })
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.email_acudiente],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
            except Exception as e:
                print(f"Error al enviar email de aprobación: {e}")
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': f'Solicitud aprobada exitosamente. Se creó el registro del niño {solicitud.nombres_nino}.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al aprobar la solicitud: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'}, status=405)


@login_required
def rechazar_solicitud_matricula(request):
    """
    2.5. Endpoint: Rechazar solicitud
    URL: POST /solicitudes/<id>/rechazar/
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener datos
            solicitud_id = request.POST.get('solicitud_id')
            motivo_rechazo = request.POST.get('motivo_rechazo', '').strip()
            
            # Validaciones
            if not solicitud_id:
                return JsonResponse({'status': 'error', 'mensaje': 'ID de solicitud no proporcionado.'}, status=400)
            
            if not motivo_rechazo:
                return JsonResponse({'status': 'error', 'mensaje': 'El motivo de rechazo es obligatorio.'}, status=400)
            
            # Obtener la solicitud
            solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
            
            # Actualizar solicitud
            solicitud.estado = 'rechazado'
            solicitud.motivo_rechazo = motivo_rechazo
            solicitud.fecha_rechazo = timezone.now()
            solicitud.save()
            
            # Enviar email de notificación
            try:
                from django.template.loader import render_to_string
                from django.core.mail import send_mail
                from django.conf import settings
                
                # 🆕 LÓGICA DIFERENCIADA PARA EMAILS
                if solicitud.tipo_solicitud == 'solicitud_padre':
                    # Email informativo para padre autenticado
                    asunto = f'Solicitud Rechazada - {hogar_madre.nombre_hogar}'
                    mensaje_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
                        <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                            <h2 style="color: #E53935; border-bottom: 3px solid #E53935; padding-bottom: 10px;">Solicitud Rechazada</h2>
                            <p>Estimado <strong>{solicitud.padre_solicitante.usuario.get_full_name()}</strong>,</p>
                            <p>Lamentamos informarte que tu solicitud de matrícula para <strong>{solicitud.nombres_nino} {solicitud.apellidos_nino}</strong> ha sido rechazada.</p>
                            
                            <div style="background-color: #FFEBEE; border-left: 4px solid #E53935; padding: 15px; margin: 20px 0;">
                                <h3 style="margin-top: 0; color: #C62828;">Motivo del Rechazo</h3>
                                <p style="margin: 0;">{motivo_rechazo}</p>
                            </div>
                            
                            <p><strong>¿Qué puedes hacer?</strong></p>
                            <ul>
                                <li>Contacta con el hogar comunitario para más detalles</li>
                                <li>Puedes enviar una nueva solicitud si corriges los problemas señalados</li>
                                <li>También puedes buscar otros hogares comunitarios disponibles</li>
                            </ul>
                            
                            <p>Puedes ver el detalle completo del rechazo en tu dashboard.</p>
                            
                            <p style="margin-top: 30px;">Saludos cordiales,<br><strong>{hogar_madre.nombre_hogar}</strong></p>
                        </div>
                    </body>
                    </html>
                    """
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.padre_solicitante.usuario.correo],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
                else:
                    # Email original para invitacion_madre
                    asunto = f'Solicitud de Matriculación Rechazada - {hogar_madre.nombre_hogar}'
                    mensaje_html = render_to_string('emails/solicitud_rechazada.html', {
                        'hogar': hogar_madre,
                        'motivo': motivo_rechazo,
                    })
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.email_acudiente],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
            except Exception as e:
                print(f"Error al enviar email de rechazo: {e}")
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': 'Solicitud rechazada. Se notificó al acudiente.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al rechazar la solicitud: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'}, status=405)


@login_required
def devolver_correccion_matricula(request):
    """
    2.6. Endpoint: Devolver para corrección
    URL: POST /solicitudes/<id>/correccion/
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener datos
            solicitud_id = request.POST.get('solicitud_id')
            campos_json = request.POST.get('campos_corregir', '[]')
            
            # Validaciones
            if not solicitud_id:
                return JsonResponse({'status': 'error', 'mensaje': 'ID de solicitud no proporcionado.'}, status=400)
            
            # Parsear campos a corregir
            try:
                campos_corregir = json.loads(campos_json)
            except json.JSONDecodeError:
                return JsonResponse({'status': 'error', 'mensaje': 'Formato de campos inválido.'}, status=400)
            
            if not campos_corregir:
                return JsonResponse({'status': 'error', 'mensaje': 'Debe seleccionar al menos un campo para corregir.'}, status=400)
            
            # Obtener la solicitud
            solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
            
            # Validar que no haya excedido los intentos
            if solicitud.intentos_correccion >= 3:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': 'Esta solicitud ya excedió el límite de 3 intentos de corrección. El acudiente debe enviar una nueva solicitud.'
                }, status=400)
            
            # Incrementar contador de intentos
            solicitud.intentos_correccion += 1
            intentos_restantes = 3 - solicitud.intentos_correccion
            
            # Actualizar solicitud
            solicitud.estado = 'correccion'
            solicitud.campos_corregir = campos_corregir
            solicitud.save()
            
            # 🆕 LÓGICA DIFERENCIADA SEGÚN TIPO DE SOLICITUD
            if solicitud.tipo_solicitud == 'solicitud_padre':
                # SOLICITUD DE PADRE: Solo notificación in-app + email informativo (sin formulario)
                try:
                    from notifications.models import Notification
                    from django.contrib.contenttypes.models import ContentType
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    # Crear notificación in-app
                    content_type = ContentType.objects.get_for_model(SolicitudMatriculacion)
                    Notification.objects.create(
                        recipient=solicitud.padre_solicitante.usuario,
                        title=f"Correcciones solicitadas: {solicitud.nombres_nino or 'Solicitud de matrícula'}",
                        message=f"La madre comunitaria ha solicitado correcciones en {len(campos_corregir)} campo(s). Revisa tu dashboard para realizar los cambios.",
                        level='warning',
                        content_type=content_type,
                        object_id=solicitud.id
                    )
                    
                    # Enviar email INFORMATIVO (sin link a formulario, solo aviso)
                    campos_nombres = {
                        'foto_nino': 'Foto del niño',
                        'carnet_vacunacion_nino': 'Carnet de vacunación',
                        'certificado_eps_nino': 'Certificado EPS',
                        'registro_civil_nino': 'Registro civil',
                        'nombres_nino': 'Nombres del niño',
                        'apellidos_nino': 'Apellidos del niño',
                        'documento_nino': 'Documento del niño',
                        'fecha_nacimiento_nino': 'Fecha de nacimiento',
                        'genero_nino': 'Género',
                        'tipo_sangre_nino': 'Tipo de sangre',
                        'parentesco': 'Parentesco',
                        'observaciones_nino': 'Observaciones médicas'
                    }
                    
                    campos_legibles = [campos_nombres.get(campo, campo) for campo in campos_corregir]
                    
                    asunto = f'Correcciones Solicitadas - {hogar_madre.nombre_hogar}'
                    mensaje_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f5f5f5;">
                        <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                            <h2 style="color: #FF6B35; border-bottom: 3px solid #FF6B35; padding-bottom: 10px;">Correcciones Solicitadas</h2>
                            <p>Estimado <strong>{solicitud.padre_solicitante.usuario.get_full_name()}</strong>,</p>
                            <p>La madre comunitaria de <strong>{hogar_madre.nombre_hogar}</strong> ha revisado su solicitud de matrícula para <strong>{solicitud.nombres_nino or 'su hijo/a'}</strong> y requiere que corrija la siguiente información:</p>
                            
                            <div style="background-color: #FFF3E0; border-left: 4px solid #FF6B35; padding: 15px; margin: 20px 0;">
                                <h3 style="margin-top: 0; color: #E65100;">Campos a corregir:</h3>
                                <ul style="margin-bottom: 0;">
                                    {''.join([f'<li><strong>{campo}</strong></li>' for campo in campos_legibles])}
                                </ul>
                            </div>
                            
                            <div style="background-color: #E3F2FD; border-left: 4px solid #2196F3; padding: 15px; margin: 20px 0;">
                                <p style="margin: 0;"><strong>📝 ¿Cómo corregir?</strong></p>
                                <p style="margin: 10px 0 0 0;">Ingresa a tu <strong>dashboard</strong> en la plataforma ICBF Conecta. Allí encontrarás tu solicitud pendiente con los campos que debes actualizar.</p>
                            </div>
                            
                            <p style="color: #666; font-size: 14px; margin-top: 30px; border-top: 1px solid #ddd; padding-top: 20px;">
                                Intentos restantes: <strong>{intentos_restantes} de 3</strong><br>
                                Si excede el límite de intentos, deberá enviar una nueva solicitud.
                            </p>
                            
                            <p style="margin-top: 30px;">Saludos cordiales,<br><strong>{hogar_madre.nombre_hogar}</strong></p>
                        </div>
                    </body>
                    </html>
                    """
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.padre_solicitante.usuario.correo],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
                    
                except Exception as e:
                    print(f"Error al enviar notificación para solicitud de padre: {e}")
                    import traceback
                    traceback.print_exc()
                
            else:
                # INVITACIÓN DE MADRE: Email con link al formulario público (flujo original)
                try:
                    from django.template.loader import render_to_string
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    asunto = f'Solicitud de Matriculación - Correcciones Requeridas - {hogar_madre.nombre_hogar}'
                    
                    # Generar link para volver al formulario
                    protocolo = 'https' if request.is_secure() else 'http'
                    dominio = request.get_host()
                    link_formulario = f"{protocolo}://{dominio}/matricula/publico/{solicitud.token}/"
                    
                    mensaje_html = render_to_string('emails/solicitud_correccion.html', {
                        'hogar': hogar_madre,
                        'campos': campos_corregir,
                        'link': link_formulario,
                        'intentos_usados': solicitud.intentos_correccion,
                        'intentos_restantes': intentos_restantes,
                    })
                    
                    send_mail(
                        asunto,
                        '',
                        settings.DEFAULT_FROM_EMAIL,
                        [solicitud.email_acudiente],
                        fail_silently=True,
                        html_message=mensaje_html
                    )
                except Exception as e:
                    print(f"Error al enviar email de corrección: {e}")
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': f'Se solicitó corrección de {len(campos_corregir)} campo(s). El acudiente fue notificado.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al solicitar correcciones: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'}, status=405)


def cancelar_solicitud_usuario(request, token):
    """
    Endpoint: Permite al acudiente cancelar su propia solicitud
    URL: POST /matricula/publico/<token>/cancelar/
    """
    if request.method == 'POST':
        try:
            # Buscar la solicitud por token
            try:
                solicitud = SolicitudMatriculacion.objects.get(token=token)
            except SolicitudMatriculacion.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Solicitud no encontrada.'
                }, status=404)
            
            # Verificar que la solicitud puede ser cancelada
            if solicitud.estado not in ['pendiente', 'correccion']:
                return JsonResponse({
                    'success': False,
                    'error': f'No se puede cancelar una solicitud en estado "{solicitud.get_estado_display()}".'
                }, status=400)
            
            # Obtener motivo opcional
            motivo = request.POST.get('motivo', '').strip()
            
            # Cancelar la solicitud
            if solicitud.cancelar_por_usuario(motivo):
                # Notificar a la madre comunitaria
                try:
                    from notifications.models import Notification
                    
                    Notification.objects.create(
                        usuario=solicitud.hogar.madre_comunitaria,
                        solicitud=solicitud,
                        tipo_notificacion='solicitud_cancelada',
                        mensaje=f'El acudiente {solicitud.email_acudiente} canceló su solicitud de matrícula.',
                        nivel='info'
                    )
                except Exception as e:
                    print(f"Error al crear notificación: {e}")
                
                return JsonResponse({
                    'success': True,
                    'mensaje': 'Su solicitud ha sido cancelada exitosamente.'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'No se pudo cancelar la solicitud.'
                }, status=400)
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'Error al cancelar la solicitud: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


def eliminar_solicitud_matricula(request):
    """
    2.8. Endpoint: Eliminar solicitud de matriculación
    URL: POST /solicitudes/eliminar/
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener ID de la solicitud
            solicitud_id = request.POST.get('solicitud_id')
            
            if not solicitud_id:
                return JsonResponse({'status': 'error', 'mensaje': 'ID de solicitud no proporcionado.'}, status=400)
            
            # Obtener la solicitud y verificar que pertenezca al hogar
            solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
            
            # Guardar datos para el mensaje
            email_acudiente = solicitud.email_acudiente
            nombre_nino = f"{solicitud.nombres_nino or ''} {solicitud.apellidos_nino or ''}".strip() or "Sin nombre"
            
            # Eliminar las notificaciones relacionadas con esta solicitud
            from django.contrib.contenttypes.models import ContentType
            from notifications.models import Notification
            
            content_type = ContentType.objects.get_for_model(SolicitudMatriculacion)
            Notification.objects.filter(content_type=content_type, object_id=solicitud_id).delete()
            
            # Eliminar la solicitud (esto también eliminará los archivos asociados)
            solicitud.delete()
            
            # Opcional: Enviar email de notificación al acudiente
            try:
                from django.template.loader import render_to_string
                from django.core.mail import send_mail
                from django.conf import settings
                
                asunto = f'Solicitud de Matriculación Cancelada - {hogar_madre.nombre_hogar}'
                
                mensaje_html = f"""
                <html>
                <body style="font-family: Arial, sans-serif; padding: 20px;">
                    <h2 style="color: #dc3545;">Solicitud Cancelada</h2>
                    <p>Estimado acudiente,</p>
                    <p>Le informamos que la solicitud de matriculación para <strong>{nombre_nino}</strong> ha sido cancelada por el hogar comunitario <strong>{hogar_madre.nombre_hogar}</strong>.</p>
                    <p>Si desea volver a solicitar la matriculación, puede contactar directamente con el hogar comunitario.</p>
                    <br>
                    <p>Saludos cordiales,<br><strong>{hogar_madre.nombre_hogar}</strong></p>
                </body>
                </html>
                """
                
                send_mail(
                    asunto,
                    '',
                    settings.DEFAULT_FROM_EMAIL,
                    [email_acudiente],
                    fail_silently=True,
                    html_message=mensaje_html
                )
            except Exception as e:
                print(f"Error al enviar email de cancelación: {e}")
            
            return JsonResponse({
                'status': 'ok',
                'mensaje': f'La solicitud de {nombre_nino} ha sido eliminada exitosamente.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'mensaje': f'Error al eliminar la solicitud: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'mensaje': 'Método no permitido.'}, status=405)


def historial_solicitud(request, solicitud_id):
    """
    2.9. Endpoint: Obtener historial de cambios de una solicitud
    URL: GET /solicitudes/<id>/historial/
    """
    try:
        from core.models import HistorialCambio
        
        # Verificar que sea madre comunitaria
        if request.user.rol.nombre_rol != 'madre_comunitaria':
            return JsonResponse({'status': 'error', 'mensaje': 'No tienes permisos.'}, status=403)
        
        # Obtener el hogar de la madre
        hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
        if not hogar_madre:
            return JsonResponse({'status': 'error', 'mensaje': 'No tienes un hogar asignado.'}, status=400)
        
        # Verificar que la solicitud pertenezca al hogar
        solicitud = get_object_or_404(SolicitudMatriculacion, id=solicitud_id, hogar=hogar_madre)
        
        # Obtener historial de cambios
        cambios = HistorialCambio.objects.filter(
            tipo_modelo='solicitud',
            objeto_id=solicitud_id
        ).order_by('-fecha_cambio')[:50]  # Últimos 50 cambios
        
        historial = []
        for cambio in cambios:
            historial.append({
                'campo': cambio.campo_modificado,
                'valor_anterior': cambio.valor_anterior,
                'valor_nuevo': cambio.valor_nuevo,
                'accion': cambio.accion,
                'fecha': cambio.fecha_cambio.strftime('%d/%m/%Y %H:%M'),
                'usuario': cambio.usuario.get_full_name() if cambio.usuario else 'Sistema',
                'observaciones': cambio.observaciones or ''
            })
        
        return JsonResponse({
            'status': 'ok',
            'historial': historial,
            'total': len(historial)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'mensaje': f'Error al obtener historial: {str(e)}'
        }, status=500)


def formulario_matricula_publico(request, token):
    """
    2.7. Endpoint: Procesar formulario público del acudiente
    URL: GET/POST /matricula/publico/<token>/
    """
    # Buscar la solicitud por token
    try:
        solicitud = SolicitudMatriculacion.objects.get(token=token)
    except SolicitudMatriculacion.DoesNotExist:
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'invalido',
            'mensaje_error': 'El enlace de invitación no es válido.',
        })
    
    # VALIDACIÓN MEJORADA DEL ESTADO DE LA SOLICITUD
    
    # 1. Si ya fue aprobada o token usado → Bloquear completamente
    if solicitud.estado in ['aprobado', 'token_usado']:
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'aprobado',
            'mensaje_error': '¡Felicidades! Esta solicitud ya fue aprobada y el niño está matriculado.',
            'solicitud': solicitud,
        })
    
    # 2. Si fue rechazada → Bloquear completamente
    if solicitud.estado == 'rechazado':
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'rechazado',
            'mensaje_error': 'Esta solicitud fue rechazada por el hogar comunitario.',
            'motivo_rechazo': solicitud.motivo_rechazo,
            'solicitud': solicitud,
        })
    
    # 3. Si fue cancelada por expiración → Bloquear
    if solicitud.estado == 'cancelado_expiracion':
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'cancelado_expiracion',
            'mensaje_error': 'Esta solicitud fue cancelada por expiración del plazo.',
            'motivo_cancelacion': solicitud.motivo_cancelacion,
            'solicitud': solicitud,
        })
    
    # 4. Si fue cancelada por usuario → Bloquear
    if solicitud.estado == 'cancelado_usuario':
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'cancelado_usuario',
            'mensaje_error': 'Esta solicitud fue cancelada.',
            'motivo_cancelacion': solicitud.motivo_cancelacion,
            'solicitud': solicitud,
        })
    
    # 5. Si expiró el token y no está en corrección → Cancelar automáticamente
    if solicitud.fecha_expiracion < timezone.now() and solicitud.estado != 'correccion':
        # Cancelar automáticamente por expiración
        solicitud.cancelar_por_expiracion()
        
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'expirado',
            'mensaje_error': 'El enlace ha expirado.',
            'solicitud': solicitud,
            'puede_renovar': True,
        })
    
    # 6. Verificar límite de intentos en corrección
    if solicitud.estado == 'correccion' and solicitud.intentos_correccion >= 3:
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': False,
            'estado_solicitud': 'limite_excedido',
            'mensaje_error': 'Ha excedido el límite de 3 intentos de corrección. Por favor, contacte al hogar comunitario.',
            'solicitud': solicitud,
        })
    
    # GET: Mostrar formulario
    if request.method == 'GET':
        # Obtener todas las discapacidades disponibles
        discapacidades = Discapacidad.objects.all().order_by('nombre')
        
        # 🆕 Obtener departamentos para el select
        from core.models import Departamento
        departamentos = Departamento.objects.all().order_by('nombre')
        
        # 🆕 Detectar si es solicitud iniciada por padre
        es_solicitud_padre = solicitud.tipo_solicitud == 'solicitud_padre'
        
        return render(request, 'public/formulario_matricula_publico.html', {
            'token_valido': True,
            'solicitud': solicitud,
            'en_correccion': solicitud.estado == 'correccion',
            'campos_corregir': solicitud.campos_corregir or [],
            'intentos_restantes': 3 - solicitud.intentos_correccion if solicitud.estado == 'correccion' else 3,
            'discapacidades': discapacidades,
            'departamentos': departamentos,
            # 🆕 Nuevos contextos
            'es_solicitud_padre': es_solicitud_padre,
            'mostrar_solo_nino': es_solicitud_padre,  # Si es solicitud de padre, solo mostrar campos del niño
        })
    
    # POST: Procesar formulario
    if request.method == 'POST':
        try:
            # Validar token nuevamente
            if not solicitud.is_valido():
                return JsonResponse({
                    'success': False,
                    'error': 'El enlace ha expirado.'
                }, status=400)
            
            # Datos del niño
            solicitud.nombres_nino = request.POST.get('nombres_nino', '').strip()
            solicitud.apellidos_nino = request.POST.get('apellidos_nino', '').strip()
            solicitud.documento_nino = request.POST.get('documento_nino', '').strip()
            
            # VALIDACIÓN 4: Verificar que no exista un niño ya matriculado con ese documento en el hogar
            if solicitud.documento_nino:
                nino_ya_existe = Nino.objects.filter(
                    documento=solicitud.documento_nino,
                    hogar=solicitud.hogar
                ).exists()
                
                if nino_ya_existe:
                    return JsonResponse({
                        'success': False,
                        'error': f'Ya existe un niño matriculado con el documento {solicitud.documento_nino} en este hogar.'
                    }, status=400)
            
            fecha_nac = request.POST.get('fecha_nacimiento_nino')
            if fecha_nac:
                solicitud.fecha_nacimiento_nino = datetime.strptime(fecha_nac, '%Y-%m-%d').date()
            solicitud.genero_nino = request.POST.get('genero_nino', '')
            solicitud.tipo_sangre_nino = request.POST.get('tipo_sangre_nino', '')
            solicitud.parentesco = request.POST.get('parentesco', '')
            solicitud.observaciones_nino = request.POST.get('observaciones_nino', '')
            
            # Discapacidad
            tiene_discapacidad = request.POST.get('tiene_discapacidad', 'false')
            solicitud.tiene_discapacidad = tiene_discapacidad == 'true'
            
            if solicitud.tiene_discapacidad:
                # Obtener IDs de tipos de discapacidad seleccionados
                tipos_ids = request.POST.getlist('tipos_discapacidad')
                solicitud.tipos_discapacidad = [int(id) for id in tipos_ids if id.isdigit()]
                solicitud.otra_discapacidad = request.POST.get('otra_discapacidad', '').strip()
            else:
                solicitud.tipos_discapacidad = None
                solicitud.otra_discapacidad = None
            
            # 🆕 LÓGICA DIFERENCIADA: Datos del padre
            # Si es solicitud de padre, los datos del padre ya están pre-llenados y NO se editan
            es_solicitud_padre = solicitud.tipo_solicitud == 'solicitud_padre'
            
            if not es_solicitud_padre:
                # INVITACIÓN DE MADRE: Procesar datos del padre normalmente
                solicitud.tipo_documento_padre = request.POST.get('tipo_documento_padre', '')
                solicitud.documento_padre = request.POST.get('documento_padre', '').strip()
                solicitud.nombres_padre = request.POST.get('nombres_padre', '').strip()
                solicitud.apellidos_padre = request.POST.get('apellidos_padre', '').strip()
                solicitud.telefono_padre = request.POST.get('telefono_padre', '')
                solicitud.correo_padre = request.POST.get('correo_padre', '').strip()
            
                # VALIDACIÓN 5: Verificar que el correo del padre no esté ya registrado con otro documento
                if solicitud.correo_padre:
                    usuario_con_correo = Usuario.objects.filter(correo=solicitud.correo_padre).first()
                    if usuario_con_correo and usuario_con_correo.documento != solicitud.documento_padre:
                        return JsonResponse({
                            'success': False,
                            'error': f'El correo {solicitud.correo_padre} ya está registrado con otro documento. Si ya tienes cuenta, usa el mismo documento.'
                        }, status=400)
                
                # 🆕 Datos geográficos del padre
                departamento_id = request.POST.get('departamento_padre', '').strip()
                ciudad_id = request.POST.get('ciudad_padre', '').strip()
                localidad_id = request.POST.get('localidad_bogota_padre', '').strip()
                
                if departamento_id:
                    from core.models import Departamento
                    solicitud.departamento_padre_id = departamento_id
                if ciudad_id:
                    from core.models import Municipio
                    solicitud.ciudad_padre_id = ciudad_id
                if localidad_id:
                    from core.models import LocalidadBogota
                    solicitud.localidad_bogota_padre_id = localidad_id
                
                solicitud.direccion_padre = request.POST.get('direccion_padre', '')
                solicitud.barrio_padre = request.POST.get('barrio_padre', '')
                solicitud.ocupacion_padre = request.POST.get('ocupacion_padre', '')
                solicitud.nivel_educativo_padre = request.POST.get('nivel_educativo_padre', '')
                
                # Contraseña (hashear) - Solo para invitaciones de madre
                password = request.POST.get('password_padre', '').strip()
                password_confirm = request.POST.get('password_confirm', '').strip()
                
                # La contraseña es requerida siempre (no se almacena hasta aprobar)
                if not password:
                    return JsonResponse({
                        'success': False,
                        'error': 'La contraseña es obligatoria.'
                    }, status=400)
                
                if password != password_confirm:
                    return JsonResponse({
                        'success': False,
                        'error': 'Las contraseñas no coinciden.'
                    }, status=400)
                
                if len(password) < 8:
                    return JsonResponse({
                        'success': False,
                        'error': 'La contraseña debe tener al menos 8 caracteres.'
                    }, status=400)
                
                solicitud.password_padre = make_password(password)
            else:
                # SOLICITUD DE PADRE: Los datos del padre ya están pre-llenados, no se procesan
                # Solo se permite actualizar documentos si se están corrigiendo
                pass
            
            # Documentos (archivos)
            if 'foto_nino' in request.FILES:
                solicitud.foto_nino = request.FILES['foto_nino']
            
            if 'carnet_vacunacion_nino' in request.FILES:
                solicitud.carnet_vacunacion_nino = request.FILES['carnet_vacunacion_nino']
            
            if 'certificado_eps_nino' in request.FILES:
                solicitud.certificado_eps_nino = request.FILES['certificado_eps_nino']
            
            if 'registro_civil_nino' in request.FILES:
                solicitud.registro_civil_nino = request.FILES['registro_civil_nino']
            
            if 'documento_identidad_padre' in request.FILES:
                solicitud.documento_identidad_padre = request.FILES['documento_identidad_padre']
            
            if 'clasificacion_sisben_padre' in request.FILES:
                solicitud.clasificacion_sisben_padre = request.FILES['clasificacion_sisben_padre']
            
            # Validar que se cargaron archivos nuevos para campos marcados en corrección
            if solicitud.estado == 'correccion' and solicitud.campos_corregir:
                campos_archivos = [
                    'foto_nino', 'carnet_vacunacion_nino', 'certificado_eps_nino',
                    'registro_civil_nino', 'documento_identidad_padre', 'clasificacion_sisben_padre'
                ]
                
                for campo in solicitud.campos_corregir:
                    if campo in campos_archivos:
                        # Verificar que se cargó un archivo nuevo
                        if campo not in request.FILES:
                            nombres_legibles = {
                                'foto_nino': 'Foto del Niño',
                                'carnet_vacunacion_nino': 'Carnet de Vacunación',
                                'certificado_eps_nino': 'Certificado EPS',
                                'registro_civil_nino': 'Registro Civil',
                                'documento_identidad_padre': 'Documento de Identidad del Acudiente',
                                'clasificacion_sisben_padre': 'Clasificación SISBEN'
                            }
                            return JsonResponse({
                                'success': False,
                                'error': f'Debe cargar un archivo nuevo para: {nombres_legibles.get(campo, campo)}'
                            }, status=400)
            
            # Validar documentos obligatorios (solo si no es corrección o si es un documento a corregir)
            # 🆕 Para solicitudes de padre, solo validar documentos del niño
            if es_solicitud_padre:
                documentos_obligatorios = [
                    ('foto_nino', 'Foto del niño'),
                    ('carnet_vacunacion_nino', 'Carnet de vacunación'),
                    ('certificado_eps_nino', 'Certificado EPS'),
                    ('registro_civil_nino', 'Registro civil'),
                ]
            else:
                documentos_obligatorios = [
                    ('foto_nino', 'Foto del niño'),
                    ('carnet_vacunacion_nino', 'Carnet de vacunación'),
                    ('certificado_eps_nino', 'Certificado EPS'),
                    ('registro_civil_nino', 'Registro civil'),
                    ('documento_identidad_padre', 'Documento de identidad del acudiente'),
                    ('clasificacion_sisben_padre', 'Clasificación SISBEN'),
                ]
            
            for campo, nombre in documentos_obligatorios:
                if not getattr(solicitud, campo):
                    return JsonResponse({
                        'success': False,
                        'error': f'El documento "{nombre}" es obligatorio.'
                    }, status=400)
            
            # Cambiar estado a pendiente (si estaba en corrección)
            es_actualizacion = solicitud.estado == 'correccion'
            if es_actualizacion:
                solicitud.estado = 'pendiente'
                solicitud.campos_corregir = None
            
            # Guardar
            solicitud.save()
            
            # Crear notificación para la madre comunitaria
            try:
                from notifications.models import Notification
                from django.contrib.contenttypes.models import ContentType
                
                # Buscar el usuario de la madre comunitaria asociada al hogar
                madre_usuario = None
                if hasattr(solicitud.hogar, 'madre') and solicitud.hogar.madre:
                    madre_usuario = solicitud.hogar.madre.usuario
                
                if madre_usuario:
                    # Crear notificación
                    if es_actualizacion:
                        titulo = f"Solicitud Actualizada: {solicitud.nombres_nino} {solicitud.apellidos_nino}"
                        mensaje = f"El acudiente ha actualizado la solicitud de matrícula según las correcciones solicitadas."
                    else:
                        titulo = f"Nueva Solicitud de Matrícula: {solicitud.nombres_nino} {solicitud.apellidos_nino}"
                        mensaje = f"Se ha recibido una nueva solicitud de matrícula. Revisa los documentos y datos proporcionados."
                    
                    content_type = ContentType.objects.get_for_model(SolicitudMatriculacion)
                    
                    Notification.objects.create(
                        recipient=madre_usuario,
                        title=titulo,
                        message=mensaje,
                        level='info',
                        content_type=content_type,
                        object_id=solicitud.id
                    )
                    print(f"✓ Notificación creada para {madre_usuario.get_full_name()}")
                else:
                    print(f"⚠ No se encontró usuario de madre para hogar {solicitud.hogar.nombre_hogar}")
            except Exception as e:
                print(f"❌ Error al crear notificación: {e}")
                import traceback
                traceback.print_exc()
            
            # Enviar email de confirmación
            try:
                from django.template.loader import render_to_string
                from django.core.mail import send_mail
                from django.conf import settings
                
                asunto = f'Solicitud de Matriculación Recibida - {solicitud.hogar.nombre_hogar}'
                mensaje_html = render_to_string('emails/solicitud_confirmacion.html', {
                    'hogar': solicitud.hogar,
                    'nombre_nino': solicitud.nombres_nino,
                })
                
                send_mail(
                    asunto,
                    '',
                    settings.DEFAULT_FROM_EMAIL,
                    [solicitud.email_acudiente],
                    fail_silently=True,
                    html_message=mensaje_html
                )
            except Exception as e:
                print(f"Error al enviar email de confirmación: {e}")
            
            return JsonResponse({
                'success': True,
                'mensaje': 'Solicitud enviada exitosamente. Será revisada por el hogar comunitario.'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'Error al procesar el formulario: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)


# =====================================================
# 👨‍👧 VISTA DE ASISTENCIA PARA PADRES
# =====================================================
@login_required
def padre_historial_asistencia(request, nino_id):
    """
    Vista para que los padres vean el historial de asistencia de sus hijos.
    Solo pueden ver la asistencia de sus propios hijos.
    """
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')
    
    try:
        # Verificar que el niño pertenece al padre autenticado
        padre = Padre.objects.get(usuario=request.user)
        nino = get_object_or_404(Nino, id=nino_id, padre=padre)
        
        # Obtener historial de asistencia
        historial = Asistencia.objects.filter(nino=nino).order_by('-fecha')
        
        # Filtro por rango de fechas
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            historial = historial.filter(fecha__range=[start_date, end_date])
        
        # Vincular novedad por fecha y niño
        for asistencia in historial:
            novedad = Novedad.objects.filter(nino=nino, fecha=asistencia.fecha).first()
            asistencia.novedad_id = novedad.id if novedad else None
        
        # Datos para calendario
        eventos = [
            {
                "title": a.estado,
                "start": a.fecha.strftime("%Y-%m-%d"),
                "color": (
                    "#28a745" if a.estado == "Presente" else
                    "#dc3545" if a.estado == "Ausente" else
                    "#6f42c1"
                )
            }
            for a in historial
        ]
        
        # Calcular estadísticas
        total = historial.count()
        presentes = historial.filter(estado="Presente").count()
        ausentes = historial.filter(estado="Ausente").count()
        justificados = historial.filter(estado="Justificado").count()
        
        def porcentaje(valor):
            return round((valor / total) * 100) if total > 0 else 0
        
        # Calcular si hay ausencias críticas
        ausencias_sin_justificar = historial.filter(estado="Ausente").count()
        tiene_alerta = ausencias_sin_justificar >= 3
        
        alerta_ausencias = {
            'tiene_alerta': tiene_alerta,
            'ausencias_sin_justificar': ausencias_sin_justificar,
            'nivel': 'grave' if ausencias_sin_justificar >= 5 else 'warning' if ausencias_sin_justificar >= 3 else 'info',
            'porcentaje_ausencias': porcentaje(ausentes),
        }
        
        return render(request, 'padre/asistencia_historial.html', {
            'nino': nino,
            'historial': historial,
            'presentes': presentes,
            'ausentes': ausentes,
            'justificados': justificados,
            'porc_presentes': porcentaje(presentes),
            'porc_ausentes': porcentaje(ausentes),
            'porc_justificados': porcentaje(justificados),
            'eventos_json': json.dumps(eventos, cls=json.JSONEncoder),
            'start_date': start_date,
            'end_date': end_date,
            'alerta_ausencias': alerta_ausencias,
        })
    except Padre.DoesNotExist:
        return redirect('padre_dashboard')
    except Nino.DoesNotExist:
        return redirect('padre_dashboard')


# ==========================================
# VISTAS DE ERROR PERSONALIZADAS
# ==========================================

def custom_404(request, exception=None):
    """
    Vista personalizada para errores 404 (Página No Encontrada)
    """
    from django.conf import settings
    
    context = {
        'request_path': request.path,
        'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
        'exception': str(exception) if exception else None,
        'settings': settings,
    }
    
    return render(request, 'errors/404.html', context, status=404)


def custom_500(request):
    """
    Vista personalizada para errores 500 (Error del Servidor)
    """
    from django.conf import settings
    import sys
    
    # Obtener información del último error si está disponible
    exc_type, exc_value, exc_traceback = sys.exc_info()
    
    context = {
        'request_path': request.path if hasattr(request, 'path') else 'Desconocido',
        'timestamp': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
        'exception': str(exc_value) if exc_value else 'Error interno del servidor',
        'settings': settings,
    }
    
    return render(request, 'errors/500.html', context, status=500)


# ============================================================================
# 🆕 VISTAS AJAX PARA GEOGRAFÍA DE COLOMBIA
# ============================================================================

def ajax_cargar_municipios(request):
    """
    Vista AJAX para cargar municipios según el departamento seleccionado
    GET: /ajax/cargar-municipios/?departamento_id=X
    Returns: JSON con lista de municipios [{id, nombre, es_capital}, ...]
    """
    from core.models import Municipio
    
    departamento_id = request.GET.get('departamento_id')
    
    if not departamento_id:
        return JsonResponse({'error': 'Departamento no especificado'}, status=400)
    
    municipios = Municipio.objects.filter(departamento_id=departamento_id).order_by('nombre')
    
    data = [
        {
            'id': mun.id,
            'nombre': mun.nombre,
            'es_capital': mun.es_capital,
        }
        for mun in municipios
    ]
    
    return JsonResponse(data, safe=False)


def ajax_cargar_localidades_bogota(request):
    """
    Vista AJAX para cargar localidades de Bogotá si el municipio seleccionado es Bogotá D.C.
    GET: /ajax/cargar-localidades-bogota/?municipio_id=X
    Returns: JSON con lista de localidades [{id, numero, nombre}, ...]
    """
    from core.models import Municipio, LocalidadBogota
    
    municipio_id = request.GET.get('municipio_id')
    
    if not municipio_id:
        return JsonResponse({'error': 'Municipio no especificado'}, status=400)
    
    try:
        municipio = Municipio.objects.get(id=municipio_id)
        
        # Solo devolver localidades si es Bogotá D.C.
        if municipio.nombre == 'Bogotá D.C.' or municipio.departamento.nombre == 'Bogotá D.C.':
            localidades = LocalidadBogota.objects.all().order_by('numero')
            
            data = [
                {
                    'id': loc.id,
                    'numero': loc.numero,
                    'nombre': loc.nombre,
                }
                for loc in localidades
            ]
            
            return JsonResponse({'es_bogota': True, 'localidades': data}, safe=False)
        else:
            return JsonResponse({'es_bogota': False, 'localidades': []}, safe=False)
            
    except Municipio.DoesNotExist:
        return JsonResponse({'error': 'Municipio no encontrado'}, status=404)


# ========================================================================================
# 🔍 VALIDACIONES AJAX EN TIEMPO REAL
# ========================================================================================

def validar_nombre_hogar(request):
    """
    Vista AJAX para validar si un nombre de hogar ya existe en la base de datos.
    GET: /ajax/validar-nombre-hogar/?nombre=Hogar Los Angeles&hogar_id=5 (opcional)
    Returns: JSON con {existe: true/false, mensaje: "..."}
    """
    from core.models import HogarComunitario
    
    nombre = request.GET.get('nombre', '').strip()
    hogar_id = request.GET.get('hogar_id', None)  # Para excluir el hogar actual en modo edición
    
    if not nombre:
        return JsonResponse({'existe': False, 'mensaje': ''})
    
    # Buscar hogares con el mismo nombre (case-insensitive)
    hogares = HogarComunitario.objects.filter(nombre_hogar__iexact=nombre)
    
    # Si estamos editando, excluir el hogar actual
    if hogar_id:
        hogares = hogares.exclude(id=hogar_id)
    
    if hogares.exists():
        hogar_existente = hogares.first()
        mensaje = f'Ya existe un hogar con el nombre "{nombre}" en {hogar_existente.ciudad.nombre}'
        return JsonResponse({
            'existe': True, 
            'mensaje': mensaje,
            'hogar_id': hogar_existente.id
        })
    else:
        return JsonResponse({
            'existe': False, 
            'mensaje': 'Nombre disponible'
        })


def validar_documento_madre(request):
    """
    Vista AJAX para validar si un documento de agente educativo ya existe.
    GET: /ajax/validar-documento-madre/?documento=123456789&usuario_id=5 (opcional)
    Returns: JSON con {existe: true/false, mensaje: "..."}
    """
    from core.models import Usuario
    
    documento = request.GET.get('documento', '').strip()
    usuario_id = request.GET.get('usuario_id', None)  # Para excluir el usuario actual en modo edición
    
    if not documento:
        return JsonResponse({'existe': False, 'mensaje': ''})
    
    # Buscar usuarios con el mismo documento
    usuarios = Usuario.objects.filter(documento=documento)
    
    # Si estamos editando, excluir el usuario actual
    if usuario_id:
        usuarios = usuarios.exclude(id=usuario_id)
    
    if usuarios.exists():
        usuario_existente = usuarios.first()
        mensaje = f'El documento {documento} ya está registrado para {usuario_existente.nombres} {usuario_existente.apellidos}'
        return JsonResponse({
            'existe': True, 
            'mensaje': mensaje,
            'usuario_id': usuario_existente.id,
            'nombre_completo': f'{usuario_existente.nombres} {usuario_existente.apellidos}'
        })
    else:
        return JsonResponse({
            'existe': False, 
            'mensaje': 'Documento disponible'
        })


# ========================================================================================
# 🏠 SISTEMA DE VISITAS TÉCNICAS PARA HABILITACIÓN DE HOGARES
# ========================================================================================

@login_required
def listar_hogares_pendientes_visita(request):
    """
    Lista todos los hogares que están pendientes de visita técnica o con visita agendada.
    """
    hogares = HogarComunitario.objects.filter(
        estado__in=['pendiente_visita', 'visita_agendada', 'en_evaluacion']
    ).select_related(
        'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
    ).prefetch_related('visitas_tecnicas').order_by('-fecha_registro')
    
    # Paginación
    paginator = Paginator(hogares, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_pendientes': hogares.filter(estado='pendiente_visita').count(),
        'total_agendadas': hogares.filter(estado='visita_agendada').count(),
        'total_evaluacion': hogares.filter(estado='en_evaluacion').count(),
    }
    
    return render(request, 'admin/visitas/listar_hogares_pendientes.html', context)


@login_required
def agendar_visita_tecnica(request, hogar_id):
    """
    Agenda una visita técnica para un hogar comunitario y envía correo a la madre.
    """
    hogar = get_object_or_404(HogarComunitario, id=hogar_id)
    
    if request.method == 'POST':
        form = AgendarVisitaTecnicaForm(request.POST)
        if form.is_valid():
            visita = form.save(commit=False)
            visita.creado_por = request.user
            visita.save()
            
            # Actualizar estado del hogar
            hogar.estado = 'visita_agendada'
            hogar.save()
            
            # Enviar correo a la madre
            enviar_correo_visita_agendada(visita)
            
            messages.success(
                request,
                f'Visita agendada exitosamente para {hogar.nombre_hogar}. '
                f'Se ha enviado un correo a {hogar.madre.usuario.correo}'
            )
            return redirect('listar_hogares_pendientes_visita')
    else:
        form = AgendarVisitaTecnicaForm(initial={'hogar': hogar})
    
    context = {
        'form': form,
        'hogar': hogar,
    }
    
    return render(request, 'admin/visitas/agendar_visita.html', context)


@login_required
def crear_acta_visita(request, visita_id):
    """
    Formulario para crear el Acta de Visita Técnica (V1).
    Dividido en pasos para facilitar el llenado.
    """
    visita = get_object_or_404(VisitaTecnica, id=visita_id)
    
    # Verificar si ya existe un acta
    try:
        acta = visita.acta
        editing = True
    except ActaVisitaTecnica.DoesNotExist:
        acta = None
        editing = False
    
    if request.method == 'POST':
        if acta:
            form = ActaVisitaTecnicaForm(request.POST, request.FILES, instance=acta)
        else:
            form = ActaVisitaTecnicaForm(request.POST, request.FILES)
        
        if form.is_valid():
            acta_obj = form.save(commit=False)
            acta_obj.visita = visita
            acta_obj.completado_por = request.user
            acta_obj.save()
            
            # Actualizar estado de la visita
            visita.estado = 'completada'
            visita.fecha_realizacion = timezone.now()
            visita.save()
            
            # Actualizar estado del hogar según resultado
            hogar = visita.hogar
            if acta_obj.resultado_visita == 'aprobado':
                hogar.estado = 'activo'
                hogar.fecha_habilitacion = timezone.now()
                hogar.capacidad_maxima = acta_obj.capacidad_recomendada
                # Actualizar geolocalización verificada
                hogar.geolocalizacion_lat = acta_obj.geolocalizacion_lat_verificada
                hogar.geolocalizacion_lon = acta_obj.geolocalizacion_lon_verificada
                hogar.save()
                enviar_correo_hogar_aprobado(hogar, acta_obj)
                
            elif acta_obj.resultado_visita == 'aprobado_condiciones':
                hogar.estado = 'activo'
                hogar.fecha_habilitacion = timezone.now()
                hogar.capacidad_maxima = acta_obj.capacidad_recomendada
                hogar.geolocalizacion_lat = acta_obj.geolocalizacion_lat_verificada
                hogar.geolocalizacion_lon = acta_obj.geolocalizacion_lon_verificada
                hogar.save()
                enviar_correo_hogar_aprobado_condiciones(hogar, acta_obj)
                
            elif acta_obj.resultado_visita == 'rechazado':
                hogar.estado = 'rechazado'
                hogar.save()
                enviar_correo_hogar_rechazado(hogar, acta_obj)
                
            elif acta_obj.resultado_visita == 'requiere_segunda_visita':
                hogar.estado = 'en_evaluacion'
                hogar.save()
            
            messages.success(request, f'Acta de visita guardada exitosamente. Estado del hogar: {hogar.get_estado_display()}')
            return redirect('ver_acta_visita', acta_id=acta_obj.id)
    else:
        if acta:
            form = ActaVisitaTecnicaForm(instance=acta)
        else:
            form = ActaVisitaTecnicaForm()
    
    # Obtener el paso actual
    current_step = request.GET.get('step', '1')
    
    context = {
        'form': form,
        'visita': visita,
        'hogar': visita.hogar,
        'current_step': current_step,
        'editing': editing,
    }
    
    return render(request, 'admin/visitas/crear_acta.html', context)


@login_required
def ver_acta_visita(request, acta_id):
    """
    Vista completa del acta de visita técnica.
    """
    acta = get_object_or_404(ActaVisitaTecnica, id=acta_id)
    
    context = {
        'acta': acta,
        'visita': acta.visita,
        'hogar': acta.visita.hogar,
    }
    
    return render(request, 'admin/visitas/ver_acta.html', context)


@login_required
def descargar_acta_pdf(request, acta_id):
    """
    Genera y descarga el acta de visita en formato PDF.
    """
    acta = get_object_or_404(ActaVisitaTecnica, id=acta_id)
    
    template = get_template('admin/visitas/acta_pdf.html')
    context = {
        'acta': acta,
        'visita': acta.visita,
        'hogar': acta.visita.hogar,
    }
    html = template.render(context)
    
    # Crear PDF
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Acta_Visita_{acta.visita.hogar.nombre_hogar}.pdf"'
        return response
    
    return HttpResponse('Error al generar PDF', status=400)


@login_required
def listar_visitas_tecnicas(request):
    """
    Lista todas las visitas técnicas con filtros.
    """
    visitas = VisitaTecnica.objects.select_related(
        'hogar__madre__usuario', 'visitador', 'creado_por'
    ).order_by('-fecha_programada')
    
    # Filtros
    estado = request.GET.get('estado')
    if estado:
        visitas = visitas.filter(estado=estado)
    
    tipo = request.GET.get('tipo')
    if tipo:
        visitas = visitas.filter(tipo_visita=tipo)
    
    visitador_id = request.GET.get('visitador')
    if visitador_id:
        visitas = visitas.filter(visitador_id=visitador_id)
    
    # Paginación
    paginator = Paginator(visitas, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Lista de visitadores para el filtro
    visitadores = Usuario.objects.filter(rol='administrador').order_by('nombres')
    
    context = {
        'page_obj': page_obj,
        'visitadores': visitadores,
        'estado_filtro': estado,
        'tipo_filtro': tipo,
    }
    
    return render(request, 'admin/visitas/listar_visitas.html', context)


# ========================================================================================
# 📧 FUNCIONES DE ENVÍO DE CORREOS
# ========================================================================================

def enviar_correo_visita_agendada(visita):
    """
    Envía correo a la madre comunitaria informando la visita agendada.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    
    hogar = visita.hogar
    madre = hogar.madre
    
    subject = f'Visita Técnica Agendada - {hogar.nombre_hogar}'
    
    context = {
        'madre': madre,
        'hogar': hogar,
        'visita': visita,
    }
    
    message = render_to_string('emails/visita_agendada.html', context)
    
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [madre.usuario.correo],
            fail_silently=False,
            html_message=message
        )
        
        # Actualizar registro de envío
        visita.correo_enviado = True
        visita.fecha_envio_correo = timezone.now()
        visita.save()
        
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False


def enviar_correo_hogar_aprobado(hogar, acta):
    """
    Envía correo a la madre cuando el hogar es aprobado.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    
    madre = hogar.madre
    
    subject = f'¡Felicidades! Tu hogar ha sido habilitado - {hogar.nombre_hogar}'
    
    context = {
        'madre': madre,
        'hogar': hogar,
        'acta': acta,
    }
    
    message = render_to_string('emails/hogar_aprobado.html', context)
    
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [madre.usuario.correo],
            fail_silently=False,
            html_message=message
        )
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False


def enviar_correo_hogar_aprobado_condiciones(hogar, acta):
    """
    Envía correo cuando el hogar es aprobado con condiciones.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    
    madre = hogar.madre
    
    subject = f'Hogar Aprobado con Condiciones - {hogar.nombre_hogar}'
    
    context = {
        'madre': madre,
        'hogar': hogar,
        'acta': acta,
        'condiciones': acta.condiciones_aprobacion,
    }
    
    message = render_to_string('emails/hogar_aprobado_condiciones.html', context)
    
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [madre.usuario.correo],
            fail_silently=False,
            html_message=message
        )
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False


def enviar_correo_hogar_rechazado(hogar, acta):
    """
    Envía correo cuando el hogar es rechazado.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.conf import settings
    
    madre = hogar.madre
    
    subject = f'Resultado de Visita Técnica - {hogar.nombre_hogar}'
    
    context = {
        'madre': madre,
        'hogar': hogar,
        'acta': acta,
    }
    
    message = render_to_string('emails/hogar_rechazado.html', context)
    
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [madre.usuario.correo],
            fail_silently=False,
            html_message=message
        )
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False


# ============================================================================
# 🆕 NUEVAS VISTAS - SISTEMA DE DOS FASES (FORMULARIO 2)
# ============================================================================

@login_required
@rol_requerido('administrador')
def completar_visita_tecnica(request, hogar_id):
    """
    Vista para completar el Formulario 2 (Visita Técnica).
    
    Esta vista se accede cuando:
    1. La visita técnica ya ocurrió
    2. El administrador necesita completar la revisión del hogar
    
    Solo es accesible si:
    - El hogar está en estado 'pendiente_revision' o 'en_revision'
    - El formulario técnico NO está completo (formulario_completo = False)
    """
    hogar = get_object_or_404(HogarComunitario, pk=hogar_id)
    
    # Verificar que el hogar no esté ya completado
    if hogar.formulario_completo:
        messages.warning(
            request, 
            f'El hogar "{hogar.nombre_hogar}" ya tiene el formulario técnico completo. '
            'Use la opción de edición si necesita modificar los datos.'
        )
        return redirect('detalle_hogar', hogar_id=hogar.id)
    
    # Verificar que el hogar esté en un estado válido para completar
    estados_validos = ['pendiente_revision', 'en_revision', 'pendiente_visita']
    if hogar.estado not in estados_validos:
        messages.error(
            request,
            f'No se puede completar la visita técnica. '
            f'Estado actual: {hogar.get_estado_display()}. '
            f'Estados válidos: Pendiente de Revisión, En Revisión.'
        )
        return redirect('detalle_hogar', hogar_id=hogar.id)
    
    if request.method == 'POST':
        form = HogarFormulario2Form(request.POST, request.FILES, instance=hogar)
        
        if form.is_valid():
            # El formulario automáticamente:
            # - Calcula la capacidad: piso(área_m²/2)
            # - Establece formulario_completo = True
            # - Cambia estado a 'en_revision'
            hogar_actualizado = form.save()
            
            # Mensajes informativos según el área
            area = hogar_actualizado.area_social_m2
            capacidad = hogar_actualizado.capacidad_calculada
            
            if area < 24:
                messages.error(
                    request,
                    f'⚠️ HOGAR NO APTO: El área social ({area} m²) es inferior al mínimo requerido (24 m²). '
                    f'Este hogar NO PUEDE SER APROBADO.'
                )
            elif 24 <= area < 30:
                messages.success(
                    request,
                    f'✅ Visita técnica completada exitosamente. '
                    f'Área: {area} m². Capacidad calculada: {capacidad} niños. '
                    f'Estado: En Revisión. Puede proceder a aprobar el hogar.'
                )
            else:  # area >= 30
                messages.success(
                    request,
                    f'✅ Visita técnica completada exitosamente. '
                    f'Área: {area} m². Capacidad máxima: {capacidad} niños (límite 15). '
                    f'Estado: En Revisión. Puede proceder a aprobar el hogar.'
                )
            
            return redirect('lista_hogares_revision')
        else:
            messages.error(
                request, 
                'Por favor corrija los errores del formulario antes de continuar.'
            )
    else:
        form = HogarFormulario2Form(instance=hogar)
    
    context = {
        'form': form,
        'hogar': hogar,
        'titulo': f'Visita Técnica - {hogar.nombre_hogar}',
        'es_nuevo': not hogar.area_social_m2,  # True si es primera vez
    }
    return render(request, 'admin/hogar_formulario2.html', context)


@login_required
@rol_requerido('administrador')
def lista_hogares_revision(request):
    """
    Panel principal de gestión de hogares.
    
    Muestra todos los hogares con filtros avanzados y alertas de visitas.
    Estados: pendiente_revision, en_revision, aprobado, rechazado, en_mantenimiento
    """
    from datetime import timedelta
    from django.utils import timezone
    
    # Filtros
    estado_filter = request.GET.get('estado', '')
    regional_filter = request.GET.get('regional', '')
    localidad_filter = request.GET.get('localidad', '')
    buscar = request.GET.get('buscar', '')
    alerta_filter = request.GET.get('alerta', '')  # visitas próximas, vencidas
    
    # Query base
    hogares = HogarComunitario.objects.select_related(
        'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
    ).order_by('-fecha_registro')
    
    # Aplicar filtros
    if estado_filter:
        hogares = hogares.filter(estado=estado_filter)
    
    if regional_filter:
        hogares = hogares.filter(regional_id=regional_filter)
    
    if localidad_filter:
        hogares = hogares.filter(localidad_bogota_id=localidad_filter)
    
    if buscar:
        hogares = hogares.filter(
            Q(nombre_hogar__icontains=buscar) |
            Q(direccion__icontains=buscar) |
            Q(barrio__icontains=buscar) |
            Q(madre__usuario__nombres__icontains=buscar) |
            Q(madre__usuario__apellidos__icontains=buscar)
        )
    
    # Filtros de alertas de visitas
    hoy = timezone.now().date()
    if alerta_filter == 'proximas':
        # Visitas en los próximos 7 días
        fecha_limite = hoy + timedelta(days=7)
        hogares = hogares.filter(
            fecha_primera_visita__gte=hoy,
            fecha_primera_visita__lte=fecha_limite
        )
    elif alerta_filter == 'vencidas':
        # Visitas que ya pasaron pero el hogar no está aprobado
        hogares = hogares.filter(
            fecha_primera_visita__lt=hoy,
            estado__in=['pendiente_revision', 'en_revision']
        )
    
    # Contar por estado
    total_hogares = HogarComunitario.objects.count()
    hogares_pendientes = HogarComunitario.objects.filter(estado='pendiente_revision').count()
    hogares_en_revision = HogarComunitario.objects.filter(estado='en_revision').count()
    hogares_aprobados = HogarComunitario.objects.filter(estado='aprobado').count()
    hogares_rechazados = HogarComunitario.objects.filter(estado='rechazado').count()
    hogares_mantenimiento = HogarComunitario.objects.filter(estado='en_mantenimiento').count()
    
    # Alertas de visitas
    visitas_proximas = HogarComunitario.objects.filter(
        fecha_primera_visita__gte=hoy,
        fecha_primera_visita__lte=hoy + timedelta(days=7),
        estado__in=['pendiente_revision', 'en_revision']
    ).count()
    
    visitas_vencidas = HogarComunitario.objects.filter(
        fecha_primera_visita__lt=hoy,
        estado__in=['pendiente_revision', 'en_revision']
    ).count()
    
    # Paginación
    paginator = Paginator(hogares, 20)
    page = request.GET.get('page', 1)
    hogares_paginados = paginator.get_page(page)
    
    # Datos para filtros
    regionales = Regional.objects.all().order_by('nombre')
    from core.models import LocalidadBogota
    localidades = LocalidadBogota.objects.all().order_by('nombre')
    
    context = {
        'hogares': hogares_paginados,
        'total_hogares': total_hogares,
        'hogares_pendientes': hogares_pendientes,
        'hogares_en_revision': hogares_en_revision,
        'hogares_aprobados': hogares_aprobados,
        'hogares_rechazados': hogares_rechazados,
        'hogares_mantenimiento': hogares_mantenimiento,
        'visitas_proximas': visitas_proximas,
        'visitas_vencidas': visitas_vencidas,
        'regionales': regionales,
        'localidades': localidades,
        'estado_filter': estado_filter,
        'regional_filter': regional_filter,
        'localidad_filter': localidad_filter,
        'buscar': buscar,
        'alerta_filter': alerta_filter,
        'hoy': hoy,
    }
    return render(request, 'admin/lista_hogares_revision.html', context)


@login_required
@rol_requerido('administrador')
def aprobar_rechazar_hogar(request, hogar_id):
    """
    Vista para aprobar o rechazar un hogar después de la revisión.
    
    Solo accesible si:
    - El hogar está en estado 'en_revision'
    - El formulario técnico está completo
    - El área cumple con los requisitos (≥24 m² para aprobar)
    """
    hogar = get_object_or_404(HogarComunitario, pk=hogar_id)
    
    # Verificar que el hogar esté listo para revisión final
    if hogar.estado not in ['en_revision', 'pendiente_revision']:
        messages.error(
            request,
            f'Este hogar no está en estado de revisión. '
            f'Estado actual: {hogar.get_estado_display()}'
        )
        return redirect('detalle_hogar', hogar_id=hogar.id)
    
    # Verificar que el formulario técnico esté completo
    if not hogar.formulario_completo:
        messages.error(
            request,
            'Debe completar la visita técnica (Formulario 2) antes de aprobar o rechazar el hogar.'
        )
        return redirect('completar_visita_tecnica', hogar_id=hogar.id)
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        observaciones = request.POST.get('observaciones', '').strip()
        
        if accion == 'aprobar':
            # Verificar área mínima antes de aprobar
            if hogar.area_social_m2 and hogar.area_social_m2 < 24:
                messages.error(
                    request,
                    f'⚠️ NO SE PUEDE APROBAR: El área social ({hogar.area_social_m2} m²) '
                    f'es inferior al mínimo requerido (24 m²). '
                    f'Debe rechazar el hogar o solicitar modificaciones.'
                )
                return redirect('aprobar_rechazar_hogar', hogar_id=hogar.id)
            
            # Aprobar hogar
            hogar.estado = 'aprobado'
            hogar.save()
            
            messages.success(
                request,
                f'✅ Hogar "{hogar.nombre_hogar}" APROBADO exitosamente. '
                f'Capacidad autorizada: {hogar.capacidad_calculada} niños. '
                f'La madre comunitaria ha sido notificada.'
            )
            
            # TODO: Enviar correo de aprobación
            
        elif accion == 'rechazar':
            if not observaciones:
                messages.error(
                    request,
                    'Debe proporcionar observaciones al rechazar un hogar.'
                )
                return redirect('aprobar_rechazar_hogar', hogar_id=hogar.id)
            
            # Rechazar hogar
            hogar.estado = 'rechazado'
            hogar.save()
            
            messages.warning(
                request,
                f'Hogar "{hogar.nombre_hogar}" RECHAZADO. '
                f'Se ha notificado a la madre comunitaria.'
            )
            
            # TODO: Guardar observaciones y enviar correo de rechazo
            
        elif accion == 'mantenimiento':
            hogar.estado = 'en_mantenimiento'
            hogar.save()
            
            messages.info(
                request,
                f'Hogar "{hogar.nombre_hogar}" marcado como EN MANTENIMIENTO.'
            )
        
        return redirect('lista_hogares_revision')
    
    # GET - Mostrar formulario de aprobación/rechazo
    context = {
        'hogar': hogar,
        'puede_aprobar': hogar.area_social_m2 and hogar.area_social_m2 >= 24,
        'area_insuficiente': hogar.area_social_m2 and hogar.area_social_m2 < 24,
    }
    return render(request, 'admin/aprobar_rechazar_hogar.html', context)


@login_required
@rol_requerido('administrador')
def detalle_hogar(request, hogar_id):
    """
    Vista detallada de un hogar comunitario.
    Muestra toda la información de ambos formularios si están completos.
    """
    hogar = get_object_or_404(
        HogarComunitario.objects.select_related(
            'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
        ),
        pk=hogar_id
    )
    
    # Obtener convivientes del hogar
    from core.models import ConvivienteHogar
    convivientes = ConvivienteHogar.objects.filter(hogar=hogar).order_by('fecha_registro')
    
    context = {
        'hogar': hogar,
        'convivientes': convivientes,
        'tiene_formulario_completo': hogar.formulario_completo,
        'puede_completar_visita': not hogar.formulario_completo and hogar.estado in ['pendiente_revision', 'en_revision', 'pendiente_visita'],
        'puede_aprobar': hogar.formulario_completo and hogar.estado == 'en_revision',
    }
    return render(request, 'admin/detalle_hogar.html', context)


# ============================================================================
# 🆕 DASHBOARD DEL PADRE - VISTA DE HOGARES COMUNITARIOS
# ============================================================================

@login_required
@rol_requerido('padre')
def padre_ver_hogares(request):
    """
    Vista para que los padres vean todos los hogares comunitarios disponibles.
    Permite filtrar por localidad y ver información completa.
    """
    # Filtros
    localidad_filter = request.GET.get('localidad', '')
    regional_filter = request.GET.get('regional', '')
    buscar = request.GET.get('buscar', '')
    
    # Solo hogares aprobados
    hogares = HogarComunitario.objects.filter(
        estado='aprobado'
    ).select_related(
        'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
    ).order_by('nombre_hogar')
    
    # Aplicar filtros
    if localidad_filter:
        hogares = hogares.filter(localidad_bogota_id=localidad_filter)
    
    if regional_filter:
        hogares = hogares.filter(regional_id=regional_filter)
    
    if buscar:
        hogares = hogares.filter(
            Q(nombre_hogar__icontains=buscar) |
            Q(direccion__icontains=buscar) |
            Q(barrio__icontains=buscar) |
            Q(madre__usuario__nombres__icontains=buscar) |
            Q(madre__usuario__apellidos__icontains=buscar)
        )
    
    # Estadísticas generales
    total_hogares = HogarComunitario.objects.filter(estado='aprobado').count()
    total_capacidad = sum([h.capacidad_calculada or 0 for h in HogarComunitario.objects.filter(estado='aprobado')])
    
    # Calcular capacidad disponible (aproximado - restar niños matriculados)
    hogares_con_ninos = []
    for hogar in hogares:
        ninos_count = Nino.objects.filter(hogar=hogar).count()
        capacidad_disponible = (hogar.capacidad_calculada or 0) - ninos_count
        hogares_con_ninos.append({
            'hogar': hogar,
            'ninos_count': ninos_count,
            'capacidad_disponible': max(0, capacidad_disponible),
            'ocupacion_pct': int((ninos_count / hogar.capacidad_calculada * 100)) if hogar.capacidad_calculada else 0
        })
    
    # Paginación
    paginator = Paginator(hogares_con_ninos, 12)
    page = request.GET.get('page', 1)
    hogares_paginados = paginator.get_page(page)
    
    # Datos para filtros
    from core.models import LocalidadBogota
    localidades = LocalidadBogota.objects.all().order_by('nombre')
    regionales = Regional.objects.all().order_by('nombre')
    
    context = {
        'hogares': hogares_paginados,
        'total_hogares': total_hogares,
        'total_capacidad': total_capacidad,
        'localidades': localidades,
        'regionales': regionales,
        'localidad_filter': localidad_filter,
        'regional_filter': regional_filter,
        'buscar': buscar,
    }
    return render(request, 'padre/hogares_disponibles.html', context)


@login_required
@rol_requerido('padre')
def padre_detalle_hogar(request, hogar_id):
    """
    Vista detallada de un hogar comunitario para padres.
    Muestra información completa, lista de niños, fotos, etc.
    """
    hogar = get_object_or_404(
        HogarComunitario.objects.select_related(
            'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
        ),
        pk=hogar_id,
        estado='aprobado'  # Solo hogares aprobados visibles para padres
    )
    
    # Obtener convivientes
    from core.models import ConvivienteHogar
    convivientes = ConvivienteHogar.objects.filter(hogar=hogar).order_by('nombre_completo')
    
    # Obtener niños del hogar
    ninos = Nino.objects.filter(hogar=hogar).select_related('padre__usuario')
    
    # Calcular capacidad disponible
    capacidad_ocupada = ninos.count()
    capacidad_disponible = (hogar.capacidad_calculada or 0) - capacidad_ocupada
    
    context = {
        'hogar': hogar,
        'convivientes': convivientes,
        'ninos': ninos,
        'capacidad_ocupada': capacidad_ocupada,
        'capacidad_disponible': max(0, capacidad_disponible),
        'ocupacion_pct': int((capacidad_ocupada / hogar.capacidad_calculada * 100)) if hogar.capacidad_calculada else 0,
    }
    return render(request, 'padre/detalle_hogar_publico.html', context)


@login_required
@rol_requerido('padre')
def padre_dashboard_mejorado(request):
    """
    Dashboard mejorado para padres con estadísticas y gráficas.
    """
    import json
    
    try:
        padre = Padre.objects.get(usuario=request.user)
        
        # Niños del padre
        ninos = Nino.objects.filter(padre=padre).select_related('hogar', 'hogar__madre__usuario')
        
        # Estadísticas de hogares
        total_hogares = HogarComunitario.objects.filter(estado='aprobado').count()
        hogares_activos = HogarComunitario.objects.filter(
            estado='aprobado',
            nino__padre=padre
        ).distinct().count()
        
        hogares_pendientes = HogarComunitario.objects.filter(
            estado__in=['pendiente_revision', 'en_revision']
        ).count()
        
        # Estadísticas por localidad
        from django.db.models import Count
        from core.models import LocalidadBogota
        
        hogares_por_localidad = HogarComunitario.objects.filter(
            estado='aprobado',
            localidad_bogota__isnull=False
        ).values('localidad_bogota__nombre').annotate(
            total=Count('id')
        ).order_by('-total')[:5]
        
        # Convertir datos para Chart.js
        localidades_json = json.dumps([
            {
                'localidad': item['localidad_bogota__nombre'],
                'total': item['total']
            }
            for item in hogares_por_localidad
        ])
        
        # Última asistencia de los niños
        ninos_data = []
        for nino in ninos:
            ultima_asistencia = Asistencia.objects.filter(
                nino=nino
            ).order_by('-fecha').first()
            
            ninos_data.append({
                'nino': nino,
                'get_full_name': nino.get_full_name(),
                'calcular_edad': nino.calcular_edad(),
                'documento': nino.documento,
                'hogar': nino.hogar,
                'ultima_asistencia': ultima_asistencia,
            })
        
        context = {
            'padre': padre,
            'ninos': ninos_data,
            'total_hogares': total_hogares,
            'hogares_activos': hogares_activos,
            'hogares_pendientes': hogares_pendientes,
            'hogares_por_localidad': hogares_por_localidad,
            'localidades_json': localidades_json,
        }
        return render(request, 'padre/dashboard_mejorado.html', context)
        
    except Padre.DoesNotExist:
        messages.error(request, 'No se encontró información de padre asociada a su usuario.')
        return redirect('home')


# ---------- DASHBOARD DE HOGARES COMUNITARIOS ----------
@login_required
@rol_requerido('administrador')
def hogares_dashboard(request):
    """Dashboard para gestionar hogares comunitarios con vista por localidad."""
    from django.db.models import Q, Count
    from collections import defaultdict
    
    # Filtros desde GET params
    filtro_localidad = request.GET.get('localidad', '')
    filtro_estado = request.GET.get('estado', '')
    busqueda = request.GET.get('q', '').strip()
    
    # Query base con relaciones pre-cargadas y anotaciones
    hogares = HogarComunitario.objects.select_related(
        'madre__usuario', 'regional', 'ciudad', 'localidad_bogota'
    ).annotate(
        ninos_count=Count('ninos', filter=Q(ninos__estado='activo'))
    )
    
    # Aplicar filtros
    if filtro_localidad:
        hogares = hogares.filter(localidad_bogota__nombre__icontains=filtro_localidad)
    
    if filtro_estado:
        hogares = hogares.filter(estado=filtro_estado)
    
    if busqueda:
        hogares = hogares.filter(
            Q(nombre_hogar__icontains=busqueda) |
            Q(madre__usuario__nombres__icontains=busqueda) |
            Q(madre__usuario__apellidos__icontains=busqueda) |
            Q(direccion__icontains=busqueda)
        )
    
    # Ordenar por localidad y fecha de registro
    hogares = hogares.order_by('localidad_bogota__numero', '-fecha_registro')
    
    # Obtener todas las localidades de Bogotá para el filtro
    from core.models import LocalidadBogota
    localidades = LocalidadBogota.objects.all().order_by('numero')
    
    # Agrupar hogares por localidad
    hogares_por_localidad = defaultdict(list)
    for hogar in hogares:
        if hogar.localidad_bogota:
            # Usar formato "Ciudad - Localidad" para hogares en Bogotá
            clave_localidad = f"{hogar.ciudad.nombre} - {hogar.localidad_bogota.nombre}"
        elif hogar.ciudad:
            # Para hogares sin localidad asignada
            clave_localidad = hogar.ciudad.nombre
        else:
            clave_localidad = 'Sin ubicación'
        
        hogares_por_localidad[clave_localidad].append(hogar)
    
    # Convertir a diccionario normal ordenado
    hogares_por_localidad = dict(sorted(hogares_por_localidad.items()))
    
    context = {
        'hogares_por_localidad': hogares_por_localidad,
        'localidades': localidades,
        'filtro_localidad': filtro_localidad,
        'filtro_estado': filtro_estado,
        'busqueda': busqueda,
    }
    
    return render(request, 'admin/hogares_dashboard.html', context)


# ---------- REALIZAR VISITA TÉCNICA ----------
@login_required
@rol_requerido('administrador')
def realizar_visita_tecnica(request, hogar_id):
    """
    Vista para realizar la visita técnica y completar el acta de evaluación.
    Solo accesible en la fecha programada o después.
    """
    hogar = get_object_or_404(HogarComunitario, id=hogar_id)
    
    # Obtener la visita programada
    visita = hogar.visitas_tecnicas.filter(
        estado='agendada'
    ).order_by('fecha_programada').first()
    
    if not visita:
        messages.error(request, 'No hay visitas programadas para este hogar.')
        return redirect('hogares_dashboard')
    
    # Verificar si se puede realizar la visita (fecha programada o posterior)
    puede_realizar_visita = visita.fecha_programada.date() <= timezone.now().date()
    
    if request.method == 'POST' and puede_realizar_visita:
        try:
            with transaction.atomic():
                # Crear el acta de visita
                acta = ActaVisitaTecnica.objects.create(
                    visita=visita,
                    # Geolocalización
                    geolocalizacion_lat_verificada=request.POST.get('geolocalizacion_lat_verificada'),
                    geolocalizacion_lon_verificada=request.POST.get('geolocalizacion_lon_verificada'),
                    direccion_verificada=request.POST.get('direccion_verificada'),
                    direccion_coincide='direccion_coincide' in request.POST,
                    observaciones_direccion=request.POST.get('observaciones_direccion', ''),
                    estrato_verificado=request.POST.get('estrato_verificado'),
                    estrato_coincide='estrato_coincide' in request.POST,
                    foto_recibo_servicio=request.FILES.get('foto_recibo_servicio'),
                    
                    # Servicios básicos
                    tiene_agua_potable='tiene_agua_potable' in request.POST,
                    agua_continua='agua_continua' in request.POST,
                    agua_legal='agua_legal' in request.POST,
                    tiene_energia='tiene_energia' in request.POST,
                    energia_continua='energia_continua' in request.POST,
                    energia_legal='energia_legal' in request.POST,
                    tiene_alcantarillado='tiene_alcantarillado' in request.POST,
                    manejo_excretas_adecuado='manejo_excretas_adecuado' in request.POST,
                    
                    # Infraestructura
                    estado_pisos=request.POST.get('estado_pisos'),
                    estado_paredes=request.POST.get('estado_paredes'),
                    estado_techos=request.POST.get('estado_techos'),
                    ventilacion_adecuada='ventilacion_adecuada' in request.POST,
                    iluminacion_natural_adecuada='iluminacion_natural_adecuada' in request.POST,
                    observaciones_infraestructura=request.POST.get('observaciones_infraestructura', ''),
                    
                    # Riesgos
                    proximidad_rios='proximidad_rios' in request.POST,
                    proximidad_deslizamientos='proximidad_deslizamientos' in request.POST,
                    proximidad_trafico_intenso='proximidad_trafico_intenso' in request.POST,
                    proximidad_contaminacion='proximidad_contaminacion' in request.POST,
                    nivel_riesgo_general=request.POST.get('nivel_riesgo_general'),
                    descripcion_riesgos=request.POST.get('descripcion_riesgos', ''),
                    
                    # Espacios
                    area_social_largo=request.POST.get('area_social_largo'),
                    area_social_ancho=request.POST.get('area_social_ancho'),
                    foto_area_social_medidas=request.FILES.get('foto_area_social_medidas'),
                    tiene_patio_cubierto='tiene_patio_cubierto' in request.POST,
                    patio_largo=request.POST.get('patio_largo') if 'tiene_patio_cubierto' in request.POST else None,
                    patio_ancho=request.POST.get('patio_ancho') if 'tiene_patio_cubierto' in request.POST else None,
                    foto_patio_medidas=request.FILES.get('foto_patio_medidas'),
                    
                    # Baños
                    num_banos_verificado=request.POST.get('num_banos_verificado'),
                    estado_higiene_banos=request.POST.get('estado_higiene_banos'),
                    foto_bano_1=request.FILES.get('foto_bano_1'),
                    foto_bano_2=request.FILES.get('foto_bano_2'),
                    
                    # Fachada
                    foto_fachada=request.FILES.get('foto_fachada'),
                    foto_fachada_numeracion=request.FILES.get('foto_fachada_numeracion'),
                    
                    # Capacidad
                    capacidad_calculada=request.POST.get('capacidad_calculada'),
                    capacidad_recomendada=request.POST.get('capacidad_recomendada'),
                    justificacion_capacidad=request.POST.get('justificacion_capacidad', ''),
                    
                    # Conclusión
                    resultado_visita=request.POST.get('recomendacion_habilitacion'),
                    condiciones_aprobacion=request.POST.get('condiciones_aprobacion', ''),
                    observaciones_generales=request.POST.get('observaciones_generales', ''),
                    completado_por=request.user
                )
                
                # Calcular áreas
                area_social_total = float(acta.area_social_largo) * float(acta.area_social_ancho)
                acta.area_social_total = area_social_total
                
                if acta.tiene_patio_cubierto and acta.patio_largo and acta.patio_ancho:
                    acta.patio_total = float(acta.patio_largo) * float(acta.patio_ancho)
                
                acta.save()
                
                # Actualizar el estado de la visita
                visita.estado = 'completada'
                visita.fecha_realizacion = timezone.now()
                visita.visitador = request.user
                visita.save()
                
                # Actualizar el hogar según la recomendación
                recomendacion = request.POST.get('recomendacion_habilitacion')
                capacidad_recomendada = int(request.POST.get('capacidad_recomendada'))
                if recomendacion == 'aprobado':
                    hogar.estado = 'aprobado'
                    hogar.capacidad_maxima = capacidad_recomendada
                    hogar.capacidad_calculada = capacidad_recomendada  # Sincronizar
                    hogar.capacidad = capacidad_recomendada  # Sincronizar
                    hogar.fecha_habilitacion = timezone.now()
                    hogar.formulario_completo = True
                    
                    # Enviar correo de aprobación con credenciales
                    enviar_correo_aprobacion(hogar, request.user)
                    
                elif recomendacion == 'aprobado_condicional' or recomendacion == 'aprobado_condiciones':
                    hogar.estado = 'en_revision'
                    hogar.capacidad_maxima = capacidad_recomendada
                    hogar.capacidad_calculada = capacidad_recomendada  # Sincronizar
                    hogar.capacidad = capacidad_recomendada  # Sincronizar
                    
                else:  # rechazado
                    hogar.estado = 'rechazado'
                
                # Actualizar ubicación del hogar
                hogar.geolocalizacion_lat = acta.geolocalizacion_lat_verificada
                hogar.geolocalizacion_lon = acta.geolocalizacion_lon_verificada
                hogar.num_habitaciones = acta.num_banos_verificado
                hogar.save()
                
                messages.success(request, f'✅ Acta de visita guardada exitosamente. Hogar actualizado a estado: {hogar.get_estado_display()}')
                return redirect('detalle_hogar', hogar_id=hogar.id)
                
        except Exception as e:
            messages.error(request, f'❌ Error al guardar el acta: {str(e)}')
            return redirect('hogares_dashboard')
    
    context = {
        'hogar': hogar,
        'visita': visita,
        'puede_realizar_visita': puede_realizar_visita,
    }
    
    return render(request, 'admin/visita_tecnica_form.html', context)


# ---------- ENVIAR CORREO DE APROBACIÓN ----------
def enviar_correo_aprobacion(hogar, aprobador):
    """
    Envía correo al agente educativo cuando su hogar es aprobado.
    Incluye credenciales de acceso al sistema.
    """
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.utils.html import strip_tags
    
    usuario = hogar.madre.usuario
    
    # Generar contraseña temporal (documento)
    contraseña_temporal = usuario.documento
    
    # Si el usuario no tiene contraseña, configurarla
    if not usuario.password or usuario.password == '':
        usuario.set_password(contraseña_temporal)
        usuario.save()
    
    # Mensaje HTML
    mensaje_html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(135deg, #004080 0%, #0066cc 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
            .content {{ background: white; padding: 30px; border: 1px solid #ddd; }}
            .credentials {{ background: #e8f5e9; border-left: 5px solid #4caf50; padding: 20px; margin: 20px 0; border-radius: 5px; }}
            .credentials h3 {{ color: #2e7d32; margin-top: 0; }}
            .credential-item {{ margin: 10px 0; font-size: 16px; }}
            .credential-item strong {{ display: inline-block; width: 120px; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; border-radius: 0 0 10px 10px; font-size: 12px; color: #666; }}
            .btn {{ display: inline-block; padding: 12px 30px; background: #007bff; color: white; text-decoration: none; border-radius: 5px; margin: 20px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎉 ¡Felicitaciones!</h1>
                <p>Su Hogar Comunitario ha sido APROBADO</p>
            </div>
            
            <div class="content">
                <p>Estimado/a <strong>{usuario.nombres} {usuario.apellidos}</strong>,</p>
                
                <p>Nos complace informarle que su hogar comunitario <strong>"{hogar.nombre_hogar}"</strong> ha sido aprobado exitosamente tras la visita técnica realizada.</p>
                
                <div class="credentials">
                    <h3>📋 Datos del Hogar Aprobado:</h3>
                    <div class="credential-item"><strong>Hogar:</strong> {hogar.nombre_hogar}</div>
                    <div class="credential-item"><strong>Dirección:</strong> {hogar.direccion}, {hogar.barrio}</div>
                    <div class="credential-item"><strong>Capacidad Máxima:</strong> {hogar.capacidad_maxima} niños</div>
                    <div class="credential-item"><strong>Regional:</strong> {hogar.regional.nombre}</div>
                    <div class="credential-item"><strong>Ciudad:</strong> {hogar.ciudad.nombre}</div>
                </div>
                
                <div class="credentials">
                    <h3>🔑 Credenciales de Acceso al Sistema</h3>
                    <p>Ya puedes ingresar a la plataforma ICBF Conecta con las siguientes credenciales:</p>
                    <div class="credential-item"><strong>Usuario:</strong> {usuario.documento}</div>
                    <div class="credential-item"><strong>Contraseña:</strong> {contraseña_temporal}</div>
                    <p style="margin-top: 15px; font-size: 14px; color: #666;">
                        <em>⚠️ Por seguridad, te recomendamos cambiar tu contraseña en tu primer ingreso.</em>
                    </p>
                </div>
                
                <div style="text-align: center;">
                    <a href="http://localhost:8000/login/" class="btn">Ingresar a ICBF Conecta</a>
                </div>
                
                <h3 style="color: #004080; margin-top: 30px;">📝 Próximos Pasos:</h3>
                <ol>
                    <li>Ingresa al sistema con tus credenciales</li>
                    <li>Completa tu perfil y revisa la información del hogar</li>
                    <li>Comienza el proceso de matriculación de niños</li>
                    <li>Mantén actualizada la información del hogar</li>
                </ol>
                
                <p style="margin-top: 20px;">Aprobado por: <strong>{aprobador.nombres} {aprobador.apellidos}</strong></p>
                <p>Fecha de aprobación: <strong>{timezone.now().strftime('%d de %B de %Y')}</strong></p>
            </div>
            
            <div class="footer">
                <p><strong>ICBF Conecta</strong></p>
                <p>Sistema de Gestión de Hogares Comunitarios</p>
                <p>Instituto Colombiano de Bienestar Familiar</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    mensaje_texto = strip_tags(mensaje_html)
    
    try:
        send_mail(
            '🎉 Hogar Comunitario Aprobado - Credenciales de Acceso',
            mensaje_texto,
            'noreply@icbf.gov.co',
            [usuario.correo],
            html_message=mensaje_html,
            fail_silently=False
        )
        return True
    except Exception as e:
        print(f"Error al enviar correo de aprobación: {e}")
        return False


# ---------- PROGRAMAR NUEVA VISITA ----------
@login_required
@rol_requerido('administrador')
def programar_visita(request, hogar_id):
    """
    Programa una nueva visita técnica para un hogar.
    """
    hogar = get_object_or_404(HogarComunitario, id=hogar_id)
    
    if request.method == 'POST':
        fecha_programada = request.POST.get('fecha_programada')
        tipo_visita = request.POST.get('tipo_visita', 'V2')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            from datetime import datetime
            fecha_visita_dt = datetime.strptime(fecha_programada, '%Y-%m-%d')
            fecha_visita_date = fecha_visita_dt.date()
            
            # Validar que la fecha no sea en el pasado
            from datetime import date
            if fecha_visita_date < date.today():
                messages.error(request, '❌ La fecha de visita no puede ser en el pasado.')
                return redirect('hogares_dashboard')
            
            # Actualizar fecha_primera_visita si no existe (hogar sin visita programada)
            if not hogar.fecha_primera_visita:
                hogar.fecha_primera_visita = fecha_visita_date
                hogar.save()
            
            # Crear la visita
            visita = VisitaTecnica.objects.create(
                hogar=hogar,
                fecha_programada=fecha_visita_dt,
                tipo_visita=tipo_visita,
                estado='agendada',
                creado_por=request.user,
                observaciones_agenda=observaciones
            )
            
            # Enviar correo de notificación
            from django.core.mail import send_mail
            usuario = hogar.madre.usuario
            
            mensaje = f"""
            Estimado/a {usuario.nombres} {usuario.apellidos},
            
            Se ha programado una nueva visita técnica para su hogar "{hogar.nombre_hogar}":
            
            Fecha: {visita.fecha_programada.strftime('%d de %B de %Y a las %H:%M')}
            Tipo: {visita.get_tipo_visita_display()}
            Observaciones: {observaciones}
            
            Por favor esté disponible en esa fecha.
            
            Saludos,
            ICBF Conecta
            """
            
            send_mail(
                'Nueva Visita Técnica Programada - ICBF Conecta',
                mensaje,
                'noreply@icbf.gov.co',
                [usuario.correo],
                fail_silently=True
            )
            
            visita.correo_enviado = True
            visita.fecha_envio_correo = timezone.now()
            visita.save()
            
            messages.success(request, '✅ Visita programada exitosamente. Se envió notificación por correo.')
            return redirect('hogares_dashboard')
            
        except Exception as e:
            messages.error(request, f'❌ Error al programar visita: {str(e)}')
    
    context = {
        'hogar': hogar,
    }
    
    return render(request, 'admin/programar_visita.html', context)


# ------------------------
# API: Actualizar Visitas del Hogar (Sistema de Gestión de Visitas)
# ------------------------
@login_required
@require_http_methods(["POST"])
def actualizar_visitas_hogar(request, hogar_id):
    """
    API para actualizar las fechas de visitas de un hogar comunitario.
    Recibe: ultima_visita, proxima_visita, observaciones
    Devuelve: JSON con success/error
    
    Lógica de negocio:
    - Si es la primera visita registrada (ultima_visita != null y estado_aptitud == 'no_apto'):
      -> Cambia estado_aptitud a 'apto'
    - Si proxima_visita no se proporciona, la calcula automáticamente:
      -> proxima_visita = ultima_visita + 365 días
      -> Ajusta a día laboral (lunes-viernes, no festivo)
    - Actualiza observaciones_visita
    """
    try:
        # Verificar permisos
        if request.user.rol.nombre not in ['administrador', 'madre_comunitaria']:
            return JsonResponse({
                'success': False,
                'error': 'No tienes permisos para actualizar visitas'
            }, status=403)
        
        # Obtener el hogar
        hogar = get_object_or_404(HogarComunitario, id=hogar_id)
        
        # Verificar que el usuario tiene acceso a este hogar
        if request.user.rol.nombre == 'madre_comunitaria':
            if hogar.madre.usuario != request.user:
                return JsonResponse({
                    'success': False,
                    'error': 'No tienes acceso a este hogar'
                }, status=403)
        
        # Parsear JSON del body
        data = json.loads(request.body)
        
        # Obtener datos del request
        ultima_visita_str = data.get('ultima_visita')
        proxima_visita_str = data.get('proxima_visita')
        observaciones = data.get('observaciones', '')
        
        # Validar que al menos ultima_visita esté presente
        if not ultima_visita_str:
            return JsonResponse({
                'success': False,
                'error': 'Debe proporcionar la fecha de la última visita'
            }, status=400)
        
        # Convertir fechas
        try:
            ultima_visita = datetime.strptime(ultima_visita_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Formato de fecha inválido para última visita (use YYYY-MM-DD)'
            }, status=400)
        
        # Validar que la fecha no sea futura
        if ultima_visita > date.today():
            return JsonResponse({
                'success': False,
                'error': 'La fecha de la última visita no puede ser futura'
            }, status=400)
        
        # Detectar si es la primera visita
        es_primera_visita = (hogar.ultima_visita is None and hogar.estado_aptitud == 'no_apto')
        
        # Actualizar ultima_visita
        hogar.ultima_visita = ultima_visita
        
        # Si es la primera visita, cambiar estado a "Apto"
        if es_primera_visita:
            hogar.estado_aptitud = 'apto'
        
        # Calcular próxima visita si no se proporcionó
        if proxima_visita_str:
            try:
                proxima_visita = datetime.strptime(proxima_visita_str, '%Y-%m-%d').date()
                
                # Validar que la próxima visita no sea en el pasado
                if proxima_visita < date.today():
                    return JsonResponse({
                        'success': False,
                        'error': 'La próxima visita no puede ser programada en el pasado'
                    }, status=400)
                
                # Validar que no sea más de 1 mes en el futuro
                un_mes_despues = date.today() + timedelta(days=30)
                if proxima_visita > un_mes_despues:
                    return JsonResponse({
                        'success': False,
                        'error': 'La próxima visita no puede ser programada con más de 1 mes de anticipación'
                    }, status=400)
                
                hogar.proxima_visita = proxima_visita
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Formato de fecha inválido para próxima visita (use YYYY-MM-DD)'
                }, status=400)
        else:
            # Auto-calcular próxima visita (1 año después, en día laboral)
            hogar.proxima_visita = calcular_proxima_visita(ultima_visita)
        
        # Actualizar observaciones
        hogar.observaciones_visita = observaciones
        
        # Guardar cambios
        hogar.save()
        
        # Preparar respuesta
        mensaje = '✅ Visita registrada exitosamente'
        if es_primera_visita:
            mensaje += '. El hogar ahora tiene estado "Apto"'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'hogar': {
                'id': hogar.id,
                'nombre': hogar.nombre_hogar,
                'ultima_visita': hogar.ultima_visita.strftime('%Y-%m-%d') if hogar.ultima_visita else None,
                'proxima_visita': hogar.proxima_visita.strftime('%Y-%m-%d') if hogar.proxima_visita else None,
                'estado_aptitud': hogar.estado_aptitud,
                'observaciones': hogar.observaciones_visita
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar datos JSON'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error inesperado: {str(e)}'
        }, status=500)


def calcular_proxima_visita(fecha_base):
    """
    Calcula la próxima fecha de visita técnica:
    - Suma 365 días a la fecha base
    - Ajusta a día laboral (lunes-viernes)
    - Evita festivos nacionales (Colombia)
    
    Args:
        fecha_base: datetime.date - Fecha de la última visita
    
    Returns:
        datetime.date - Fecha calculada para próxima visita
    """
    # Sumar 1 año (365 días)
    proxima = fecha_base + timedelta(days=365)
    
    # Festivos fijos de Colombia (simplificado - en producción usar librería)
    festivos_colombia_2024 = [
        date(2024, 1, 1),   # Año Nuevo
        date(2024, 1, 8),   # Reyes Magos (movido)
        date(2024, 3, 25),  # San José (movido)
        date(2024, 3, 28),  # Jueves Santo
        date(2024, 3, 29),  # Viernes Santo
        date(2024, 5, 1),   # Día del Trabajo
        date(2024, 5, 13),  # Ascensión (movido)
        date(2024, 6, 3),   # Corpus Christi (movido)
        date(2024, 6, 10),  # Sagrado Corazón (movido)
        date(2024, 7, 1),   # San Pedro y San Pablo (movido)
        date(2024, 7, 20),  # Independencia
        date(2024, 8, 7),   # Batalla de Boyacá
        date(2024, 8, 19),  # Asunción (movido)
        date(2024, 10, 14), # Día de la Raza (movido)
        date(2024, 11, 4),  # Todos los Santos (movido)
        date(2024, 11, 11), # Independencia de Cartagena (movido)
        date(2024, 12, 8),  # Inmaculada Concepción
        date(2024, 12, 25), # Navidad
    ]
    
    # Ajustar a día laboral (evitar sábados, domingos y festivos)
    while proxima.weekday() >= 5 or proxima in festivos_colombia_2024:
        proxima += timedelta(days=1)
        
        # Si avanzamos al siguiente año, actualizar lista de festivos
        # (En producción, usar librería como workalendar o API de festivos)
        if proxima.year > 2024:
            # Por simplicidad, solo verificar fines de semana en años futuros
            # TODO: Implementar librería de festivos para todos los años
            while proxima.weekday() >= 5:
                proxima += timedelta(days=1)
            break
    
    return proxima


@login_required
@rol_requerido('administrador')
def activar_hogar(request, hogar_id):
    """
    Vista para activar un hogar después de la primera visita técnica.
    Solo disponible el día programado de la visita.
    
    Procesa el formulario de evaluación y determina:
    - Si el hogar es APTO (aprobado) → estado 'activo'
    - Si el hogar es NO APTO → estado 'pendiente_visita'
    - Envía email de notificación si es activado
    """
    hogar = get_object_or_404(HogarComunitario, id=hogar_id)
    fecha_hoy = date.today()
    
    # Permitir activación de hogares que NO estén ya activos o aprobados
    # Incluye: pendiente_visita, rechazado, en_revision, etc.
    if hogar.estado in ['activo', 'aprobado']:
        messages.warning(request, 'Este hogar ya está activo y aprobado.')
        return redirect('hogares_dashboard')
    
    # Manejar diferentes escenarios de fecha
    if not hogar.fecha_primera_visita:
        # Hogar SIN fecha programada - Activación directa
        messages.info(request, 
            f'📋 Este hogar NO tenía visita programada. '
            f'Se registrará la fecha de hoy ({fecha_hoy.strftime("%d de %B de %Y")}) como fecha de visita.'
        )
    elif fecha_hoy < hogar.fecha_primera_visita:
        # Activación ANTICIPADA
        messages.info(request, 
            f'📅 Visita programada para: {hogar.fecha_primera_visita.strftime("%d de %B de %Y")}. '
            f'Estás realizando la evaluación de forma anticipada.'
        )
    elif fecha_hoy > hogar.fecha_primera_visita:
        # Activación CON RETRASO
        messages.warning(request, 
            f'⚠️ La fecha programada era: {hogar.fecha_primera_visita.strftime("%d de %B de %Y")}. '
            f'Estás realizando la evaluación con retraso.'
        )
    # Si fecha_hoy == fecha_primera_visita, no mostrar mensaje (es el día correcto)
    
    if request.method == 'POST':
        try:
            # Capturar datos del formulario
            tipo_vivienda = request.POST.get('tipo_vivienda')
            ubicacion = request.POST.get('ubicacion')
            sin_riesgo = request.POST.get('sin_riesgo')
            
            # Servicios públicos (checkboxes)
            servicios = {
                'acueducto': request.POST.get('servicio_acueducto') == '1',
                'alcantarillado': request.POST.get('servicio_alcantarillado') == '1',
                'energia': request.POST.get('servicio_energia') == '1',
                'gas': request.POST.get('servicio_gas') == '1',
                'internet': request.POST.get('servicio_internet') == '1',
                'telefono': request.POST.get('servicio_telefono') == '1',
            }
            
            # Espacios
            espacios = {
                'sala': request.POST.get('espacio_sala') == '1',
                'comedor': request.POST.get('espacio_comedor') == '1',
                'cocina': request.POST.get('espacio_cocina') == '1',
                'patio': request.POST.get('espacio_patio') == '1',
            }
            espacio_suficiente = request.POST.get('espacio_suficiente')
            
            # Condiciones
            higiene = request.POST.get('higiene')
            orden = request.POST.get('orden')
            estado_vivienda = request.POST.get('estado_vivienda')
            ventilacion = request.POST.get('ventilacion')
            iluminacion = request.POST.get('iluminacion')
            
            # Aspectos familiares
            familia_acuerdo = request.POST.get('familia_acuerdo')
            dinamica_familiar = request.POST.get('dinamica_familiar', '')
            
            # Observaciones y recomendación
            observaciones_generales = request.POST.get('observaciones_generales', '')
            capacidad_calculada = int(request.POST.get('capacidad_calculada', 0))
            recomendacion = request.POST.get('recomendacion')
            
            # ✅ Validar capacidad permitida (12-15 niños según normativa ICBF)
            if capacidad_calculada < 12 or capacidad_calculada > 15:
                messages.error(request, 
                    f'❌ Error: La capacidad debe estar entre 12 y 15 niños. '
                    f'Valor ingresado: {capacidad_calculada}. '
                    f'Verifica el formulario y selecciona una capacidad válida.'
                )
                return redirect('activar_hogar', hogar_id=hogar.id)
            
            # Construir observaciones completas
            observaciones_completas = f"""
=== EVALUACIÓN DE PRIMERA VISITA TÉCNICA ===
Fecha: {fecha_hoy.strftime('%d/%m/%Y')}

VIVIENDA:
- Tipo: {tipo_vivienda}
- Ubicación: {ubicacion}
- Sin zonas de riesgo: {sin_riesgo}
- Estado: {estado_vivienda}

SERVICIOS PÚBLICOS:
- Acueducto: {'✓' if servicios['acueducto'] else '✗'}
- Alcantarillado: {'✓' if servicios['alcantarillado'] else '✗'}
- Energía: {'✓' if servicios['energia'] else '✗'}
- Gas: {'✓' if servicios['gas'] else '✗'}
- Internet: {'✓' if servicios['internet'] else '✗'}
- Teléfono: {'✓' if servicios['telefono'] else '✗'}

ESPACIOS:
- Sala: {'✓' if espacios['sala'] else '✗'}
- Comedor: {'✓' if espacios['comedor'] else '✗'}
- Cocina: {'✓' if espacios['cocina'] else '✗'}
- Patio: {'✓' if espacios['patio'] else '✗'}
- Espacio suficiente: {espacio_suficiente}

CONDICIONES:
- Higiene: {higiene}
- Orden: {orden}
- Ventilación: {ventilacion}
- Iluminación: {iluminacion}

FAMILIA:
- Acuerdo familiar: {familia_acuerdo}
- Dinámica: {dinamica_familiar}

CAPACIDAD CALCULADA: {capacidad_calculada} niños

OBSERVACIONES:
{observaciones_generales}

RECOMENDACIÓN: {recomendacion.upper()}
"""
            
            # Actualizar campos del hogar
            hogar.ultima_visita = fecha_hoy
            hogar.observaciones_visita = observaciones_completas
            hogar.capacidad = capacidad_calculada
            hogar.capacidad_calculada = capacidad_calculada  # Sincronizar ambos campos
            
            # Si no tenía fecha_primera_visita, asignar la de hoy
            if not hogar.fecha_primera_visita:
                hogar.fecha_primera_visita = fecha_hoy
            
            # Determinar estado según recomendación
            if recomendacion == 'aprobado':
                hogar.estado_aptitud = 'apto'
                hogar.estado = 'activo'
                hogar.proxima_visita = calcular_proxima_visita(fecha_hoy)
                
                # Guardar cambios
                hogar.save()
                
                # Enviar email de activación
                enviar_email_activacion(hogar)
                
                messages.success(request, 
                    f'✅ ¡Hogar ACTIVADO exitosamente! '
                    f'El hogar "{hogar.nombre_hogar}" ahora está en estado ACTIVO. '
                    f'Se ha enviado un correo de notificación al agente educativo.'
                )
                
            elif recomendacion == 'aprobado_condiciones':
                hogar.estado_aptitud = 'apto'
                hogar.estado = 'activo'
                hogar.proxima_visita = calcular_proxima_visita(fecha_hoy)
                hogar.save()
                
                messages.warning(request, 
                    f'⚠️ Hogar activado CON CONDICIONES. '
                    f'Revisar observaciones: {observaciones_generales[:100]}...'
                )
                
            elif recomendacion == 'no_aprobado':
                hogar.estado_aptitud = 'no_apto'
                hogar.estado = 'pendiente_visita'
                hogar.save()
                
                messages.error(request, 
                    f'❌ Hogar NO APROBADO. '
                    f'El hogar permanece en estado "Pendiente de Visita". '
                    f'Se requiere corregir las deficiencias antes de programar nueva visita.'
                )
                
            elif recomendacion == 'requiere_nueva_visita':
                hogar.estado_aptitud = 'no_apto'
                hogar.estado = 'pendiente_visita'
                hogar.fecha_primera_visita = None  # Permitir reprogramación
                hogar.save()
                
                messages.info(request, 
                    f'🔄 Se requiere NUEVA VISITA. '
                    f'Puede programar una nueva fecha desde el dashboard.'
                )
            
            return redirect('hogares_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el formulario: {str(e)}')
            return redirect('activar_hogar', hogar_id=hogar.id)
    
    # GET: Mostrar formulario
    context = {
        'hogar': hogar,
        'fecha_hoy': fecha_hoy,
        'es_administrador': request.user.rol.nombre_rol == 'administrador',
    }
    
    return render(request, 'admin/formulario_activacion_hogar.html', context)


@login_required
@rol_requerido('administrador')
def registrar_visita(request, hogar_id):
    """
    Vista para registrar visitas de seguimiento en hogares YA ACTIVOS.
    Usa el mismo formulario de evaluación que la activación.
    """
    hogar = get_object_or_404(HogarComunitario, id=hogar_id)
    fecha_hoy = date.today()
    
    # Verificar que el hogar esté activo
    if hogar.estado not in ['activo', 'aprobado']:
        messages.warning(request, 'Solo puedes registrar visitas en hogares activos.')
        return redirect('hogares_dashboard')
    
    if request.method == 'POST':
        try:
            # Capturar datos del formulario (mismo código que activar_hogar)
            tipo_vivienda = request.POST.get('tipo_vivienda')
            ubicacion = request.POST.get('ubicacion')
            sin_riesgo = request.POST.get('sin_riesgo')
            
            servicios = {
                'acueducto': request.POST.get('servicio_acueducto') == '1',
                'alcantarillado': request.POST.get('servicio_alcantarillado') == '1',
                'energia': request.POST.get('servicio_energia') == '1',
                'gas': request.POST.get('servicio_gas') == '1',
                'internet': request.POST.get('servicio_internet') == '1',
                'telefono': request.POST.get('servicio_telefono') == '1',
            }
            
            espacios = {
                'sala': request.POST.get('espacio_sala') == '1',
                'comedor': request.POST.get('espacio_comedor') == '1',
                'cocina': request.POST.get('espacio_cocina') == '1',
                'patio': request.POST.get('espacio_patio') == '1',
            }
            espacio_suficiente = request.POST.get('espacio_suficiente')
            
            higiene = request.POST.get('higiene')
            orden = request.POST.get('orden')
            estado_vivienda = request.POST.get('estado_vivienda')
            ventilacion = request.POST.get('ventilacion')
            iluminacion = request.POST.get('iluminacion')
            
            familia_acuerdo = request.POST.get('familia_acuerdo')
            dinamica_familiar = request.POST.get('dinamica_familiar', '')
            
            observaciones_generales = request.POST.get('observaciones_generales', '')
            capacidad_calculada = int(request.POST.get('capacidad_calculada', hogar.capacidad))
            recomendacion = request.POST.get('recomendacion')
            
            # Validar capacidad
            if capacidad_calculada < 12 or capacidad_calculada > 15:
                messages.error(request, 
                    f'❌ Error: La capacidad debe estar entre 12 y 15 niños. '
                    f'Valor ingresado: {capacidad_calculada}.'
                )
                return redirect('registrar_visita', hogar_id=hogar.id)
            
            # Construir observaciones
            observaciones_completas = f"""
=== VISITA DE SEGUIMIENTO ===
Fecha: {fecha_hoy.strftime('%d/%m/%Y')}

VIVIENDA:
- Tipo: {tipo_vivienda}
- Ubicación: {ubicacion}
- Sin zonas de riesgo: {sin_riesgo}
- Estado: {estado_vivienda}

SERVICIOS PÚBLICOS:
- Acueducto: {'✓' if servicios['acueducto'] else '✗'}
- Alcantarillado: {'✓' if servicios['alcantarillado'] else '✗'}
- Energía: {'✓' if servicios['energia'] else '✗'}
- Gas: {'✓' if servicios['gas'] else '✗'}
- Internet: {'✓' if servicios['internet'] else '✗'}
- Teléfono: {'✓' if servicios['telefono'] else '✗'}

ESPACIOS:
- Sala: {'✓' if espacios['sala'] else '✗'}
- Comedor: {'✓' if espacios['comedor'] else '✗'}
- Cocina: {'✓' if espacios['cocina'] else '✗'}
- Patio: {'✓' if espacios['patio'] else '✗'}
- Espacio suficiente: {espacio_suficiente}

CONDICIONES:
- Higiene: {higiene}
- Orden: {orden}
- Ventilación: {ventilacion}
- Iluminación: {iluminacion}

FAMILIA:
- Acuerdo familiar: {familia_acuerdo}
- Dinámica: {dinamica_familiar}

CAPACIDAD: {capacidad_calculada} niños

OBSERVACIONES:
{observaciones_generales}

RESULTADO: {recomendacion.upper()}
"""
            
            # Actualizar hogar
            hogar.ultima_visita = fecha_hoy
            hogar.observaciones_visita = observaciones_completas
            hogar.capacidad = capacidad_calculada
            hogar.capacidad_calculada = capacidad_calculada  # Sincronizar ambos campos
            
            # Determinar acción según recomendación
            if recomendacion == 'aprobado':
                hogar.estado_aptitud = 'apto'
                hogar.estado = 'activo'
                hogar.proxima_visita = calcular_proxima_visita(fecha_hoy)
                hogar.save()
                
                messages.success(request, 
                    f'✅ Visita registrada exitosamente. '
                    f'El hogar continúa ACTIVO. '
                    f'Próxima visita: {hogar.proxima_visita.strftime("%d/%m/%Y")}'
                )
                
            elif recomendacion == 'aprobado_condiciones':
                hogar.estado_aptitud = 'apto'
                hogar.estado = 'activo'
                hogar.proxima_visita = calcular_proxima_visita(fecha_hoy)
                hogar.save()
                
                messages.warning(request, 
                    f'⚠️ Visita registrada CON CONDICIONES. '
                    f'Revisar observaciones: {observaciones_generales[:100]}...'
                )
                
            elif recomendacion == 'no_aprobado':
                hogar.estado_aptitud = 'no_apto'
                hogar.estado = 'rechazado'
                hogar.save()
                
                messages.error(request, 
                    f'❌ Hogar NO APTO. '
                    f'El estado ha cambiado a "Rechazado". '
                    f'Se requiere corrección antes de próxima visita.'
                )
                
            elif recomendacion == 'requiere_nueva_visita':
                hogar.estado_aptitud = 'apto'
                hogar.estado = 'activo'
                # Programar visita más próxima (30 días)
                hogar.proxima_visita = fecha_hoy + timedelta(days=30)
                hogar.save()
                
                messages.info(request, 
                    f'🔄 Se requiere NUEVA VISITA en 30 días. '
                    f'Próxima visita: {hogar.proxima_visita.strftime("%d/%m/%Y")}'
                )
            
            return redirect('hogares_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el formulario: {str(e)}')
            return redirect('registrar_visita', hogar_id=hogar.id)
    
    # GET: Mostrar formulario
    context = {
        'hogar': hogar,
        'fecha_hoy': fecha_hoy,
        'es_administrador': request.user.rol.nombre_rol == 'administrador',
        'es_seguimiento': True,  # 🆕 Flag para indicar que es visita de seguimiento
    }
    
    return render(request, 'admin/formulario_activacion_hogar.html', context)


def enviar_email_activacion(hogar):
    """
    Envía email de notificación cuando un hogar es activado.
    """
    try:
        madre = hogar.madre
        usuario = madre.usuario
        
        asunto = f'✅ Hogar Activado - ICBF Conecta'
        
        mensaje = f"""
¡Hola {usuario.nombres}!

Nos complace informarte que tu Hogar Comunitario ha sido ACTIVADO exitosamente.

📋 DETALLES DEL HOGAR:
- Nombre: {hogar.nombre_hogar}
- Dirección: {hogar.direccion}
- Estado: ACTIVO
- Capacidad aprobada: {hogar.capacidad} niños
- Fecha de activación: {date.today().strftime('%d de %B de %Y')}

🔐 ACCESO AL SISTEMA:
Ahora puedes acceder al sistema ICBF Conecta con tus credenciales:

- Usuario: {usuario.numero_documento}
- Contraseña temporal: 123456

Por favor, cambia tu contraseña en tu primer inicio de sesión.

📅 PRÓXIMA VISITA TÉCNICA:
Tu próxima visita está programada para: {hogar.proxima_visita.strftime('%d de %B de %Y')}

Si tienes alguna pregunta, no dudes en contactarnos.

¡Felicidades y bienvenida al programa ICBF Conecta!

---
Sistema ICBF Conecta
Este es un correo automático, por favor no responder.
"""
        
        send_mail(
            asunto,
            mensaje,
            settings.EMAIL_HOST_USER,
            [usuario.correo],
            fail_silently=False,
        )
        
    except Exception as e:
        print(f"Error al enviar email de activación: {str(e)}")

